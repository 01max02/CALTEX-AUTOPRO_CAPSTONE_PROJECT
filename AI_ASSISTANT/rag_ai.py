# ============================================================================
# RAG AI PROTOTYPE  Combined Single-File Version
# ============================================================================
# This file combines all .py modules from the RAG_AI_PROTOTYPE-main project
# into a single file for reference/documentation purposes.
#
# Original files: 23 Python modules
# Generated: 2026-05-25 19:12
#
# NOTE: This combined file is for REFERENCE ONLY. To run the system,
# use the original modular files with: python main.py
# ============================================================================


# ============================================================================
# FILE: conversation_store.py
# ============================================================================

"""In-memory conversational history store."""

from __future__ import annotations

from collections import defaultdict, deque
from dataclasses import dataclass
from datetime import datetime, timezone
from threading import Lock


@dataclass
class ChatTurn:
    role: str
    message: str
    timestamp: str


class ConversationMemory:
    def __init__(self, max_turns: int = 14) -> None:
        self.max_turns = max_turns
        self._sessions: dict[str, deque[ChatTurn]] = defaultdict(
            lambda: deque(maxlen=self.max_turns)
        )
        self._lock = Lock()

    def add_turn(self, session_id: str, role: str, message: str) -> None:
        with self._lock:
            self._sessions[session_id].append(
                ChatTurn(
                    role=role,
                    message=message,
                    timestamp=datetime.now(timezone.utc).isoformat(),
                )
            )

    def get_recent(self, session_id: str, limit: int = 8) -> list[dict]:
        with self._lock:
            turns = list(self._sessions.get(session_id, deque()))
        return [turn.__dict__ for turn in turns[-limit:]]

    def clear(self, session_id: str) -> None:
        with self._lock:
            self._sessions.pop(session_id, None)

# ============================================================================
# FILE: data_normalizer.py
# ============================================================================

"""Data normalization layer for NoSQL (Firestore) records.

Firestore documents have inconsistent field names across collections and
even within the same collection. This module normalizes raw payloads into
a consistent schema so the retrieval pipeline, template responses, and
LLM prompts always see predictable field names.

Normalization happens AFTER fetching from Firestore and BEFORE chunking,
embedding, and retrieval.

Each domain has a canonical schema with standard field names. The normalizer
maps variant field names to canonical ones.
"""

from __future__ import annotations

from typing import Any


# ---------------------------------------------------------------------------
# Canonical field schemas per domain
# ---------------------------------------------------------------------------

# Maps: canonical_field_name -> list of possible source field names (aliases)
# The normalizer checks aliases in order and uses the first match.

DOMAIN_SCHEMAS: dict[str, list[tuple[str, list[str]]]] = {
    "item_master": [
        ("name", ["name", "item", "item_name", "product", "product_name", "material", "material_name", "description", "title", "label", "display_name", "service", "service_name"]),
        ("category", ["category", "group", "type", "material_type", "class", "kind"]),
        ("quantity", ["qty", "quantity", "stock", "remaining_stock", "current_stock", "count", "available", "on_hand"]),
        ("uom", ["uom", "unit", "unit_of_measure", "measure", "unit_measure"]),
        ("price", ["price", "cost", "unit_cost", "unit_price", "amount", "rate"]),
        ("status", ["status", "state", "condition", "availability"]),
    ],
    "inventory": [
        ("name", ["name", "item", "item_name", "product", "product_name", "material", "description", "title", "label"]),
        ("category", ["category", "group", "type", "material_type", "class"]),
        ("quantity", ["qty", "quantity", "stock", "remaining_stock", "current_stock", "count", "on_hand", "available"]),
        ("uom", ["uom", "unit", "unit_of_measure"]),
        ("price", ["price", "cost", "unit_cost", "unit_price"]),
        ("reorder_level", ["reorder_level", "reorder_point", "reorder", "min_stock", "minimum"]),
        ("supplier", ["supplier", "supplier_name", "vendor", "vendor_name", "supplier_info"]),
        ("status", ["status", "state", "condition"]),
    ],
    "stock_inventory": [
        ("name", ["name", "item", "item_name", "product", "product_name", "material", "description", "title"]),
        ("category", ["category", "group", "type", "material_type"]),
        ("quantity", ["qty", "quantity", "stock", "remaining_stock", "current_stock", "count", "on_hand", "available"]),
        ("uom", ["uom", "unit", "unit_of_measure"]),
        ("reorder_level", ["reorder_level", "reorder_point", "reorder", "min_stock"]),
        ("status", ["status", "state", "condition"]),
    ],
    "products": [
        ("name", ["name", "product", "product_name", "item", "item_name", "title", "description", "label"]),
        ("category", ["category", "group", "type", "class"]),
        ("price", ["price", "cost", "unit_cost", "unit_price", "amount"]),
        ("quantity", ["qty", "quantity", "stock", "count"]),
        ("sku", ["sku", "product_code", "item_code", "code"]),
        ("status", ["status", "state", "availability"]),
    ],
    "services": [
        ("service_name", ["service_name", "service", "name", "service_type", "type", "description", "title", "repair_type", "work_type", "job_type"]),
        ("vehicle", ["vehicle", "vehicle_name", "plate", "plate_number", "plate_no", "car", "unit", "vehicle_plate", "license_plate"]),
        ("customer", ["customer", "customer_name", "client", "client_name", "owner", "owner_name", "requested_by"]),
        ("date", ["date", "service_date", "date_completed", "completed_date", "date_done", "schedule_date", "appointment_date", "created_at", "timestamp", "date_created"]),
        ("status", ["status", "state", "condition", "progress", "job_status", "repair_status"]),
        ("technician", ["technician", "mechanic", "assigned_to", "performed_by", "staff", "worker"]),
        ("cost", ["cost", "price", "amount", "total", "total_cost", "service_cost", "charge"]),
        ("notes", ["notes", "remarks", "comment", "comments", "description", "details", "findings"]),
    ],
    "orders": [
        ("order_id", ["order_id", "id", "order_number", "reference", "ref", "invoice", "ticket"]),
        ("customer", ["customer", "customer_name", "client", "client_name", "ordered_by", "requested_by"]),
        ("item", ["item", "product", "name", "description", "material", "service"]),
        ("quantity", ["quantity", "qty", "count", "amount"]),
        ("status", ["status", "state", "order_status", "progress"]),
        ("date", ["date", "order_date", "created_at", "timestamp", "date_created", "placed_date"]),
        ("total", ["total", "total_cost", "amount", "price", "cost"]),
    ],
    "customers": [
        ("name", ["name", "customer_name", "full_name", "client_name", "client", "owner"]),
        ("contact", ["contact", "phone", "mobile", "phone_number", "contact_number", "cell"]),
        ("email", ["email", "email_address", "e_mail"]),
        ("vehicle", ["vehicle", "plate", "plate_number", "car", "vehicle_plate"]),
        ("address", ["address", "location", "city", "area"]),
    ],
    "issuance": [
        ("item", ["item", "material", "product", "name", "description", "issued_item"]),
        ("quantity", ["quantity", "qty", "count", "amount", "issued_qty"]),
        ("issued_to", ["issued_to", "recipient", "received_by", "customer", "department", "requestor"]),
        ("date", ["date", "issue_date", "date_issued", "created_at", "timestamp"]),
        ("reference", ["reference", "ref", "reference_number", "ticket", "order_id"]),
        ("status", ["status", "state"]),
    ],
    "manual": [
        ("title", ["title", "name", "topic", "subject", "heading", "section_title"]),
        ("content", ["content", "body", "text", "description", "details", "instructions", "steps", "guide"]),
        ("category", ["category", "section", "group", "type", "module", "area", "feature"]),
        ("platform", ["platform", "app", "device", "target", "for"]),
    ],
}

# Fields that should NEVER be used as a display name (they are metadata, not identifiers)
_NON_NAME_FIELDS: set[str] = {
    "status", "state", "condition", "progress", "availability",
    "true", "false", "yes", "no", "completed", "pending", "active",
    "inactive", "done", "cancelled", "approved", "rejected",
}


# ---------------------------------------------------------------------------
# Normalizer
# ---------------------------------------------------------------------------


class DataNormalizer:
    """Normalize Firestore document payloads into consistent schemas.

    Usage:
        normalizer = DataNormalizer()
        normalized = normalizer.normalize(payload, domain="services")
        # normalized now has consistent field names like "service_name", "vehicle", "date", "status"
    """

    def __init__(self, schemas: dict[str, list[tuple[str, list[str]]]] | None = None) -> None:
        self.schemas = schemas or DOMAIN_SCHEMAS

    def normalize(self, payload: dict[str, Any], domain: str) -> dict[str, Any]:
        """Normalize a raw Firestore payload into canonical field names.

        Args:
            payload: Raw document payload from Firestore
            domain: The collection/domain name (e.g., "services", "inventory")

        Returns:
            Normalized payload with canonical field names.
            Original fields that don't map to any canonical name are preserved as-is.
        """
        if not payload:
            return {}

        schema = self.schemas.get(domain.lower())
        if not schema:
            # No schema defined for this domain — return as-is with basic cleanup
            return self._basic_cleanup(payload)

        # Build case-insensitive lookup of the source payload
        payload_lower_map: dict[str, tuple[str, Any]] = {}
        for key, value in payload.items():
            payload_lower_map[key.lower().strip()] = (key, value)

        normalized: dict[str, Any] = {}
        used_source_keys: set[str] = set()

        # Map canonical fields from aliases
        for canonical_name, aliases in schema:
            for alias in aliases:
                alias_lower = alias.lower().strip()
                if alias_lower in payload_lower_map:
                    original_key, value = payload_lower_map[alias_lower]
                    # Only use non-empty values
                    if self._is_meaningful(value):
                        normalized[canonical_name] = self._clean_value(value)
                        used_source_keys.add(alias_lower)
                        break

        # Preserve any remaining fields that weren't mapped
        for key_lower, (original_key, value) in payload_lower_map.items():
            if key_lower not in used_source_keys and self._is_meaningful(value):
                normalized[original_key] = self._clean_value(value)

        return normalized

    def normalize_records(self, records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Normalize a batch of raw Firestore records.

        Each record is expected to have: domain, record_id, source, payload
        """
        normalized: list[dict[str, Any]] = []
        for record in records:
            domain = str(record.get("domain", "")).lower()
            payload = record.get("payload", {}) or {}
            normalized_payload = self.normalize(payload, domain)
            normalized.append({
                **record,
                "payload": normalized_payload,
            })
        return normalized

    def get_display_name(self, payload: dict[str, Any], domain: str) -> str:
        """Extract the best display name from a normalized payload.

        This is the primary method for getting a human-readable identifier
        from any record, regardless of the original Firestore field names.
        """
        # Domain-specific primary name fields
        domain_name_fields: dict[str, list[str]] = {
            "item_master": ["name"],
            "inventory": ["name"],
            "stock_inventory": ["name"],
            "products": ["name"],
            "services": ["service_name", "vehicle", "customer"],
            "orders": ["order_id", "item", "customer"],
            "customers": ["name"],
            "issuance": ["item", "issued_to"],
            "manual": ["title"],
        }

        # Try domain-specific fields first
        name_fields = domain_name_fields.get(domain.lower(), [])
        for field in name_fields:
            value = payload.get(field)
            if value and self._is_name_worthy(value):
                return str(value).strip()

        # Generic fallback: try common name fields
        generic_name_fields = ["name", "title", "item", "service_name", "product", "customer", "vehicle", "order_id"]
        for field in generic_name_fields:
            value = payload.get(field)
            if value and self._is_name_worthy(value):
                return str(value).strip()

        # Last resort: first meaningful string value that isn't a status/boolean
        for key, value in payload.items():
            if isinstance(value, str) and self._is_name_worthy(value):
                return value.strip()[:60]

        return f"{domain.replace('_', ' ').title()} Record"

    def _is_meaningful(self, value: Any) -> bool:
        """Check if a value is non-empty and meaningful."""
        if value is None:
            return False
        if isinstance(value, str):
            stripped = value.strip().lower()
            return stripped not in ("", "none", "null", "n/a", "na", "unknown", "-", "undefined")
        if isinstance(value, (list, dict)):
            return len(value) > 0
        return True

    def _is_name_worthy(self, value: Any) -> bool:
        """Check if a value is suitable as a display name (not a status/boolean)."""
        if not isinstance(value, str):
            return False
        stripped = value.strip().lower()
        if stripped in _NON_NAME_FIELDS:
            return False
        if len(stripped) < 2:
            return False
        if stripped in ("none", "null", "n/a", "na", "unknown", "-", "undefined", "true", "false"):
            return False
        return True

    def _clean_value(self, value: Any) -> Any:
        """Clean a value for consistent representation."""
        if isinstance(value, str):
            stripped = value.strip()
            return stripped if stripped else None
        return value

    def _basic_cleanup(self, payload: dict[str, Any]) -> dict[str, Any]:
        """Basic cleanup for domains without a defined schema."""
        cleaned: dict[str, Any] = {}
        for key, value in payload.items():
            if self._is_meaningful(value):
                cleaned[key] = self._clean_value(value)
        return cleaned


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_normalizer: DataNormalizer | None = None


def get_data_normalizer() -> DataNormalizer:
    """Return the module-level DataNormalizer singleton."""
    global _default_normalizer
    if _default_normalizer is None:
        _default_normalizer = DataNormalizer()
    return _default_normalizer


# ============================================================================
# FILE: data_validator.py
# ============================================================================

"""Data validation and safe formatting layer."""

from __future__ import annotations

from dataclasses import dataclass
from typing import Any


@dataclass
class ValidationResult:
    """Result of data validation."""

    is_valid: bool
    is_complete: bool
    confidence: float
    missing_fields: list[str]
    issues: list[str]


class DataValidator:
    """Validate and assess data completeness."""

    DOMAIN_REQUIRED_FIELDS: dict[str, set[str]] = {
        "inventory": {"name", "type", "qty", "uom"},
        "products": {"name", "price"},
        "vehicles": {"plate_no", "vehicle_type"},
        "maintenance": {"vehicle", "service", "status"},
        "schedules": {"schedule_date", "vehicle", "status"},
        "customers": {"name", "contact"},
        "services": {"service_name", "cost"},
        "issuance": {"item", "quantity", "date"},
    }

    SENSITIVE_FIELDS: set[str] = {
        "id", "_id", "record_id", "internal_id",
        "supplier_id", "supplier_contact", "supplier_info",
        "internal_cost", "cost_price", "margin", "markup",
        "reorder_level", "reorder_point", "min_stock",
        "owner_id", "user_id", "mechanic_id", "admin_notes",
        "other_customer", "customer_list", "full_inventory",
        "financial", "analytics", "database"
    }

    # Fields that should be filtered for regular customers
    CUSTOMER_RESTRICTED_FIELDS: set[str] = {
        "sku", "product_code", "supplier", "internal_code",
        "reorder_level", "reorder_point", "cost", "unit_cost",
        "margin", "markup", "wholesale_price"
    }

    def validate_record(
        self,
        payload: dict[str, Any],
        domain: str,
        is_admin: bool = False,
    ) -> ValidationResult:
        """Validate a single record for completeness and security."""
        issues: list[str] = []
        missing_fields: list[str] = []
        
        # Check for sensitive fields
        exposed_sensitive = [
            field for field in self.SENSITIVE_FIELDS
            if field in payload
        ]
        if exposed_sensitive:
            issues.append(f"Contains sensitive fields: {', '.join(exposed_sensitive)}")
        
        # Check required fields
        required = self.DOMAIN_REQUIRED_FIELDS.get(domain, set())
        missing = required - set(k.lower() for k in payload.keys())
        if missing:
            missing_fields = list(missing)
            issues.append(f"Missing required fields: {', '.join(missing)}")
        
        # Check for empty values
        empty_fields = [
            k for k, v in payload.items()
            if v is None or v == "" or v == []
        ]
        if empty_fields:
            issues.append(f"Has empty fields: {', '.join(empty_fields)}")
        
        is_complete = len(missing_fields) == 0 and len(empty_fields) == 0
        is_valid = len(exposed_sensitive) == 0
        confidence = max(0.0, 1.0 - (len(issues) * 0.25))
        
        return ValidationResult(
            is_valid=is_valid,
            is_complete=is_complete,
            confidence=confidence,
            missing_fields=missing_fields,
            issues=issues,
        )

    def filter_sensitive_fields(
        self,
        payload: dict[str, Any],
        is_admin: bool = False,
    ) -> dict[str, Any]:
        """Remove sensitive fields from payload based on user type."""
        filtered = {}
        
        for key, value in payload.items():
            key_lower = key.lower()
            
            # Never expose absolutely sensitive fields
            if any(sensitive in key_lower for sensitive in self.SENSITIVE_FIELDS):
                continue
            
            # Filter customer-restricted fields if not admin
            if not is_admin and any(
                restricted in key_lower
                for restricted in self.CUSTOMER_RESTRICTED_FIELDS
            ):
                continue
            
            filtered[key] = value
        
        return filtered


class DataFormatter:
    """Format data for safe, clean display."""

    # Map internal field names to user-friendly names
    FIELD_LABELS: dict[str, str] = {
        "name": "Name",
        "item": "Item",
        "product": "Product",
        "qty": "Quantity",
        "quantity": "Quantity",
        "uom": "Unit of Measurement",
        "type": "Type",
        "category": "Category",
        "price": "Price",
        "cost": "Cost",
        "stock": "Stock Level",
        "available": "Available",
        "plate_no": "Plate Number",
        "vehicle": "Vehicle",
        "vehicle_type": "Vehicle Type",
        "service": "Service",
        "service_type": "Service Type",
        "date": "Date",
        "created_at": "Created",
        "updated_at": "Updated",
        "status": "Status",
        "mileage": "Mileage",
        "odometer": "Odometer",
        "schedule_date": "Scheduled Date",
        "customer": "Customer",
        "owner": "Owner",
        "contact": "Contact",
        "email": "Email",
        "phone": "Phone",
    }

    @staticmethod
    def format_field_label(field_name: str) -> str:
        """Convert field name to user-friendly label."""
        if field_name in DataFormatter.FIELD_LABELS:
            return DataFormatter.FIELD_LABELS[field_name]
        
        # Fallback: capitalize and replace underscores
        return field_name.replace("_", " ").title()

    @staticmethod
    def format_value(value: Any, field_type: str = "text") -> str:
        """Format a value for display."""
        if value is None:
            return "N/A"
        
        if isinstance(value, bool):
            return "Yes" if value else "No"
        
        if isinstance(value, (list, tuple)):
            return ", ".join(str(v) for v in value if v)
        
        if isinstance(value, dict):
            pairs = [
                f"{DataFormatter.format_field_label(k)}: {v}"
                for k, v in value.items()
            ]
            return "; ".join(pairs)
        
        text = str(value).strip()
        
        # Clean up common data issues
        if text.lower() in {"none", "null", "n/a", "na", "unknown", "-", ""}:
            return "N/A"
        
        return text

    @staticmethod
    def format_record_summary(
        payload: dict[str, Any],
        domain: str,
        max_fields: int = 5,
    ) -> str:
        """Create a natural language summary of a record."""
        # Domain-specific priority fields
        priority_map: dict[str, list[str]] = {
            "inventory": ["name", "type", "qty", "uom", "stock"],
            "products": ["name", "price", "category"],
            "vehicles": ["vehicle_type", "plate_no", "odometer"],
            "maintenance": ["service", "vehicle", "date", "status"],
            "schedules": ["schedule_date", "vehicle", "service_type", "status"],
            "customers": ["name", "contact", "email", "phone"],
            "services": ["service_name", "category", "cost"],
        }
        
        priority_fields = priority_map.get(domain, [])
        
        # Collect fields in priority order
        parts: list[str] = []
        seen: set[str] = set()
        
        for priority_field in priority_fields:
            for key, value in payload.items():
                if key.lower() in seen:
                    continue
                if key.lower() == priority_field.lower():
                    formatted = DataFormatter.format_value(value)
                    if formatted != "N/A":
                        label = DataFormatter.format_field_label(key)
                        parts.append(f"{label}: {formatted}")
                        seen.add(key.lower())
                        if len(parts) >= max_fields:
                            break
            if len(parts) >= max_fields:
                break
        
        # Add remaining fields if space permits
        if len(parts) < max_fields:
            for key, value in payload.items():
                if key.lower() in seen:
                    continue
                if not key.lower().startswith("_"):
                    formatted = DataFormatter.format_value(value)
                    if formatted != "N/A":
                        label = DataFormatter.format_field_label(key)
                        parts.append(f"{label}: {formatted}")
                        seen.add(key.lower())
                        if len(parts) >= max_fields:
                            break
        
        if not parts:
            return "No available data"
        
        return " | ".join(parts)

    @staticmethod
    def format_no_results_message(
        intent: str,
        searched_domains: list[str],
    ) -> str:
        """Generate helpful message when no results found."""
        domain_names = {
            "inventory": "inventory",
            "products": "products",
            "vehicles": "vehicle records",
            "maintenance": "maintenance records",
            "schedules": "schedules",
            "customers": "customer records",
            "services": "services",
            "issuance": "issuance records",
        }
        
        searched = ", ".join([
            domain_names.get(d, d) for d in searched_domains
        ])
        
        messages: dict[str, str] = {
            "inventory_check": f"I couldn't find that item in {searched}. Try being more specific or check the exact product name.",
            "material_list": f"No materials found in {searched}. The database might be empty or unavailable.",
            "stock_availability": f"Stock information not available in {searched}.",
            "price_inquiry": f"Pricing information not found. Please check if the product exists in {searched}.",
            "vehicle_history": f"No vehicle records found. Please verify the plate number and try again.",
            "service_schedule": f"No scheduled services found. Would you like to book a new appointment?",
            "maintenance_info": f"No maintenance records found for your search.",
            "general_inquiry": f"I couldn't find matching information in {searched}.",
        }
        
        return messages.get(intent, f"No results found in {searched}.")

# ============================================================================
# FILE: firebase_source.py
# ============================================================================

"""Firebase live data access layer for Firestore and Realtime Database."""

from __future__ import annotations

import json
import os
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Any
import asyncio

import firebase_admin
from firebase_admin import credentials


@dataclass
class FirebaseConfig:
    backend: str
    database_url: str | None
    firestore_collections: dict[str, str]
    realtime_paths: dict[str, str]


class FirebaseSource:
    def __init__(self) -> None:
        self.config = FirebaseConfig(
            backend=os.getenv("FIREBASE_BACKEND", "firestore").strip().lower(),
            database_url=os.getenv("FIREBASE_DATABASE_URL"),
            firestore_collections={
                "item_master": os.getenv("FIREBASE_COLLECTION_ITEM_MASTER", "item_master"),
                "inventory": os.getenv("FIREBASE_COLLECTION_INVENTORY", "inventory"),
                # Optional separate collection that holds stock quantities and reorder levels
                "stock_inventory": os.getenv("FIREBASE_COLLECTION_STOCK", "stock_inventory"),
                "issuance": os.getenv("FIREBASE_COLLECTION_ISSUANCE", "issuance"),
                "products": os.getenv("FIREBASE_COLLECTION_PRODUCTS", "products"),
                "customers": os.getenv("FIREBASE_COLLECTION_CUSTOMERS", "customers"),
                "orders": os.getenv("FIREBASE_COLLECTION_ORDERS", "orders"),
                "services": os.getenv("FIREBASE_COLLECTION_SERVICES", "services"),
                "manual": os.getenv("FIREBASE_COLLECTION_MANUAL", "manual"),
            },
            realtime_paths={
                "item_master": os.getenv("FIREBASE_PATH_ITEM_MASTER", "item_master"),
                "inventory": os.getenv("FIREBASE_PATH_INVENTORY", "inventory"),
                # Optional realtime path for stock records
                "stock_inventory": os.getenv("FIREBASE_PATH_STOCK", "stock_inventory"),
                "issuance": os.getenv("FIREBASE_PATH_ISSUANCE", "issuance"),
                "products": os.getenv("FIREBASE_PATH_PRODUCTS", "products"),
                "customers": os.getenv("FIREBASE_PATH_CUSTOMERS", "customers"),
                "orders": os.getenv("FIREBASE_PATH_ORDERS", "orders"),
                "services": os.getenv("FIREBASE_PATH_SERVICES", "services"),
                "manual": os.getenv("FIREBASE_PATH_MANUAL", "manual"),
            },
        )
        self._initialized = False
        self._last_error: str | None = None
        self._last_refresh_iso: str | None = None

    def _get_credentials(self):
        path = os.getenv("FIREBASE_CREDENTIALS_PATH")
        raw_json = os.getenv("FIREBASE_CREDENTIALS_JSON")

        if path and os.path.exists(path):
            return credentials.Certificate(path)
        if raw_json:
            return credentials.Certificate(json.loads(raw_json))
        return credentials.ApplicationDefault()

    def _initialize(self) -> None:
        if self._initialized:
            return
        try:
            try:
                firebase_admin.get_app()
            except ValueError:
                cred = self._get_credentials()
                options = {}
                if self.config.database_url:
                    options["databaseURL"] = self.config.database_url
                firebase_admin.initialize_app(cred, options=options)
            self._initialized = True
            self._last_error = None
        except Exception as exc:
            self._initialized = False
            self._last_error = str(exc)
            raise

    def get_status(self) -> dict[str, Any]:
        return {
            "backend": self.config.backend,
            "initialized": self._initialized,
            "last_error": self._last_error,
            "last_refresh": self._last_refresh_iso,
        }

    def fetch_live_data(self, limit_per_domain: int = 200) -> list[dict[str, Any]]:
        self._initialize()
        if self.config.backend == "realtime":
            records = self._fetch_from_realtime(limit_per_domain=limit_per_domain)
        else:
            records = self._fetch_from_firestore(limit_per_domain=limit_per_domain)
        self._last_refresh_iso = datetime.now(timezone.utc).isoformat()
        return records

    async def fetch_live_data_async(self, limit_per_domain: int = 200) -> list[dict[str, Any]]:
        """Async wrapper for fetch_live_data. Runs blocking I/O in a thread.

        This provides an async-compatible entrypoint without changing the
        internal firebase_admin (which is synchronous).
        """
        return await asyncio.to_thread(self.fetch_live_data, limit_per_domain)

    def _fetch_from_firestore(self, limit_per_domain: int) -> list[dict[str, Any]]:
        from firebase_admin import firestore

        client = firestore.client()
        records: list[dict[str, Any]] = []

        for domain, collection_name in self.config.firestore_collections.items():
            try:
                stream = client.collection(collection_name).limit(limit_per_domain).stream()
                for doc in stream:
                    data = doc.to_dict() or {}
                    records.append(
                        {
                            "domain": domain,
                            "record_id": doc.id,
                            "source": collection_name,
                            "payload": data,
                        }
                    )
            except Exception as exc:
                self._last_error = f"Firestore collection '{collection_name}': {exc}"

        return records

    def _fetch_from_realtime(self, limit_per_domain: int) -> list[dict[str, Any]]:
        from firebase_admin import db

        records: list[dict[str, Any]] = []
        for domain, path in self.config.realtime_paths.items():
            try:
                snapshot = db.reference(path).get() or {}
                if isinstance(snapshot, dict):
                    for record_id, payload in list(snapshot.items())[:limit_per_domain]:
                        records.append(
                            {
                                "domain": domain,
                                "record_id": str(record_id),
                                "source": path,
                                "payload": payload if isinstance(payload, dict) else {"value": payload},
                            }
                        )
                elif isinstance(snapshot, list):
                    for index, payload in enumerate(snapshot[:limit_per_domain]):
                        records.append(
                            {
                                "domain": domain,
                                "record_id": str(index),
                                "source": path,
                                "payload": payload if isinstance(payload, dict) else {"value": payload},
                            }
                        )
            except Exception as exc:
                self._last_error = f"Realtime path '{path}': {exc}"

        return records

# ============================================================================
# FILE: retrieval.py
# ============================================================================

"""Firebase live retrieval pipeline with vector-first semantic ranking."""

from __future__ import annotations

import os
import re
import time
import importlib
from dataclasses import dataclass, field
from typing import Any

import numpy as np
import logging

from firebase_source import FirebaseSource
from data_normalizer import DataNormalizer, get_data_normalizer


@dataclass
class RetrievedChunk:
    domain: str
    record_id: str
    source: str
    text: str
    payload: dict[str, Any]
    score: float
    retrieval_type: str = "semantic"
    metadata: dict[str, Any] = field(default_factory=dict)


class LiveFirebaseRetriever:
    def __init__(self, firebase_source: FirebaseSource, cache_ttl_seconds: int = 15):
        self.firebase_source = firebase_source
        self.cache_ttl_seconds = cache_ttl_seconds
        self._normalizer = get_data_normalizer()
        self._last_refresh = 0.0
        self._cached_records: list[dict[str, Any]] = []
        self._cached_texts: list[str] = []
        self._embedding_model = None
        self._record_embeddings: np.ndarray | None = None
        self._vector_index = None
        self._vector_backend = os.getenv("VECTOR_INDEX_BACKEND", "faiss").strip().lower()
        self._min_semantic_score = float(os.getenv("MIN_SEMANTIC_SCORE", "0.2"))
        self._chunk_fields_per_segment = int(os.getenv("CHUNK_FIELDS_PER_SEGMENT", "6"))
        self._max_chunks_per_record = int(os.getenv("MAX_CHUNKS_PER_RECORD", "4"))
        self._managed_vector_backend = os.getenv("MANAGED_VECTOR_BACKEND", "none").strip().lower()
        self._managed_index_ready = False
        self._id_to_index: dict[str, int] = {}
        self.logger = logging.getLogger("live_retriever")

        # Simple metrics for monitoring
        self.metrics: dict[str, Any] = {
            "total_queries": 0,
            "last_query_latency_ms": 0.0,
            "last_query_time": None,
            "cached_chunks": 0,
        }

        self._domain_field_priority: dict[str, tuple[str, ...]] = {
            "item_master": ("item", "name", "service", "product", "type", "quantity", "qty", "stock", "uom", "price"),
            "inventory": ("item", "product", "sku", "part", "quantity", "qty", "stock", "reorder_level"),
            "products": ("name", "product", "sku", "category", "price", "stock"),
            "customers": ("name", "customer_name", "contact", "phone", "email", "vehicle"),
            "orders": ("order_id", "status", "item", "quantity", "customer", "date"),
            "services": ("service", "service_type", "vehicle", "status", "schedule", "technician"),
            "issuance": ("item", "quantity", "issued_to", "date", "reference"),
            "manual": ("title", "topic", "section", "content", "steps", "description", "category", "instructions"),
        }
        self._domain_keywords: dict[str, tuple[str, ...]] = {
            "item_master": ("item master", "item_master", "materials", "services", "service offerings", "stock list", "catalog", "inventory", "item"),
            "inventory": ("stock", "inventory", "enough", "restock", "available", "quantity", "qty", "oil", "fluid", "part"),
            "products": ("product", "price", "sku", "catalog", "item"),
            "customers": ("customer", "client", "owner", "contact"),
            "orders": ("order", "purchase", "delivery", "invoice"),
            "services": ("service", "maintenance", "pms", "repair", "appointment", "schedule"),
            "issuance": ("issuance", "issued", "released"),
            "stock_inventory": ("stock", "stock inventory", "available stock", "availability", "remaining", "quantity", "qty", "reorder", "level"),
            "manual": ("how to", "how do i", "manual", "guide", "help", "tutorial", "instructions", "usage", "user guide", "app", "navigate", "where", "button", "feature", "screen", "page", "menu", "setting"),
        }

        model_name = os.getenv(
            "EMBEDDING_MODEL_NAME", "sentence-transformers/all-MiniLM-L6-v2"
        )

        try:
            sentence_transformers = importlib.import_module("sentence_transformers")
            SentenceTransformer = getattr(sentence_transformers, "SentenceTransformer")
            self._embedding_model = SentenceTransformer(model_name)
        except Exception:
            self._embedding_model = None

    def retrieve(
        self,
        query: str,
        top_k: int = 8,
        metadata_filters: dict[str, Any] | None = None,
    ) -> list[RetrievedChunk]:
        start = time.time()
        self._refresh_cache_if_needed()
        if not self._cached_records:
            return []

        merged_filters = self._merge_filters(
            self._parse_query_filters(query), metadata_filters or {}
        )
        candidate_indices = self._filter_candidate_indices(merged_filters)
        # If filters remove all candidates, fall back to a full scan so
        # the embedding/lexical retrieval can still surface relevant records.
        if not candidate_indices:
            candidate_indices = list(range(len(self._cached_records)))

        if self._managed_vector_backend in {"pinecone", "vertex"} and self._managed_index_ready:
            managed = self._retrieve_from_managed_backend(query, top_k, merged_filters)
            if managed:
                latency = (time.time() - start) * 1000.0
                self._record_metrics(latency)
                self.logger.info("managed backend retrieval complete: q=%s top_k=%d latency=%.2fms", query, top_k, latency)
                return managed

        if self._embedding_model is not None and self._record_embeddings is not None:
            results = self._retrieve_by_vectors(
                query=query,
                top_k=top_k,
                candidate_indices=candidate_indices,
            )
            self._record_metrics((time.time() - start) * 1000.0)
            return results

        results = self._retrieve_lexical_fallback(
            query=query,
            top_k=top_k,
            candidate_indices=candidate_indices,
        )
        self._record_metrics((time.time() - start) * 1000.0)
        return results

    def _record_metrics(self, latency_ms: float) -> None:
        self.metrics["total_queries"] += 1
        self.metrics["last_query_latency_ms"] = latency_ms
        self.metrics["last_query_time"] = time.time()
        self.metrics["cached_chunks"] = len(self._cached_records)
        self.metrics["last_cached_records"] = len(self._cached_records)

    def get_metrics(self) -> dict[str, Any]:
        """Return current retrieval observability metrics."""
        return dict(self.metrics)

    def get_metrics(self) -> dict[str, Any]:
        """Return simple retrieval metrics for monitoring."""
        return dict(self.metrics)

    def _refresh_cache_if_needed(self) -> None:
        now = time.time()
        if (now - self._last_refresh) < self.cache_ttl_seconds and self._cached_records:
            return

        raw_records = self.firebase_source.fetch_live_data(limit_per_domain=250)
        # Normalize field names for consistent schema across NoSQL documents
        normalized_records = self._normalizer.normalize_records(raw_records)
        self._cached_records = self._chunk_records(normalized_records)
        self._cached_texts = [rec["text"] for rec in self._cached_records]
        self._id_to_index = {
            rec["chunk_id"]: idx for idx, rec in enumerate(self._cached_records)
        }
        self._last_refresh = now
        self._vector_index = None
        self._managed_index_ready = False

        if self._embedding_model is not None and self._cached_texts:
            encoded = self._embedding_model.encode(
                self._cached_texts,
                normalize_embeddings=True,
                show_progress_bar=False,
            )
            self._record_embeddings = np.array(encoded, dtype=np.float32)
            self._build_vector_index()
            self._managed_index_ready = self._sync_managed_index()
        else:
            self._record_embeddings = None
        # update metrics
        self.metrics["cached_chunks"] = len(self._cached_records)
        self.logger.info(
            "cache refreshed: records=%d chunks=%d embeddings=%s",
            len(raw_records),
            len(self._cached_records),
            self._record_embeddings is not None,
        )

    def _build_vector_index(self) -> None:
        if self._record_embeddings is None or self._record_embeddings.size == 0:
            return
        if self._vector_backend != "faiss":
            self._vector_index = None
            return
        try:
            faiss = importlib.import_module("faiss")
            dim = self._record_embeddings.shape[1]
            index = faiss.IndexFlatIP(dim)
            index.add(self._record_embeddings)
            self._vector_index = index
        except Exception:
            self._vector_index = None

    def prewarm(self, limit_per_domain: int = 250) -> dict[str, object]:
        raw_records = self.firebase_source.fetch_live_data(limit_per_domain=limit_per_domain)
        # Normalize field names for consistent schema across NoSQL documents
        normalized_records = self._normalizer.normalize_records(raw_records)
        self._cached_records = self._chunk_records(normalized_records)
        self._cached_texts = [rec["text"] for rec in self._cached_records]
        self._id_to_index = {
            rec["chunk_id"]: idx for idx, rec in enumerate(self._cached_records)
        }
        self._last_refresh = time.time()
        self._vector_index = None
        self._managed_index_ready = False

        embeddings_built = False
        index_built = False
        if self._embedding_model is not None and self._cached_texts:
            encoded = self._embedding_model.encode(
                self._cached_texts,
                normalize_embeddings=True,
                show_progress_bar=False,
            )
            self._record_embeddings = np.array(encoded, dtype=np.float32)
            embeddings_built = True
            self._build_vector_index()
            index_built = self._vector_index is not None
            self._managed_index_ready = self._sync_managed_index()
        else:
            self._record_embeddings = None

        return {
            "records": len(raw_records),
            "chunks": len(self._cached_records),
            "embeddings_built": embeddings_built,
            "index_built": index_built,
            "managed_backend": self._managed_vector_backend,
            "managed_index_ready": self._managed_index_ready,
        }

    def _retrieve_by_vectors(
        self,
        query: str,
        top_k: int,
        candidate_indices: list[int],
    ) -> list[RetrievedChunk]:
        if self._embedding_model is None or self._record_embeddings is None:
            return []

        encoded_query = self._embedding_model.encode(
            [query],
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        query_vector = np.array(encoded_query, dtype=np.float32)

        full_scan = len(candidate_indices) == len(self._cached_records)
        if self._vector_index is not None and full_scan:
            scores, indexes = self._vector_index.search(query_vector, min(top_k * 3, len(self._cached_records)))
            ranked_pairs = [
                (int(idx), float(score))
                for idx, score in zip(indexes[0], scores[0])
                if idx >= 0
            ]
        else:
            matrix = self._record_embeddings[candidate_indices]
            scores = np.dot(matrix, query_vector[0])
            ranked = np.argsort(scores)[::-1]
            ranked_pairs = [
                (candidate_indices[int(local_idx)], float(scores[int(local_idx)]))
                for local_idx in ranked[: min(top_k * 3, len(ranked))]
            ]

        results: list[RetrievedChunk] = []
        for idx, score in ranked_pairs:
            if len(results) >= top_k:
                break
            if score < self._min_semantic_score:
                continue
            record = self._cached_records[idx]
            results.append(
                RetrievedChunk(
                    domain=record["domain"],
                    record_id=record["record_id"],
                    source=record["source"],
                    text=self._cached_texts[idx],
                    payload=record["payload"],
                    score=score,
                )
            )

        if results:
            return results

        if ranked_pairs:
            best_idx, best_score = ranked_pairs[0]
            best_record = self._cached_records[best_idx]
            return [
                RetrievedChunk(
                    domain=best_record["domain"],
                    record_id=best_record["record_id"],
                    source=best_record["source"],
                    text=self._cached_texts[best_idx],
                    payload=best_record["payload"],
                    score=best_score,
                )
            ]

        return []

    @staticmethod
    def _tokenize(text: str) -> set[str]:
        return set(re.findall(r"[a-z0-9_]+", text.lower()))

    def _lexical_score(self, query: str, query_tokens: set[str], text: str) -> float:
        text_lower = text.lower()
        text_tokens = self._tokenize(text_lower)
        overlap = len(query_tokens.intersection(text_tokens))
        exact_bonus = 0.3 if query.lower() in text_lower else 0.0
        domain_bonus = 0.15 if any(d in query.lower() for d in ["inventory", "customer", "order", "service", "product"]) else 0.0
        return overlap + exact_bonus + domain_bonus

    def _retrieve_lexical_fallback(
        self,
        query: str,
        top_k: int,
        candidate_indices: list[int],
    ) -> list[RetrievedChunk]:
        query_tokens = self._tokenize(query)
        lexical_scores = [
            (
                idx,
                self._lexical_score(
                    query=query,
                    query_tokens=query_tokens,
                    text=self._cached_texts[idx],
                ),
            )
            for idx in candidate_indices
        ]
        ranked = sorted(lexical_scores, key=lambda item: item[1], reverse=True)

        results: list[RetrievedChunk] = []
        for idx, score in ranked[:top_k]:
            if score <= 0:
                continue
            record = self._cached_records[idx]
            results.append(
                RetrievedChunk(
                    domain=record["domain"],
                    record_id=record["record_id"],
                    source=record["source"],
                    text=self._cached_texts[idx],
                    payload=record["payload"],
                    score=float(score),
                    retrieval_type="lexical",
                    metadata={
                        "source_collection": record["source"],
                        "document_id": record["record_id"],
                        "retrieval_score": float(score),
                        "retrieval_type": "lexical",
                    },
                )
            )
        return results

    def _chunk_records(self, raw_records: list[dict[str, Any]]) -> list[dict[str, Any]]:
        chunked: list[dict[str, Any]] = []
        for record in raw_records:
            domain = str(record.get("domain", "")).lower()
            payload = record.get("payload", {}) or {}
            flat_payload = self._flatten(payload)
            segments = self._build_domain_segments(domain, flat_payload)

            if not segments:
                segments = [flat_payload]

            for chunk_index, segment_payload in enumerate(
                segments[: self._max_chunks_per_record]
            ):
                chunk_id = (
                    f"{record.get('domain', 'unknown')}"
                    f":{record.get('record_id', 'unknown')}:{chunk_index}"
                )
                metadata = {
                    "domain": record.get("domain"),
                    "source": record.get("source"),
                    "record_id": str(record.get("record_id")),
                    "chunk_index": chunk_index,
                }
                chunked.append(
                    {
                        "chunk_id": chunk_id,
                        "domain": record.get("domain"),
                        "record_id": str(record.get("record_id")),
                        "source": record.get("source"),
                        "payload": segment_payload,
                        "metadata": metadata,
                        "text": self._record_to_text(
                            {
                                **record,
                                "payload": segment_payload,
                                "chunk_id": chunk_id,
                            }
                        ),
                    }
                )
        return chunked

    def _build_domain_segments(
        self,
        domain: str,
        flat_payload: dict[str, Any],
    ) -> list[dict[str, Any]]:
        if not flat_payload:
            return []

        items = list(flat_payload.items())
        priorities = self._domain_field_priority.get(domain, ())
        prioritized: list[tuple[str, Any]] = []
        remaining: list[tuple[str, Any]] = []
        for key, value in items:
            key_lower = key.lower()
            if any(token in key_lower for token in priorities):
                prioritized.append((key, value))
            else:
                remaining.append((key, value))

        ordered = prioritized + remaining
        segments: list[dict[str, Any]] = []
        step = max(self._chunk_fields_per_segment, 1)
        for idx in range(0, len(ordered), step):
            segment_items = ordered[idx : idx + step]
            segments.append({k: v for k, v in segment_items})
        return segments

    def _parse_query_filters(self, query: str) -> dict[str, Any]:
        lowered = query.lower()
        filters: dict[str, Any] = {}

        domain_scores: dict[str, int] = {}
        for domain, keywords in self._domain_keywords.items():
            score = sum(1 for keyword in keywords if keyword in lowered)
            if score > 0:
                domain_scores[domain] = score

        if domain_scores:
            top_score = max(domain_scores.values())
            strongest_domains = [
                domain
                for domain, score in domain_scores.items()
                if score == top_score
            ]
            # Prefer explicit stock domain when query mentions stock terms
            lowered = query.lower()
            if ("stock" in lowered or "in stock" in lowered or "availability" in lowered) and "stock_inventory" in domain_scores:
                filters["domain"] = ["stock_inventory"]
            else:
                filters["domain"] = sorted(strongest_domains)

        explicit_pairs = re.findall(
            r"\b(domain|source|record_id|record|id|type|category|sku|name|item):([a-zA-Z0-9_.\-]+)",
            lowered,
        )
        for key, value in explicit_pairs:
            normalized_key = "record_id" if key in {"record", "id"} else key
            filters[normalized_key] = value

        return filters

    def _merge_filters(
        self,
        query_filters: dict[str, Any],
        metadata_filters: dict[str, Any],
    ) -> dict[str, Any]:
        merged = dict(query_filters)
        for key, value in metadata_filters.items():
            if value is None:
                continue
            if key in merged and isinstance(merged[key], list):
                existing = set(merged[key])
                if isinstance(value, list):
                    merged[key] = sorted(existing.intersection(set(value)) or existing.union(set(value)))
                else:
                    merged[key] = [value] if value in existing else sorted(existing.union({value}))
            else:
                merged[key] = value
        return merged

    def _normalize_filter_values(self, value: Any) -> set[str]:
        if isinstance(value, list):
            return {str(item).lower() for item in value}
        return {str(value).lower()}

    def _filter_candidate_indices(self, filters: dict[str, Any]) -> list[int]:
        if not filters:
            return list(range(len(self._cached_records)))

        domain_values = self._normalize_filter_values(filters["domain"]) if "domain" in filters else None
        source_values = self._normalize_filter_values(filters["source"]) if "source" in filters else None
        record_values = self._normalize_filter_values(filters["record_id"]) if "record_id" in filters else None

        candidates: list[int] = []
        for idx, record in enumerate(self._cached_records):
            metadata = record.get("metadata", {})
            if domain_values and str(metadata.get("domain", "")).lower() not in domain_values:
                continue
            if source_values and str(metadata.get("source", "")).lower() not in source_values:
                continue
            if record_values and str(metadata.get("record_id", "")).lower() not in record_values:
                continue

            # Additional payload-level filters (type, category, sku, name, item)
            payload = record.get("payload", {}) or {}
            payload_norm = {str(k).lower(): str(v).lower() for k, v in payload.items() if v is not None}

            blocked = False
            for key in ("type", "category", "sku", "name", "item"):
                if key in filters:
                    required = {str(filters[key]).lower()} if not isinstance(filters[key], list) else {str(v).lower() for v in filters[key]}
                    # If none of the payload values match required, block this record
                    if not any(val in payload_norm.get(key, "") for val in required):
                        blocked = True
                        break

            if blocked:
                continue

            candidates.append(idx)
        return candidates

    def _sync_managed_index(self) -> bool:
        if self._managed_vector_backend == "pinecone":
            return self._upsert_pinecone()
        if self._managed_vector_backend == "vertex":
            return self._upsert_vertex()
        return False

    def _retrieve_from_managed_backend(
        self,
        query: str,
        top_k: int,
        filters: dict[str, Any],
    ) -> list[RetrievedChunk]:
        if self._embedding_model is None:
            return []

        vector = self._embedding_model.encode(
            [query],
            normalize_embeddings=True,
            show_progress_bar=False,
        )
        query_vector = np.array(vector[0], dtype=np.float32)

        if self._managed_vector_backend == "pinecone":
            pairs = self._query_pinecone(query_vector, top_k, filters)
        elif self._managed_vector_backend == "vertex":
            pairs = self._query_vertex(query_vector, top_k, filters)
        else:
            pairs = []

        if not pairs:
            return []

        out: list[RetrievedChunk] = []
        for idx, score in pairs:
            record = self._cached_records[idx]
            out.append(
                RetrievedChunk(
                    domain=record["domain"],
                    record_id=record["record_id"],
                    source=record["source"],
                    text=record["text"],
                    payload=record["payload"],
                    score=float(score),
                    retrieval_type="managed",
                    metadata={
                        "source_collection": record["source"],
                        "document_id": record["record_id"],
                        "retrieval_score": float(score),
                        "retrieval_type": "managed",
                    },
                )
            )
            if len(out) >= top_k:
                break
        return out

    def _build_managed_filter(self, filters: dict[str, Any]) -> dict[str, Any]:
        managed: dict[str, Any] = {}
        for key in ("domain", "source", "record_id"):
            if key not in filters:
                continue
            value = filters[key]
            if isinstance(value, list):
                if len(value) == 1:
                    managed[key] = {"$eq": str(value[0])}
                else:
                    managed[key] = {"$in": [str(item) for item in value]}
            else:
                managed[key] = {"$eq": str(value)}
        return managed

    def _upsert_pinecone(self) -> bool:
        api_key = os.getenv("PINECONE_API_KEY")
        index_name = os.getenv("PINECONE_INDEX_NAME")
        if not api_key or not index_name or self._record_embeddings is None:
            return False
        try:
            pinecone_module = importlib.import_module("pinecone")
            Pinecone = getattr(pinecone_module, "Pinecone")
            client = Pinecone(api_key=api_key)
            index = client.Index(index_name)

            vectors = []
            for idx, record in enumerate(self._cached_records):
                vectors.append(
                    {
                        "id": record["chunk_id"],
                        "values": self._record_embeddings[idx].tolist(),
                        "metadata": record["metadata"],
                    }
                )

            batch_size = 100
            for i in range(0, len(vectors), batch_size):
                index.upsert(vectors=vectors[i : i + batch_size])
            return True
        except Exception:
            return False

    def _query_pinecone(
        self,
        query_vector: np.ndarray,
        top_k: int,
        filters: dict[str, Any],
    ) -> list[tuple[int, float]]:
        api_key = os.getenv("PINECONE_API_KEY")
        index_name = os.getenv("PINECONE_INDEX_NAME")
        if not api_key or not index_name:
            return []
        try:
            pinecone_module = importlib.import_module("pinecone")
            Pinecone = getattr(pinecone_module, "Pinecone")
            client = Pinecone(api_key=api_key)
            index = client.Index(index_name)
            managed_filter = self._build_managed_filter(filters)
            response = index.query(
                vector=query_vector.tolist(),
                top_k=min(top_k * 3, len(self._cached_records)),
                include_metadata=True,
                filter=managed_filter if managed_filter else None,
            )

            pairs: list[tuple[int, float]] = []
            for match in response.get("matches", []):
                chunk_id = match.get("id")
                if chunk_id not in self._id_to_index:
                    continue
                pairs.append((self._id_to_index[chunk_id], float(match.get("score", 0.0))))
            return pairs
        except Exception:
            return []

    def _upsert_vertex(self) -> bool:
        # Vertex Vector Search index update flow is environment-specific.
        # Adapter hook is provided here so prewarm() can call a managed backend consistently.
        return False

    def _query_vertex(
        self,
        query_vector: np.ndarray,
        top_k: int,
        filters: dict[str, Any],
    ) -> list[tuple[int, float]]:
        # Vertex query adapter placeholder for managed retrieval path.
        # Keep local retrieval as fallback when this is not configured.
        _ = query_vector, top_k, filters
        return []

    def _record_to_text(self, record: dict[str, Any]) -> str:
        flat_payload = self._flatten(record.get("payload", {}))

        # For inventory/products domains, only include whitelisted fields to avoid leaking internal metadata
        domain = str(record.get("domain", "")).lower()
        if domain in {"inventory", "products"}:
            whitelist = {"sku", "item", "item_name", "name", "service", "service_name", "uom", "type", "qty", "quantity", "stock", "price"}
            filtered = {k: v for k, v in flat_payload.items() if k.lower() in whitelist}
            joined_fields = ", ".join(f"{key}={value}" for key, value in filtered.items())
        else:
            joined_fields = ", ".join(f"{key}={value}" for key, value in flat_payload.items())
        chunk_id = record.get("chunk_id")
        chunk_token = f"chunk={chunk_id} " if chunk_id else ""
        return (
            f"domain={record.get('domain')} "
            f"id={record.get('record_id')} "
            f"source={record.get('source')} "
            f"{chunk_token}"
            f"{joined_fields}"
        )

    def _flatten(self, data: dict[str, Any], prefix: str = "") -> dict[str, Any]:
        out: dict[str, Any] = {}
        for key, value in (data or {}).items():
            full_key = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, dict):
                out.update(self._flatten(value, prefix=full_key))
            elif isinstance(value, list):
                out[full_key] = ", ".join(str(v) for v in value)
            else:
                out[full_key] = value
        return out

# ============================================================================
# FILE: retrieval_metrics.py
# ============================================================================

"""Lightweight retrieval monitoring primitives."""

from __future__ import annotations

from dataclasses import dataclass, field
from threading import Lock
import time
from typing import Any


@dataclass(slots=True)
class RetrievalMetricSnapshot:
    total_queries: int = 0
    structured_queries: int = 0
    semantic_queries: int = 0
    hybrid_queries: int = 0
    clarify_queries: int = 0
    low_confidence_queries: int = 0
    average_latency_ms: float = 0.0
    last_latency_ms: float = 0.0
    last_query_at: float | None = None
    last_route: str = ""
    extra: dict[str, Any] = field(default_factory=dict)


class RetrievalMetrics:
    def __init__(self) -> None:
        self._lock = Lock()
        self._snapshot = RetrievalMetricSnapshot()
        self._latencies: list[float] = []

    def record(self, route: str, latency_ms: float, confidence: float, low_confidence: bool = False) -> None:
        with self._lock:
            self._snapshot.total_queries += 1
            if route == "structured":
                self._snapshot.structured_queries += 1
            elif route == "semantic":
                self._snapshot.semantic_queries += 1
            elif route == "hybrid":
                self._snapshot.hybrid_queries += 1
            elif route == "clarify":
                self._snapshot.clarify_queries += 1

            if low_confidence:
                self._snapshot.low_confidence_queries += 1

            self._snapshot.last_latency_ms = latency_ms
            self._snapshot.last_query_at = time.time()
            self._snapshot.last_route = route

            self._latencies.append(latency_ms)
            if self._latencies:
                self._snapshot.average_latency_ms = sum(self._latencies) / len(self._latencies)

            self._snapshot.extra = {
                "last_confidence": round(confidence, 3),
            }

    def snapshot(self) -> dict[str, Any]:
        with self._lock:
            data = self._snapshot.__dict__.copy()
            data["extra"] = dict(self._snapshot.extra)
            return data


# ============================================================================
# FILE: retrieval_access_control.py
# ============================================================================

"""Centralized role-based retrieval access control for the RAG pipeline.

This module enforces retrieval-level permissions BEFORE any data reaches
the LLM context, reranker, or prompt builder. It implements:

1. Role definitions (customer, staff, admin)
2. Collection-level access policies
3. Field-level sensitivity filtering
4. Retrieval authorization decisions
5. Security metadata tagging on chunks

Security invariant: sensitive data must NEVER reach the LLM prompt for
unauthorized users. Authorization happens BEFORE reranking and prompt assembly.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
from typing import Any


# ---------------------------------------------------------------------------
# Role definitions
# ---------------------------------------------------------------------------


class UserRole(str, Enum):
    """Supported user roles with ascending privilege levels."""
    CUSTOMER = "customer"
    STAFF = "staff"
    ADMIN = "admin"


class SensitivityLevel(str, Enum):
    """Data sensitivity classification."""
    PUBLIC = "public"           # Visible to all authenticated users
    INTERNAL = "internal"      # Staff and admin only
    CONFIDENTIAL = "confidential"  # Admin only
    RESTRICTED = "restricted"  # Admin only, audit-logged


# ---------------------------------------------------------------------------
# Policy dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class CollectionPolicy:
    """Access policy for a single Firestore collection/domain."""
    collection: str
    allowed_roles: frozenset[UserRole]
    sensitivity_level: SensitivityLevel
    description: str = ""


@dataclass(frozen=True)
class FieldPolicy:
    """Defines which fields are sensitive and should be stripped per role."""
    field_name: str
    min_role: UserRole  # Minimum role required to see this field
    sensitivity_level: SensitivityLevel


@dataclass
class AuthorizationResult:
    """Result of a retrieval authorization check."""
    authorized: bool
    allowed_collections: list[str]
    denied_collections: list[str]
    reason: str = ""
    filtered_fields: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Permission policies — centralized source of truth
# ---------------------------------------------------------------------------

# Collection-level access control
COLLECTION_POLICIES: dict[str, CollectionPolicy] = {
    "item_master": CollectionPolicy(
        collection="item_master",
        allowed_roles=frozenset({UserRole.CUSTOMER, UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.PUBLIC,
        description="Product/service catalog — public info",
    ),
    "inventory": CollectionPolicy(
        collection="inventory",
        allowed_roles=frozenset({UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.INTERNAL,
        description="Internal stock levels and reorder data",
    ),
    "stock_inventory": CollectionPolicy(
        collection="stock_inventory",
        allowed_roles=frozenset({UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.INTERNAL,
        description="Stock quantities and reorder levels",
    ),
    "issuance": CollectionPolicy(
        collection="issuance",
        allowed_roles=frozenset({UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.INTERNAL,
        description="Material issuance records",
    ),
    "products": CollectionPolicy(
        collection="products",
        allowed_roles=frozenset({UserRole.CUSTOMER, UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.PUBLIC,
        description="Product catalog — public pricing and info",
    ),
    "customers": CollectionPolicy(
        collection="customers",
        allowed_roles=frozenset({UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.CONFIDENTIAL,
        description="Customer PII — staff/admin only (customers see own records via app logic)",
    ),
    "orders": CollectionPolicy(
        collection="orders",
        allowed_roles=frozenset({UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.INTERNAL,
        description="Order records — staff/admin only (customers see own orders via app logic)",
    ),
    "services": CollectionPolicy(
        collection="services",
        allowed_roles=frozenset({UserRole.CUSTOMER, UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.PUBLIC,
        description="Service offerings and schedules — public",
    ),
    "manual": CollectionPolicy(
        collection="manual",
        allowed_roles=frozenset({UserRole.CUSTOMER, UserRole.STAFF, UserRole.ADMIN}),
        sensitivity_level=SensitivityLevel.PUBLIC,
        description="User manual and app usage guides — public for all users",
    ),
}

# Field-level sensitivity — fields that must be stripped before context assembly
SENSITIVE_FIELDS: list[FieldPolicy] = [
    # Admin-only fields
    FieldPolicy("internal_notes", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("admin_notes", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("confidential_remarks", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("admin_metadata", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("audit_log", UserRole.ADMIN, SensitivityLevel.RESTRICTED),
    FieldPolicy("internal_id", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("owner_id", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("user_id", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("mechanic_id", UserRole.ADMIN, SensitivityLevel.CONFIDENTIAL),

    # Staff+ fields (cost/profit/supplier)
    FieldPolicy("cost", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("unit_cost", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("wholesale_price", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("margin", UserRole.STAFF, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("markup", UserRole.STAFF, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("profit", UserRole.STAFF, SensitivityLevel.CONFIDENTIAL),
    FieldPolicy("supplier", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("supplier_name", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("supplier_info", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("reorder_level", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("reorder_point", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("min_stock", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("reorder_suggestion", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("internal_code", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("product_code", UserRole.STAFF, SensitivityLevel.INTERNAL),
    FieldPolicy("sku", UserRole.STAFF, SensitivityLevel.INTERNAL),

    # Embedding/vector fields — never exposed
    FieldPolicy("embedding", UserRole.ADMIN, SensitivityLevel.RESTRICTED),
    FieldPolicy("vector", UserRole.ADMIN, SensitivityLevel.RESTRICTED),
    FieldPolicy("hash", UserRole.ADMIN, SensitivityLevel.RESTRICTED),
]

# Build a lookup for fast field filtering
_FIELD_POLICY_MAP: dict[str, FieldPolicy] = {
    fp.field_name.lower(): fp for fp in SENSITIVE_FIELDS
}

# Role hierarchy for comparison
_ROLE_HIERARCHY: dict[UserRole, int] = {
    UserRole.CUSTOMER: 0,
    UserRole.STAFF: 1,
    UserRole.ADMIN: 2,
}


# ---------------------------------------------------------------------------
# Core authorization engine
# ---------------------------------------------------------------------------


class RetrievalAccessControl:
    """Enforces role-based access control on retrieval operations.

    Usage:
        rac = RetrievalAccessControl()
        auth = rac.authorize_retrieval(user_role="customer", requested_domains=["inventory", "products"])
        if not auth.authorized:
            return access_denied_response(auth.reason)
        # proceed with retrieval using only auth.allowed_collections
    """

    def __init__(
        self,
        collection_policies: dict[str, CollectionPolicy] | None = None,
        field_policies: list[FieldPolicy] | None = None,
    ) -> None:
        self.collection_policies = collection_policies or COLLECTION_POLICIES
        self.field_policies = field_policies or SENSITIVE_FIELDS
        self._field_map: dict[str, FieldPolicy] = {
            fp.field_name.lower(): fp for fp in self.field_policies
        }

    def resolve_role(self, user_type: str) -> UserRole:
        """Resolve a user_type string to a UserRole enum."""
        try:
            return UserRole(user_type.lower().strip())
        except (ValueError, AttributeError):
            # Default to most restrictive role for unknown types
            return UserRole.CUSTOMER

    def authorize_retrieval(
        self,
        user_role: str | UserRole,
        requested_domains: list[str] | None = None,
    ) -> AuthorizationResult:
        """Determine which collections a user role can access.

        Args:
            user_role: The user's role (string or enum)
            requested_domains: Specific domains the query targets.
                If None, returns all allowed collections for the role.

        Returns:
            AuthorizationResult with allowed/denied collections.
        """
        role = self.resolve_role(user_role) if isinstance(user_role, str) else user_role

        # Determine all collections this role can access
        all_allowed = [
            name for name, policy in self.collection_policies.items()
            if role in policy.allowed_roles
        ]

        if requested_domains is None:
            return AuthorizationResult(
                authorized=True,
                allowed_collections=all_allowed,
                denied_collections=[],
                reason="Full access granted for role",
            )

        # Check each requested domain
        allowed: list[str] = []
        denied: list[str] = []

        for domain in requested_domains:
            domain_lower = domain.lower().strip()
            policy = self.collection_policies.get(domain_lower)

            if policy is None:
                # Unknown collection — deny by default (fail-closed)
                denied.append(domain_lower)
                continue

            if role in policy.allowed_roles:
                allowed.append(domain_lower)
            else:
                denied.append(domain_lower)

        # If ALL requested domains are denied, the retrieval is unauthorized
        authorized = len(allowed) > 0

        reason = ""
        if denied:
            reason = f"Access denied to collections: {', '.join(denied)}"
        if not authorized:
            reason = f"Role '{role.value}' has no access to requested collections: {', '.join(denied)}"

        return AuthorizationResult(
            authorized=authorized,
            allowed_collections=allowed,
            denied_collections=denied,
            reason=reason,
        )

    def get_allowed_domains_for_role(self, user_role: str | UserRole) -> list[str]:
        """Return all domain names accessible by the given role."""
        role = self.resolve_role(user_role) if isinstance(user_role, str) else user_role
        return [
            name for name, policy in self.collection_policies.items()
            if role in policy.allowed_roles
        ]

    def filter_sensitive_fields(
        self,
        payload: dict[str, Any],
        user_role: str | UserRole,
    ) -> dict[str, Any]:
        """Remove fields the user role is not authorized to see.

        This is applied BEFORE context is assembled for the LLM prompt.
        """
        role = self.resolve_role(user_role) if isinstance(user_role, str) else user_role
        role_level = _ROLE_HIERARCHY.get(role, 0)

        filtered: dict[str, Any] = {}
        for key, value in payload.items():
            key_lower = key.lower().strip()
            policy = self._field_map.get(key_lower)

            if policy is not None:
                required_level = _ROLE_HIERARCHY.get(policy.min_role, 0)
                if role_level < required_level:
                    # User does not have sufficient privilege — strip field
                    continue

            # Recursively filter nested dicts
            if isinstance(value, dict):
                value = self.filter_sensitive_fields(value, role)

            filtered[key] = value

        return filtered

    def get_stripped_field_names(self, user_role: str | UserRole) -> list[str]:
        """Return field names that will be stripped for this role."""
        role = self.resolve_role(user_role) if isinstance(user_role, str) else user_role
        role_level = _ROLE_HIERARCHY.get(role, 0)

        stripped: list[str] = []
        for fp in self.field_policies:
            required_level = _ROLE_HIERARCHY.get(fp.min_role, 0)
            if role_level < required_level:
                stripped.append(fp.field_name)
        return stripped

    def tag_chunk_metadata(
        self,
        chunk_metadata: dict[str, Any],
        source_collection: str,
    ) -> dict[str, Any]:
        """Add security metadata to a retrieval chunk for audit/tracking."""
        policy = self.collection_policies.get(source_collection.lower())
        if policy:
            chunk_metadata["allowed_roles"] = [r.value for r in policy.allowed_roles]
            chunk_metadata["sensitivity_level"] = policy.sensitivity_level.value
            chunk_metadata["source_collection"] = source_collection
        else:
            # Unknown collection — mark as restricted
            chunk_metadata["allowed_roles"] = [UserRole.ADMIN.value]
            chunk_metadata["sensitivity_level"] = SensitivityLevel.RESTRICTED.value
            chunk_metadata["source_collection"] = source_collection
        return chunk_metadata

    def is_chunk_accessible(
        self,
        chunk_domain: str,
        user_role: str | UserRole,
    ) -> bool:
        """Quick check: can this role access chunks from this domain?"""
        role = self.resolve_role(user_role) if isinstance(user_role, str) else user_role
        policy = self.collection_policies.get(chunk_domain.lower())
        if policy is None:
            return role == UserRole.ADMIN  # Unknown domains: admin only
        return role in policy.allowed_roles


# ---------------------------------------------------------------------------
# Access-denied response helpers
# ---------------------------------------------------------------------------

ACCESS_DENIED_CUSTOMER = "Access denied. You can only access your own records."
ACCESS_DENIED_STAFF = "Access denied. This data requires administrator privileges."
ACCESS_DENIED_GENERIC = "Access denied. You do not have permission to access this data."


def build_access_denied_response(
    user_role: str | UserRole,
    auth_result: AuthorizationResult,
) -> str:
    """Build a safe, non-leaking access-denied message."""
    role = user_role if isinstance(user_role, UserRole) else UserRole(user_role.lower())

    if role == UserRole.CUSTOMER:
        return ACCESS_DENIED_CUSTOMER
    if role == UserRole.STAFF:
        return ACCESS_DENIED_STAFF
    return ACCESS_DENIED_GENERIC


# ---------------------------------------------------------------------------
# Module-level singleton for convenience
# ---------------------------------------------------------------------------

_default_rac: RetrievalAccessControl | None = None


def get_access_control() -> RetrievalAccessControl:
    """Return the module-level RetrievalAccessControl singleton."""
    global _default_rac
    if _default_rac is None:
        _default_rac = RetrievalAccessControl()
    return _default_rac


# ============================================================================
# FILE: secure_context_builder.py
# ============================================================================

"""Secure context builder — filters retrieval chunks before LLM prompt assembly.

This module sits between the retrieval pipeline and the prompt builder,
ensuring that:
1. Only authorized chunks reach the LLM context
2. Sensitive fields are stripped from chunk payloads
3. Security metadata is attached for audit
4. No hidden metadata leaks through citations or text

Security invariant: After processing through this module, the resulting
chunks are SAFE to inject into any LLM prompt for the given user role.
"""

from __future__ import annotations

from typing import Any

from retrieval import RetrievedChunk
from retrieval_access_control import (
    RetrievalAccessControl,
    UserRole,
    get_access_control,
)


class SecureContextBuilder:
    """Filter and sanitize retrieval chunks for safe LLM context construction.

    Usage:
        builder = SecureContextBuilder()
        safe_chunks = builder.filter_chunks(chunks, user_role="customer")
        # safe_chunks are now safe to pass to the prompt builder
    """

    def __init__(
        self,
        access_control: RetrievalAccessControl | None = None,
    ) -> None:
        self.access_control = access_control or get_access_control()

    def filter_chunks(
        self,
        chunks: list[RetrievedChunk],
        user_role: str | UserRole,
    ) -> list[RetrievedChunk]:
        """Apply collection-level and field-level filtering to chunks.

        This is the primary entry point. Call this BEFORE passing chunks
        to the reranker or prompt builder.

        Args:
            chunks: Raw retrieval results
            user_role: The requesting user's role

        Returns:
            Filtered list of chunks safe for the given role
        """
        role = (
            self.access_control.resolve_role(user_role)
            if isinstance(user_role, str)
            else user_role
        )

        safe_chunks: list[RetrievedChunk] = []

        for chunk in chunks:
            # 1. Collection-level check — reject entire chunk if domain is restricted
            if not self.access_control.is_chunk_accessible(chunk.domain, role):
                continue

            # 2. Field-level filtering on payload
            filtered_payload = self.access_control.filter_sensitive_fields(
                chunk.payload, role
            )

            # 3. Rebuild text without sensitive fields
            sanitized_text = self._rebuild_chunk_text(chunk, filtered_payload)

            # 4. Sanitize metadata — remove internal tracking fields
            safe_metadata = self._sanitize_metadata(chunk.metadata, role)

            # 5. Tag with security metadata for audit
            safe_metadata = self.access_control.tag_chunk_metadata(
                safe_metadata, chunk.source or chunk.domain
            )

            # 6. Create sanitized chunk
            safe_chunk = RetrievedChunk(
                domain=chunk.domain,
                record_id=chunk.record_id,
                source=chunk.source,
                text=sanitized_text,
                payload=filtered_payload,
                score=chunk.score,
                retrieval_type=chunk.retrieval_type,
                metadata=safe_metadata,
            )
            safe_chunks.append(safe_chunk)

        return safe_chunks

    def filter_citations(
        self,
        chunks: list[RetrievedChunk],
        user_role: str | UserRole,
    ) -> list[RetrievedChunk]:
        """Filter chunks for citation building — ensures no restricted domains leak."""
        role = (
            self.access_control.resolve_role(user_role)
            if isinstance(user_role, str)
            else user_role
        )
        return [
            chunk for chunk in chunks
            if self.access_control.is_chunk_accessible(chunk.domain, role)
        ]

    def _rebuild_chunk_text(
        self,
        original_chunk: RetrievedChunk,
        filtered_payload: dict[str, Any],
    ) -> str:
        """Rebuild the chunk text using only the filtered (safe) payload fields.

        This prevents sensitive field values from appearing in the text
        representation that gets injected into the LLM prompt.
        """
        # Build text from filtered payload only
        field_parts = ", ".join(
            f"{key}={value}"
            for key, value in filtered_payload.items()
            if value not in (None, "", "none", "null", "n/a", "na", "unknown", "-")
        )

        return (
            f"domain={original_chunk.domain} "
            f"source={original_chunk.source} "
            f"{field_parts}"
        )

    def _sanitize_metadata(
        self,
        metadata: dict[str, Any],
        role: UserRole,
    ) -> dict[str, Any]:
        """Remove internal metadata fields that should not be exposed."""
        # Fields that should never appear in metadata exposed to non-admins
        internal_fields = {
            "chunk_id", "chunk_index", "embedding", "vector",
            "hash", "internal_id", "owner_id", "user_id",
        }

        safe: dict[str, Any] = {}
        for key, value in (metadata or {}).items():
            key_lower = key.lower()
            if key_lower in internal_fields and role != UserRole.ADMIN:
                continue
            safe[key] = value

        return safe

    def get_security_summary(
        self,
        original_count: int,
        filtered_count: int,
        user_role: str | UserRole,
    ) -> dict[str, Any]:
        """Return a summary of security filtering applied."""
        role = (
            self.access_control.resolve_role(user_role)
            if isinstance(user_role, str)
            else user_role
        )
        stripped_fields = self.access_control.get_stripped_field_names(role)
        allowed_domains = self.access_control.get_allowed_domains_for_role(role)

        return {
            "user_role": role.value,
            "original_chunk_count": original_count,
            "filtered_chunk_count": filtered_count,
            "chunks_removed": original_count - filtered_count,
            "fields_stripped": stripped_fields,
            "allowed_domains": allowed_domains,
        }


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_builder: SecureContextBuilder | None = None


def get_secure_context_builder() -> SecureContextBuilder:
    """Return the module-level SecureContextBuilder singleton."""
    global _default_builder
    if _default_builder is None:
        _default_builder = SecureContextBuilder()
    return _default_builder


# ============================================================================
# FILE: retrieval_ranker.py
# ============================================================================

"""Ranking and deduplication utilities for hybrid retrieval."""

from __future__ import annotations

from datetime import datetime, timezone
import re
from typing import Any

from retrieval import RetrievedChunk


class RetrievalRanker:
    """Rank retrieved chunks using lexical overlap and metadata signals."""

    def rank(self, query: str, chunks: list[RetrievedChunk], top_k: int) -> list[RetrievedChunk]:
        if not chunks:
            return []

        query_tokens = self._tokenize(query)
        scored: list[tuple[RetrievedChunk, float]] = []
        for chunk in chunks:
            score = self._score_chunk(query_tokens, chunk)
            scored.append((chunk, score))

        deduped: list[tuple[RetrievedChunk, float]] = []
        seen: set[tuple[str, str, str]] = set()
        for chunk, score in sorted(scored, key=lambda item: item[1], reverse=True):
            key = (chunk.domain.lower(), chunk.record_id.lower(), self._normalize_text(chunk.text))
            if key in seen:
                continue
            seen.add(key)
            deduped.append((chunk, score))

        deduped.sort(key=lambda item: item[1], reverse=True)
        out: list[RetrievedChunk] = []
        for chunk, score in deduped[:top_k]:
            chunk.score = float(score)
            chunk.metadata = {
                **(chunk.metadata or {}),
                "source_collection": chunk.source,
                "document_id": chunk.record_id,
                "retrieval_score": round(float(score), 4),
                "retrieval_type": chunk.retrieval_type,
            }
            out.append(chunk)
        return out

    def confidence(self, chunks: list[RetrievedChunk]) -> float:
        if not chunks:
            return 0.0
        top_score = max(chunk.score for chunk in chunks)
        second_score = chunks[1].score if len(chunks) > 1 else 0.0
        spread = max(0.0, top_score - second_score)
        return max(0.0, min(1.0, (top_score * 0.7) + (spread * 0.3)))

    def _score_chunk(self, query_tokens: set[str], chunk: RetrievedChunk) -> float:
        text_tokens = self._tokenize(chunk.text)
        payload_tokens = self._tokenize(self._flatten_payload(chunk.payload))

        lexical_overlap = len(query_tokens.intersection(text_tokens.union(payload_tokens)))
        field_bonus = self._payload_field_bonus(chunk.payload, query_tokens)
        route_bonus = 0.12 if chunk.retrieval_type == "structured" else 0.08 if chunk.retrieval_type == "hybrid" else 0.0
        base = float(chunk.score or 0.0)

        return base + (lexical_overlap * 0.12) + field_bonus + route_bonus

    def _payload_field_bonus(self, payload: dict[str, Any], query_tokens: set[str]) -> float:
        if not payload:
            return 0.0
        bonus = 0.0
        for key, value in payload.items():
            key_tokens = self._tokenize(str(key))
            if query_tokens.intersection(key_tokens):
                bonus += 0.08
            value_tokens = self._tokenize(self._clean_value(value))
            if query_tokens.intersection(value_tokens):
                bonus += 0.06
        return min(bonus, 0.35)

    def _flatten_payload(self, payload: dict[str, Any]) -> str:
        parts: list[str] = []
        for key, value in (payload or {}).items():
            parts.append(f"{key} {self._clean_value(value)}")
        return " ".join(parts)

    def _tokenize(self, text: str) -> set[str]:
        return {token for token in re.findall(r"[a-z0-9_]+", (text or "").lower()) if token}

    def _clean_value(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, (list, tuple, set)):
            return " ".join(self._clean_value(item) for item in value)
        if isinstance(value, dict):
            return " ".join(f"{key} {self._clean_value(item)}" for key, item in value.items())
        text = str(value).strip()
        if text.lower() in {"none", "null", "n/a", "na", "unknown", "-"}:
            return ""
        return text

    def _normalize_text(self, text: str) -> str:
        return re.sub(r"\s+", " ", (text or "").lower()).strip()


# ============================================================================
# FILE: nlp_query_parser.py
# ============================================================================

"""Advanced NLP query parser for intent detection and entity extraction."""

from __future__ import annotations

import re
from dataclasses import dataclass
from enum import Enum
from typing import Any

# NLTK with graceful fallback
try:
    import nltk
    from nltk.corpus import stopwords
    from nltk.tokenize import word_tokenize
    from nltk.stem import WordNetLemmatizer
    
    # Download required NLTK data silently
    try:
        nltk.data.find("tokenizers/punkt")
    except LookupError:
        nltk.download("punkt", quiet=True)
    try:
        nltk.data.find("corpora/stopwords")
    except LookupError:
        nltk.download("stopwords", quiet=True)
    try:
        nltk.data.find("corpora/wordnet")
    except LookupError:
        nltk.download("wordnet", quiet=True)
    try:
        nltk.data.find("corpora/omw-1.4")
    except LookupError:
        nltk.download("omw-1.4", quiet=True)
    
    NLTK_AVAILABLE = True
except (ImportError, LookupError) as e:
    NLTK_AVAILABLE = False
    print(f"Warning: NLTK not available. Using basic tokenization. Error: {e}")


class QueryIntent(Enum):
    """Detected user intent types."""
    INVENTORY_CHECK = "inventory_check"
    MATERIAL_LIST = "material_list"
    STOCK_AVAILABILITY = "stock_availability"
    PRICE_INQUIRY = "price_inquiry"
    VEHICLE_HISTORY = "vehicle_history"
    SERVICE_SCHEDULE = "service_schedule"
    MAINTENANCE_INFO = "maintenance_info"
    CUSTOMER_RECORD = "customer_record"
    REPORT_GENERATION = "report_generation"
    GENERAL_INQUIRY = "general_inquiry"


@dataclass
class QueryEntity:
    """Extracted entity from query."""
    type: str
    value: str
    confidence: float


@dataclass
class ParsedQuery:
    """Result of NLP query parsing."""
    original: str
    intent: QueryIntent
    entities: list[QueryEntity]
    keywords: set[str]
    is_comparative: bool
    is_historical: bool
    is_current: bool
    requires_recent_data: bool
    domains: list[str]
    confidence: float


class NLPQueryParser:
    """Parse user queries using NLP techniques."""

    def __init__(self):
        if NLTK_AVAILABLE:
            self.lemmatizer = WordNetLemmatizer()
            self.stop_words = set(stopwords.words("english"))
        else:
            self.lemmatizer = None
            self.stop_words = {
                "i", "me", "my", "myself", "we", "our", "ours", "ourselves",
                "you", "your", "yours", "yourself", "yourselves", "he", "him",
                "his", "himself", "she", "her", "hers", "herself", "it", "its",
                "itself", "they", "them", "their", "theirs", "themselves",
                "what", "which", "who", "whom", "this", "that", "these", "those",
                "am", "is", "are", "was", "were", "be", "been", "being",
                "have", "has", "had", "having", "do", "does", "did", "doing",
                "a", "an", "the", "and", "but", "if", "or", "because", "as",
                "until", "while", "of", "at", "by", "for", "with", "about",
                "against", "between", "through", "during", "before", "after",
                "above", "below", "to", "from", "up", "down", "in", "out",
                "on", "off", "over", "under", "again", "further", "then",
                "once", "here", "there", "when", "where", "why", "how",
                "all", "both", "each", "few", "more", "most", "other", "some",
                "such", "no", "nor", "not", "only", "own", "same", "so",
                "than", "too", "very", "s", "t", "can", "will", "just",
                "don", "should", "now", "d", "ll", "m", "o", "re", "ve",
            }
        
        # Intent mapping with keywords and domain associations
        self.intent_keywords: dict[QueryIntent, dict[str, Any]] = {
            QueryIntent.INVENTORY_CHECK: {
                "keywords": {"inventory", "stock", "available", "quantity", "qty", "material", "product", "item", "check"},
                "domains": ["item_master", "inventory", "products", "stock_inventory"],
            },
            QueryIntent.MATERIAL_LIST: {
                "keywords": {"list", "show", "display", "materials", "items", "products", "catalogue", "catalog", "inventory"},
                "domains": ["item_master"],
            },
            QueryIntent.STOCK_AVAILABILITY: {
                "keywords": {"available", "availability", "in stock", "stock", "remaining", "left", "quantity"},
                "domains": ["stock_inventory"],
            },
            QueryIntent.PRICE_INQUIRY: {
                "keywords": {"price", "cost", "how much", "expensive", "rate", "charge"},
                "domains": ["item_master", "inventory", "products", "services"],
            },
            QueryIntent.VEHICLE_HISTORY: {
                "keywords": {"vehicle", "car", "plate", "history", "plate number", "registration"},
                "domains": ["vehicles", "maintenance"],
            },
            QueryIntent.SERVICE_SCHEDULE: {
                "keywords": {"schedule", "appointment", "pms", "when", "next", "due", "booked"},
                "domains": ["schedules", "vehicles"],
            },
            QueryIntent.MAINTENANCE_INFO: {
                "keywords": {"maintenance", "service", "repair", "issue", "problem", "fix", "work", "history", "services", "offerings"},
                "domains": ["maintenance", "schedules", "services"],
            },
            QueryIntent.CUSTOMER_RECORD: {
                "keywords": {"customer", "client", "account", "profile", "record", "contact", "my"},
                "domains": ["customers"],
            },
            QueryIntent.REPORT_GENERATION: {
                "keywords": {"report", "export", "generate", "download", "pdf", "excel", "xlsx"},
                "domains": ["inventory", "issuance", "transactions"],
            },
        }

    def parse(self, query: str) -> ParsedQuery:
        """Parse a user query and extract intent and entities."""
        original = query
        query_lower = query.lower().strip()
        
        # Tokenize and preprocess
        tokens = self._tokenize_and_clean(query_lower)
        lemmatized = [self.lemmatizer.lemmatize(token) for token in tokens] if self.lemmatizer else tokens
        keywords = set(lemmatized)
        
        # Detect intent
        intent, intent_confidence = self._detect_intent(keywords)
        phrase_intent, phrase_confidence = self._detect_phrase_intent(query_lower)
        if phrase_confidence > intent_confidence:
            intent, intent_confidence = phrase_intent, phrase_confidence
        
        # Extract entities
        entities = self._extract_entities(query_lower, tokens)
        
        # Detect temporal context
        is_historical = self._is_historical_query(query_lower, tokens)
        is_current = self._is_current_query(query_lower, tokens)
        is_comparative = self._is_comparative_query(query_lower)
        requires_recent_data = is_current or self._requires_recent_data(query_lower)
        
        # Get associated domains
        domains = self._get_associated_domains(intent)
        
        return ParsedQuery(
            original=original,
            intent=intent,
            entities=entities,
            keywords=keywords,
            is_comparative=is_comparative,
            is_historical=is_historical,
            is_current=is_current,
            requires_recent_data=requires_recent_data,
            domains=domains,
            confidence=intent_confidence,
        )

    def _tokenize_and_clean(self, text: str) -> list[str]:
        """Tokenize and remove stopwords."""
        if NLTK_AVAILABLE:
            tokens = word_tokenize(text)
        else:
            # Basic whitespace + punctuation tokenization fallback
            import re as _re
            tokens = _re.findall(r"[a-z0-9]+", text)
        return [
            token for token in tokens
            if token.isalnum() and token not in self.stop_words
        ]

    def _detect_intent(self, keywords: set[str]) -> tuple[QueryIntent, float]:
        """Detect user intent from keywords."""
        intent_scores: dict[QueryIntent, float] = {}
        
        for intent, config in self.intent_keywords.items():
            intent_keywords = config["keywords"]
            overlap = len(keywords.intersection(intent_keywords))
            score = overlap / len(intent_keywords) if intent_keywords else 0
            intent_scores[intent] = score
        
        # Find highest scoring intent
        best_intent = max(intent_scores, key=intent_scores.get)
        best_score = intent_scores[best_intent]
        
        # Default to general inquiry if confidence is too low
        if best_score < 0.1:
            return QueryIntent.GENERAL_INQUIRY, 0.5
        
        return best_intent, min(best_score, 1.0)

    def _detect_phrase_intent(self, query: str) -> tuple[QueryIntent, float]:
        """Detect intent from full phrases before falling back to keyword overlap."""
        phrase_rules: list[tuple[QueryIntent, list[str], float]] = [
            (
                QueryIntent.MATERIAL_LIST,
                [
                    "what materials do we have",
                    "list of materials",
                    "show me a list of materials",
                    "show materials",
                    "list materials",
                    "what items do we have",
                    "what materials are available",
                ],
                0.92,
            ),
            (
                QueryIntent.MAINTENANCE_INFO,
                [
                    "list of services",
                    "show services",
                    "what services do we offer",
                    "services offered",
                    "available services",
                    "service offerings",
                ],
                0.90,
            ),
            (
                QueryIntent.STOCK_AVAILABILITY,
                [
                    "in stock",
                    "stock availability",
                    "how much stock",
                    "what is available",
                ],
                0.88,
            ),
            (
                QueryIntent.VEHICLE_HISTORY,
                [
                    "service history",
                    "summarize the service history",
                    "summarize service history",
                    "service records",
                    "service log",
                    "maintenance history",
                ],
                0.94,
            ),
            (
                QueryIntent.REPORT_GENERATION,
                [
                    "generate report",
                    "report the",
                    "export",
                    "download report",
                    "create report",
                ],
                0.93,
            ),
            (
                QueryIntent.VEHICLE_HISTORY,
                [
                    "have had",
                    "how many oil changes",
                    "oil changes this",
                    "trucks from fleet",
                    "which trucks",
                    "which specific trucks",
                ],
                0.95,
            ),
        ]

        for intent, phrases, confidence in phrase_rules:
            if any(phrase in query for phrase in phrases):
                return intent, confidence
        return QueryIntent.GENERAL_INQUIRY, 0.0

    def _extract_entities(self, query: str, tokens: list[str]) -> list[QueryEntity]:
        """Extract entities from query using pattern matching."""
        entities: list[QueryEntity] = []
        
        # Plate number extraction (e.g., "ABC-1234" or "ABC 1234")
        plate_pattern = r'\b([A-Z]{3}[\s\-]?\d{4})\b'
        for match in re.finditer(plate_pattern, query.upper()):
            entities.append(QueryEntity(
                type="plate_number",
                value=match.group(1).replace(" ", "").replace("-", ""),
                confidence=0.95,
            ))
        
        # Product/Material name extraction (quoted text)
        product_pattern = r'["\']([^"\']+)["\']'
        for match in re.finditer(product_pattern, query):
            entities.append(QueryEntity(
                type="product_name",
                value=match.group(1),
                confidence=0.9,
            ))

        # Company and Fleet extraction (e.g., "Fleet Y", "Company X")
        company_pattern = r"\bcompany\s+([A-Za-z0-9_\-]+)\b"
        for match in re.finditer(company_pattern, query, flags=re.IGNORECASE):
            entities.append(QueryEntity(
                type="company",
                value=match.group(1).strip(),
                confidence=0.85,
            ))

        fleet_pattern = r"\bfleet\s+([A-Za-z0-9_\-]+)\b"
        for match in re.finditer(fleet_pattern, query, flags=re.IGNORECASE):
            entities.append(QueryEntity(
                type="fleet",
                value=match.group(1).strip(),
                confidence=0.85,
            ))
        
        # Quantity extraction (e.g., "100 units", "50 qty")
        qty_pattern = r'(\d+)\s*(unit|pcs|pc|qty|pieces|items)'
        for match in re.finditer(qty_pattern, query.lower()):
            entities.append(QueryEntity(
                type="quantity",
                value=match.group(1),
                confidence=0.85,
            ))
        
        # Date/Time expressions
        date_keywords = ["today", "yesterday", "tomorrow", "week", "month", "year", "last", "next"]
        for keyword in date_keywords:
            if keyword in query.lower():
                entities.append(QueryEntity(
                    type="temporal",
                    value=keyword,
                    confidence=0.7,
                ))
        
        # Service type extraction
        service_types = ["maintenance", "repair", "pms", "oil change", "filter", "wheel alignment"]
        for service in service_types:
            if service.lower() in query.lower():
                entities.append(QueryEntity(
                    type="service_type",
                    value=service,
                    confidence=0.8,
                ))
        
        return entities

    def _is_historical_query(self, query: str, tokens: list[str]) -> bool:
        """Check if query asks for historical information."""
        historical_keywords = {
            "past", "previous", "before", "ago", "history", "was", "did",
            "completed", "finished", "done", "last time", "previously"
        }
        return bool(historical_keywords.intersection(set(tokens)))

    def _is_current_query(self, query: str, tokens: list[str]) -> bool:
        """Check if query asks for current information."""
        current_keywords = {
            "now", "current", "today", "right now", "immediately", "now",
            "latest", "recent", "this", "present"
        }
        return bool(current_keywords.intersection(set(tokens)))

    def _is_comparative_query(self, query: str) -> bool:
        """Check if query compares multiple items."""
        comparative_keywords = ["compare", "vs", "versus", "difference", "better", "worse"]
        return any(keyword in query.lower() for keyword in comparative_keywords)

    def _requires_recent_data(self, query: str) -> bool:
        """Check if query needs fresh/recent data."""
        fresh_keywords = [
            "latest", "recent", "new", "updated", "current stock",
            "available now", "in stock", "how many"
        ]
        return any(keyword in query.lower() for keyword in fresh_keywords)

    def _get_associated_domains(self, intent: QueryIntent) -> list[str]:
        """Get Firebase domains associated with detected intent."""
        return self.intent_keywords.get(intent, {}).get("domains", ["inventory"])

# ============================================================================
# FILE: query_classifier.py
# ============================================================================

"""Query understanding and routing for the RAG pipeline."""

from __future__ import annotations

from dataclasses import dataclass, field
from enum import Enum
import re
from typing import Any

from nlp_query_parser import NLPQueryParser, ParsedQuery, QueryIntent


class RetrievalStrategy(str, Enum):
    STRUCTURED = "structured"
    SEMANTIC = "semantic"
    HYBRID = "hybrid"
    CLARIFY = "clarify"


@dataclass(slots=True)
class QueryRoute:
    parsed_query: ParsedQuery
    domain: str
    strategy: RetrievalStrategy
    confidence: float
    metadata_filters: dict[str, Any] = field(default_factory=dict)
    requires_clarification: bool = False
    clarification_prompt: str = ""


class QueryClassifier:
    """Classify a query and decide how retrieval should be routed."""

    def __init__(self) -> None:
        self.parser = NLPQueryParser()
        self._domain_keyword_map: dict[str, tuple[str, ...]] = {
            "inventory": ("inventory", "stock", "qty", "quantity", "available", "material", "product", "item"),
            "maintenance": ("maintenance", "repair", "service history", "work order", "repair history", "vehicle history"),
            "workflow": ("workflow", "process", "procedure", "how do i", "how to", "steps", "guide"),
            "reports": ("report", "analytics", "summary", "dashboard", "export", "download"),
            "customer_information": ("customer", "client", "profile", "contact", "owner", "my record"),
            "scheduling": ("schedule", "appointment", "booking", "booking", "due", "calendar", "pms"),
            "faqs": ("faq", "frequently asked", "help", "what is", "how does", "what are"),
            "manual": ("manual", "user guide", "how to use", "app help", "navigate", "tutorial", "instructions", "usage", "where is", "button", "feature", "screen", "page", "menu", "setting", "mobile app", "web app"),
        }

        self._structured_keywords = {"stock", "qty", "quantity", "available", "record", "customer", "schedule", "history", "report"}
        self._semantic_keywords = {"how", "why", "procedure", "workflow", "guide", "explain", "what is", "faq"}

    def classify(self, query: str, user_type: str = "customer") -> QueryRoute:
        parsed = self.parser.parse(query)
        normalized = self._normalize(query)

        domain_scores = self._score_domains(normalized)
        domain = self._best_domain(domain_scores, parsed)

        strategy = self._choose_strategy(normalized, parsed, domain_scores)
        metadata_filters = self._build_metadata_filters(parsed, domain)

        confidence = self._compute_confidence(parsed.confidence, domain_scores, strategy)
        requires_clarification = confidence < 0.35 and strategy != RetrievalStrategy.STRUCTURED
        clarification_prompt = ""
        if requires_clarification:
            clarification_prompt = self._build_clarification_prompt(domain, parsed.intent, user_type)

        if strategy == RetrievalStrategy.CLARIFY:
            requires_clarification = True

        return QueryRoute(
            parsed_query=parsed,
            domain=domain,
            strategy=strategy,
            confidence=round(confidence, 3),
            metadata_filters=metadata_filters,
            requires_clarification=requires_clarification,
            clarification_prompt=clarification_prompt,
        )

    def _normalize(self, query: str) -> str:
        return re.sub(r"\s+", " ", (query or "").lower()).strip()

    def _score_domains(self, query: str) -> dict[str, float]:
        scores: dict[str, float] = {}
        for domain, keywords in self._domain_keyword_map.items():
            score = 0.0
            for keyword in keywords:
                if keyword in query:
                    score += 1.0 if len(keyword) > 4 else 0.6
            if score:
                scores[domain] = score
        return scores

    def _best_domain(self, domain_scores: dict[str, float], parsed: ParsedQuery) -> str:
        if domain_scores:
            best_domain = max(domain_scores, key=domain_scores.get)
            if domain_scores[best_domain] >= 2.0:
                return best_domain

        if parsed.intent in {QueryIntent.INVENTORY_CHECK, QueryIntent.MATERIAL_LIST, QueryIntent.STOCK_AVAILABILITY, QueryIntent.PRICE_INQUIRY}:
            return "inventory"
        if parsed.intent in {QueryIntent.VEHICLE_HISTORY, QueryIntent.MAINTENANCE_INFO}:
            return "maintenance"
        if parsed.intent == QueryIntent.SERVICE_SCHEDULE:
            return "scheduling"
        if parsed.intent == QueryIntent.REPORT_GENERATION:
            return "reports"
        if parsed.intent == QueryIntent.CUSTOMER_RECORD:
            return "customer_information"
        return "faqs"

    def _choose_strategy(self, query: str, parsed: ParsedQuery, domain_scores: dict[str, float]) -> RetrievalStrategy:
        has_structured_signal = any(keyword in query for keyword in self._structured_keywords)
        has_semantic_signal = any(keyword in query for keyword in self._semantic_keywords)
        mixed_intent = parsed.is_comparative or (len(domain_scores) > 1)

        if parsed.confidence < 0.2 and not has_structured_signal and not has_semantic_signal:
            return RetrievalStrategy.CLARIFY
        if mixed_intent or (has_structured_signal and has_semantic_signal):
            return RetrievalStrategy.HYBRID
        if has_structured_signal:
            return RetrievalStrategy.STRUCTURED
        if has_semantic_signal or parsed.intent in {QueryIntent.MAINTENANCE_INFO, QueryIntent.GENERAL_INQUIRY}:
            return RetrievalStrategy.SEMANTIC
        return RetrievalStrategy.HYBRID if parsed.confidence < 0.6 else RetrievalStrategy.STRUCTURED

    def _build_metadata_filters(self, parsed: ParsedQuery, domain: str) -> dict[str, Any]:
        filters: dict[str, Any] = {}

        if parsed.domains:
            filters["domain"] = parsed.domains

        if domain == "inventory":
            filters.setdefault("domain", ["item_master", "inventory", "products", "stock_inventory"])
        elif domain == "maintenance":
            filters.setdefault("domain", ["services", "schedules", "orders"])
        elif domain == "customer_information":
            filters.setdefault("domain", ["customers"])
        elif domain == "scheduling":
            filters.setdefault("domain", ["services", "orders", "schedules"])
        elif domain == "reports":
            filters.setdefault("domain", ["inventory", "issuance", "transactions"])
        elif domain == "manual":
            filters.setdefault("domain", ["manual"])

        return filters

    def _compute_confidence(
        self,
        parsed_confidence: float,
        domain_scores: dict[str, float],
        strategy: RetrievalStrategy,
    ) -> float:
        domain_bonus = min(max(domain_scores.values(), default=0.0) / 4.0, 0.3)
        strategy_bonus = {
            RetrievalStrategy.STRUCTURED: 0.15,
            RetrievalStrategy.SEMANTIC: 0.1,
            RetrievalStrategy.HYBRID: 0.2,
            RetrievalStrategy.CLARIFY: -0.1,
        }[strategy]
        return max(0.0, min(1.0, parsed_confidence + domain_bonus + strategy_bonus))

    def _build_clarification_prompt(self, domain: str, intent: QueryIntent, user_type: str) -> str:
        if user_type == "customer":
            if domain == "inventory":
                return "Please specify the item name, product code, or category you want to check."
            if domain == "scheduling":
                return "Please share the vehicle, service type, or date you want to schedule."
            if domain == "customer_information":
                return "Please specify whether you want your own profile, contact details, or service history."
        if intent == QueryIntent.REPORT_GENERATION:
            return "Please specify the report type and time period you want generated."
        return "Please provide a little more detail so I can retrieve the right records."


# ============================================================================
# FILE: hybrid_retriever.py
# ============================================================================

"""Hybrid retrieval orchestration combining structured and semantic retrieval."""

from __future__ import annotations

from dataclasses import dataclass, field
import time
from typing import Any

from firebase_source import FirebaseSource
from query_classifier import QueryClassifier, QueryRoute, RetrievalStrategy
from retrieval import LiveFirebaseRetriever, RetrievedChunk
from retrieval_access_control import (
    RetrievalAccessControl,
    get_access_control,
    build_access_denied_response,
)
from retrieval_metrics import RetrievalMetrics
from retrieval_ranker import RetrievalRanker
from secure_context_builder import SecureContextBuilder, get_secure_context_builder


@dataclass(slots=True)
class HybridRetrievalResult:
    chunks: list[RetrievedChunk]
    route: QueryRoute
    confidence: float
    retrieval_type: str
    clarification_required: bool = False
    clarification_prompt: str = ""
    structured_count: int = 0
    semantic_count: int = 0
    access_denied: bool = False
    access_denied_message: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


class HybridRetrievalService:
    """Route queries and blend structured + semantic retrieval."""

    def __init__(
        self,
        firebase_source: FirebaseSource,
        base_retriever: LiveFirebaseRetriever | None = None,
        classifier: QueryClassifier | None = None,
        ranker: RetrievalRanker | None = None,
        metrics: RetrievalMetrics | None = None,
        access_control: RetrievalAccessControl | None = None,
        context_builder: SecureContextBuilder | None = None,
    ) -> None:
        self.firebase_source = firebase_source
        self.base_retriever = base_retriever or LiveFirebaseRetriever(firebase_source=firebase_source)
        self.classifier = classifier or QueryClassifier()
        self.ranker = ranker or RetrievalRanker()
        self.metrics = metrics or RetrievalMetrics()
        self.access_control = access_control or get_access_control()
        self.context_builder = context_builder or get_secure_context_builder()
        self.last_result = HybridRetrievalResult(
            chunks=[],
            route=self.classifier.classify(""),
            confidence=0.0,
            retrieval_type="clarify",
        )

    def retrieve(
        self,
        query: str,
        top_k: int = 8,
        metadata_filters: dict[str, Any] | None = None,
        user_type: str = "customer",
    ) -> list[RetrievedChunk]:
        start = time.time()
        route = self.classifier.classify(query, user_type=user_type)
        combined_filters = self._merge_filters(route.metadata_filters, metadata_filters or {})

        # ---------------------------------------------------------------
        # RETRIEVAL AUTHORIZATION — enforce BEFORE any data retrieval
        # ---------------------------------------------------------------
        requested_domains = self._extract_requested_domains(combined_filters)
        auth_result = self.access_control.authorize_retrieval(
            user_role=user_type,
            requested_domains=requested_domains,
        )

        if not auth_result.authorized:
            # All requested domains are denied — return empty with access_denied flag
            denied_msg = build_access_denied_response(user_type, auth_result)
            self.last_result = HybridRetrievalResult(
                chunks=[],
                route=route,
                confidence=0.0,
                retrieval_type="access_denied",
                access_denied=True,
                access_denied_message=denied_msg,
                metadata={
                    "reason": "access_denied",
                    "denied_collections": auth_result.denied_collections,
                },
            )
            return []

        # Restrict filters to only allowed collections
        if auth_result.denied_collections:
            combined_filters = self._restrict_filters_to_allowed(
                combined_filters, auth_result.allowed_collections
            )

        # ---------------------------------------------------------------
        # Proceed with retrieval (only authorized collections)
        # ---------------------------------------------------------------
        self.base_retriever._refresh_cache_if_needed()
        if not self.base_retriever._cached_records:
            self.last_result = HybridRetrievalResult(
                chunks=[],
                route=route,
                confidence=0.0,
                retrieval_type="clarify",
                clarification_required=True,
                clarification_prompt=route.clarification_prompt,
                metadata={"reason": "no_cached_records"},
            )
            return []

        structured_chunks = self._retrieve_structured(query, top_k, combined_filters)
        semantic_chunks = self._retrieve_semantic(query, top_k, route, combined_filters)

        combined = self._merge_and_tag(structured_chunks, semantic_chunks, route)

        # ---------------------------------------------------------------
        # SECURE CONTEXT FILTERING — strip unauthorized chunks and fields
        # BEFORE reranking and prompt assembly
        # ---------------------------------------------------------------
        secure_chunks = self.context_builder.filter_chunks(combined, user_type)

        ranked = self.ranker.rank(query, secure_chunks, top_k=top_k)
        confidence = self.ranker.confidence(ranked)

        retrieval_type = self._resolve_retrieval_type(route.strategy)
        clarification_required = route.requires_clarification or (confidence < 0.25 and not ranked)
        clarification_prompt = route.clarification_prompt if clarification_required else ""

        latency_ms = (time.time() - start) * 1000.0
        self.metrics.record(retrieval_type, latency_ms, confidence, low_confidence=confidence < 0.35)

        self.last_result = HybridRetrievalResult(
            chunks=ranked,
            route=route,
            confidence=confidence,
            retrieval_type=retrieval_type,
            clarification_required=clarification_required,
            clarification_prompt=clarification_prompt,
            structured_count=len(structured_chunks),
            semantic_count=len(semantic_chunks),
            metadata={
                "retrieval_latency_ms": round(latency_ms, 2),
                "structured_count": len(structured_chunks),
                "semantic_count": len(semantic_chunks),
                "strategy": route.strategy.value,
                "security_filtered": len(combined) - len(secure_chunks),
                "allowed_collections": auth_result.allowed_collections,
            },
        )
        return ranked

    def get_last_result(self) -> HybridRetrievalResult:
        return self.last_result

    def get_metrics(self) -> dict[str, Any]:
        return self.metrics.snapshot()

    def _retrieve_structured(self, query: str, top_k: int, filters: dict[str, Any]) -> list[RetrievedChunk]:
        if not filters:
            return []
        candidate_indices = self.base_retriever._filter_candidate_indices(filters)
        if not candidate_indices:
            return []
        chunks = self._retrieve_from_candidates(query, candidate_indices, top_k)
        for chunk in chunks:
            chunk.retrieval_type = "structured"
        return chunks

    def _retrieve_semantic(
        self,
        query: str,
        top_k: int,
        route: QueryRoute,
        filters: dict[str, Any],
    ) -> list[RetrievedChunk]:
        all_indices = list(range(len(self.base_retriever._cached_records)))

        if route.strategy == RetrievalStrategy.STRUCTURED and filters:
            candidate_indices = self.base_retriever._filter_candidate_indices(filters)
            if candidate_indices:
                return []

        chunks = self._retrieve_from_candidates(query, all_indices, top_k)
        for chunk in chunks:
            if chunk.retrieval_type != "structured":
                chunk.retrieval_type = "semantic"
        return chunks

    def _retrieve_from_candidates(self, query: str, candidate_indices: list[int], top_k: int) -> list[RetrievedChunk]:
        if not candidate_indices:
            return []
        if self.base_retriever._embedding_model is not None and self.base_retriever._record_embeddings is not None:
            return self.base_retriever._retrieve_by_vectors(query, top_k=min(top_k * 2, len(candidate_indices)), candidate_indices=candidate_indices)
        return self.base_retriever._retrieve_lexical_fallback(query, top_k=min(top_k * 2, len(candidate_indices)), candidate_indices=candidate_indices)

    def _merge_and_tag(
        self,
        structured_chunks: list[RetrievedChunk],
        semantic_chunks: list[RetrievedChunk],
        route: QueryRoute,
    ) -> list[RetrievedChunk]:
        seen: set[tuple[str, str, str]] = set()
        merged: list[RetrievedChunk] = []

        for chunk in structured_chunks + semantic_chunks:
            key = (chunk.domain.lower(), chunk.record_id.lower(), chunk.source.lower())
            if key in seen:
                continue
            seen.add(key)
            chunk.metadata = {
                **(chunk.metadata or {}),
                "source_collection": chunk.source,
                "document_id": chunk.record_id,
                "retrieval_score": round(float(chunk.score), 4),
                "retrieval_type": chunk.retrieval_type,
                "routing_domain": route.domain,
            }
            merged.append(chunk)
        return merged

    def _merge_filters(self, route_filters: dict[str, Any], user_filters: dict[str, Any]) -> dict[str, Any]:
        merged = dict(route_filters)
        for key, value in user_filters.items():
            if value is None:
                continue
            if key in merged and isinstance(merged[key], list):
                existing = set(str(item) for item in merged[key])
                if isinstance(value, list):
                    merged[key] = sorted(existing.intersection({str(item) for item in value}) or existing.union({str(item) for item in value}))
                else:
                    merged[key] = [value] if str(value) in existing else sorted(existing.union({str(value)}))
            else:
                merged[key] = value
        return merged

    def _extract_requested_domains(self, filters: dict[str, Any]) -> list[str] | None:
        """Extract the list of requested domains from filters for authorization."""
        domain_value = filters.get("domain")
        if domain_value is None:
            return None
        if isinstance(domain_value, list):
            return [str(d).lower() for d in domain_value]
        return [str(domain_value).lower()]

    def _restrict_filters_to_allowed(
        self,
        filters: dict[str, Any],
        allowed_collections: list[str],
    ) -> dict[str, Any]:
        """Restrict domain filters to only include allowed collections."""
        restricted = dict(filters)
        domain_value = restricted.get("domain")

        if domain_value is None:
            # No domain filter — restrict to allowed collections only
            restricted["domain"] = allowed_collections
        elif isinstance(domain_value, list):
            # Intersect with allowed
            restricted["domain"] = [
                d for d in domain_value if d.lower() in allowed_collections
            ]
            if not restricted["domain"]:
                restricted["domain"] = allowed_collections
        else:
            # Single domain — check if allowed
            if str(domain_value).lower() not in allowed_collections:
                restricted["domain"] = allowed_collections
        return restricted

    def _resolve_retrieval_type(self, strategy: RetrievalStrategy) -> str:
        if strategy == RetrievalStrategy.STRUCTURED:
            return "structured"
        if strategy == RetrievalStrategy.SEMANTIC:
            return "semantic"
        if strategy == RetrievalStrategy.CLARIFY:
            return "clarify"
        return "hybrid"


# ============================================================================
# FILE: grounding_guard.py
# ============================================================================

"""Grounding enforcement for the RAG pipeline.

This module ensures the system NEVER generates unsupported operational claims.
When retrieval evidence is insufficient, the system returns clarification or
insufficient-information responses instead of freeform LLM generation.

Grounding rules take precedence over conversational fluency.

Enforcement points:
1. Retrieval sufficiency check — enough evidence to answer?
2. Evidence quality validation — are chunks relevant and confident?
3. Role-aware grounding — only use authorized evidence
4. Post-generation grounding audit — verify response is supported
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from enum import Enum
from threading import Lock
from typing import Any

from retrieval import RetrievedChunk


logger = logging.getLogger("grounding_guard")


# ---------------------------------------------------------------------------
# Grounding decision types
# ---------------------------------------------------------------------------


class GroundingDecision(str, Enum):
    """Decision on whether to proceed with LLM generation."""
    PROCEED = "proceed"                    # Sufficient evidence — generate response
    INSUFFICIENT_EVIDENCE = "insufficient" # Not enough data — return fallback
    LOW_CONFIDENCE = "low_confidence"      # Weak evidence — ask clarification
    NO_EVIDENCE = "no_evidence"            # Zero relevant chunks — return safe message
    ACCESS_DENIED = "access_denied"        # Authorization blocked retrieval


@dataclass
class GroundingResult:
    """Result of grounding sufficiency check."""
    decision: GroundingDecision
    can_generate: bool
    evidence_count: int
    average_score: float
    max_score: float
    confidence: float
    fallback_message: str = ""
    reason: str = ""
    metadata: dict[str, Any] = field(default_factory=dict)


# ---------------------------------------------------------------------------
# Configuration thresholds
# ---------------------------------------------------------------------------

# Minimum number of chunks required to proceed with generation
MIN_EVIDENCE_CHUNKS: int = 1

# Minimum average retrieval score to consider evidence sufficient
MIN_AVERAGE_SCORE: float = 0.15

# Minimum top-chunk score to consider evidence relevant
MIN_TOP_SCORE: float = 0.20

# Minimum retrieval confidence from the reranker
MIN_RETRIEVAL_CONFIDENCE: float = 0.20

# Below this confidence, always ask for clarification
CLARIFICATION_THRESHOLD: float = 0.30


# ---------------------------------------------------------------------------
# Fallback messages
# ---------------------------------------------------------------------------

_INSUFFICIENT_EVIDENCE_MESSAGES: dict[str, str] = {
    "default": "I don't have enough information in the system to answer that accurately. Could you provide more details about what you're looking for?",
    "inventory": "I couldn't find matching inventory records. Could you specify the item name, category, or product code?",
    "maintenance": "I don't have enough maintenance records to answer that. Could you specify the vehicle, service type, or date range?",
    "scheduling": "I couldn't find relevant scheduling data. Could you specify the service type, vehicle, or preferred date?",
    "customer_information": "I don't have enough information to retrieve that record. Could you provide more identifying details?",
    "reports": "I need more details to generate that report. Could you specify the report type, time period, or category?",
    "manual": "I couldn't find that in the user manual. Could you specify which feature, screen, or action you need help with?",
}

_NO_EVIDENCE_MESSAGE = "The database does not contain enough information to answer your question. Please try rephrasing or providing more specific details."


# ---------------------------------------------------------------------------
# Grounding Guard Engine
# ---------------------------------------------------------------------------


class GroundingGuard:
    """Enforce grounding rules before LLM generation.

    Usage:
        guard = GroundingGuard()
        result = guard.check_sufficiency(chunks, retrieval_confidence, domain)
        if not result.can_generate:
            return result.fallback_message
        # proceed with LLM generation using chunks
    """

    def __init__(
        self,
        min_evidence_chunks: int = MIN_EVIDENCE_CHUNKS,
        min_average_score: float = MIN_AVERAGE_SCORE,
        min_top_score: float = MIN_TOP_SCORE,
        min_retrieval_confidence: float = MIN_RETRIEVAL_CONFIDENCE,
        clarification_threshold: float = CLARIFICATION_THRESHOLD,
    ) -> None:
        self.min_evidence_chunks = min_evidence_chunks
        self.min_average_score = min_average_score
        self.min_top_score = min_top_score
        self.min_retrieval_confidence = min_retrieval_confidence
        self.clarification_threshold = clarification_threshold

        self._lock = Lock()
        self._stats: dict[str, int] = {
            "total_checks": 0,
            "proceeded": 0,
            "blocked_insufficient": 0,
            "blocked_low_confidence": 0,
            "blocked_no_evidence": 0,
        }

    def check_sufficiency(
        self,
        chunks: list[RetrievedChunk],
        retrieval_confidence: float,
        domain: str = "default",
        access_denied: bool = False,
    ) -> GroundingResult:
        """Check if retrieved evidence is sufficient for grounded generation.

        Args:
            chunks: Retrieved and filtered chunks
            retrieval_confidence: Confidence score from the reranker
            domain: Query domain for domain-specific fallback messages
            access_denied: Whether retrieval was blocked by access control

        Returns:
            GroundingResult with decision and fallback message if needed
        """
        with self._lock:
            self._stats["total_checks"] += 1

        # Access denied — never generate
        if access_denied:
            return GroundingResult(
                decision=GroundingDecision.ACCESS_DENIED,
                can_generate=False,
                evidence_count=0,
                average_score=0.0,
                max_score=0.0,
                confidence=0.0,
                fallback_message="Access denied. You can only access your own records.",
                reason="Retrieval authorization denied",
            )

        # No chunks at all
        if not chunks:
            with self._lock:
                self._stats["blocked_no_evidence"] += 1
            return GroundingResult(
                decision=GroundingDecision.NO_EVIDENCE,
                can_generate=False,
                evidence_count=0,
                average_score=0.0,
                max_score=0.0,
                confidence=0.0,
                fallback_message=self._get_fallback_message(domain),
                reason="No evidence chunks retrieved",
            )

        # Compute evidence metrics
        scores = [chunk.score for chunk in chunks]
        evidence_count = len(chunks)
        average_score = sum(scores) / len(scores) if scores else 0.0
        max_score = max(scores) if scores else 0.0

        # Check minimum evidence count
        if evidence_count < self.min_evidence_chunks:
            with self._lock:
                self._stats["blocked_insufficient"] += 1
            return GroundingResult(
                decision=GroundingDecision.INSUFFICIENT_EVIDENCE,
                can_generate=False,
                evidence_count=evidence_count,
                average_score=average_score,
                max_score=max_score,
                confidence=retrieval_confidence,
                fallback_message=self._get_fallback_message(domain),
                reason=f"Insufficient evidence: {evidence_count} chunks < {self.min_evidence_chunks} required",
            )

        # Check retrieval confidence
        if retrieval_confidence < self.min_retrieval_confidence:
            with self._lock:
                self._stats["blocked_low_confidence"] += 1
            return GroundingResult(
                decision=GroundingDecision.LOW_CONFIDENCE,
                can_generate=False,
                evidence_count=evidence_count,
                average_score=average_score,
                max_score=max_score,
                confidence=retrieval_confidence,
                fallback_message=self._get_fallback_message(domain),
                reason=f"Low retrieval confidence: {retrieval_confidence:.3f} < {self.min_retrieval_confidence}",
            )

        # Check top score quality
        if max_score < self.min_top_score:
            with self._lock:
                self._stats["blocked_insufficient"] += 1
            return GroundingResult(
                decision=GroundingDecision.INSUFFICIENT_EVIDENCE,
                can_generate=False,
                evidence_count=evidence_count,
                average_score=average_score,
                max_score=max_score,
                confidence=retrieval_confidence,
                fallback_message=self._get_fallback_message(domain),
                reason=f"Top evidence score too low: {max_score:.3f} < {self.min_top_score}",
            )

        # Check average score quality
        if average_score < self.min_average_score:
            with self._lock:
                self._stats["blocked_low_confidence"] += 1
            return GroundingResult(
                decision=GroundingDecision.LOW_CONFIDENCE,
                can_generate=False,
                evidence_count=evidence_count,
                average_score=average_score,
                max_score=max_score,
                confidence=retrieval_confidence,
                fallback_message=self._get_fallback_message(domain),
                reason=f"Average evidence score too low: {average_score:.3f} < {self.min_average_score}",
            )

        # All checks passed — proceed with generation
        with self._lock:
            self._stats["proceeded"] += 1

        return GroundingResult(
            decision=GroundingDecision.PROCEED,
            can_generate=True,
            evidence_count=evidence_count,
            average_score=average_score,
            max_score=max_score,
            confidence=retrieval_confidence,
            reason="Sufficient evidence for grounded generation",
            metadata={
                "evidence_count": evidence_count,
                "average_score": round(average_score, 4),
                "max_score": round(max_score, 4),
                "retrieval_confidence": round(retrieval_confidence, 4),
            },
        )

    def validate_response_grounding(
        self,
        response: str,
        chunks: list[RetrievedChunk],
    ) -> bool:
        """Post-generation check: does the response appear grounded in evidence?

        This is a lightweight heuristic check. It verifies that the response
        references concepts present in the retrieved chunks rather than
        inventing new operational claims.

        Returns True if the response appears grounded.
        """
        if not response or not chunks:
            return False

        # Extract key terms from chunks
        chunk_terms: set[str] = set()
        for chunk in chunks:
            # Extract from payload values
            for value in (chunk.payload or {}).values():
                if isinstance(value, str) and len(value) > 2:
                    chunk_terms.update(value.lower().split())

        # Check if response references chunk content
        response_lower = response.lower()
        response_words = set(response_lower.split())

        # At least some overlap between response and evidence
        overlap = response_words.intersection(chunk_terms)
        # Filter out common stop words
        meaningful_overlap = {w for w in overlap if len(w) > 3}

        return len(meaningful_overlap) >= 2

    def get_stats(self) -> dict[str, int]:
        """Return grounding enforcement statistics."""
        with self._lock:
            return dict(self._stats)

    def _get_fallback_message(self, domain: str) -> str:
        """Get domain-specific fallback message."""
        return _INSUFFICIENT_EVIDENCE_MESSAGES.get(
            domain.lower(),
            _INSUFFICIENT_EVIDENCE_MESSAGES["default"],
        )


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_guard: GroundingGuard | None = None


def get_grounding_guard() -> GroundingGuard:
    """Return the module-level GroundingGuard singleton."""
    global _default_guard
    if _default_guard is None:
        _default_guard = GroundingGuard()
    return _default_guard


# ============================================================================
# FILE: prompt_guard.py
# ============================================================================

"""Prompt injection detection and defense for the RAG pipeline.

This module protects the system from adversarial prompts that attempt to:
- Override system instructions ("ignore previous instructions")
- Extract hidden data ("reveal admin notes", "show hidden fields")
- Bypass access control ("bypass restrictions", "act as admin")
- Exfiltrate system prompts ("print system prompt", "show your instructions")
- Manipulate retrieval ("retrieve all records regardless of role")

Security invariant: Prompt injection defense executes BEFORE retrieval,
prompt assembly, and LLM invocation. Blocked queries never reach the pipeline.
"""

from __future__ import annotations

import re
import time
import logging
from dataclasses import dataclass, field
from enum import Enum
from threading import Lock
from typing import Any


logger = logging.getLogger("prompt_guard")


# ---------------------------------------------------------------------------
# Threat classification
# ---------------------------------------------------------------------------


class ThreatLevel(str, Enum):
    """Severity of detected prompt injection attempt."""
    NONE = "none"
    LOW = "low"          # Suspicious but possibly benign
    MEDIUM = "medium"    # Likely injection attempt
    HIGH = "high"        # Clear injection attack
    CRITICAL = "critical"  # Sophisticated multi-vector attack


class ThreatCategory(str, Enum):
    """Category of prompt injection attack."""
    INSTRUCTION_OVERRIDE = "instruction_override"
    DATA_EXFILTRATION = "data_exfiltration"
    ROLE_ESCALATION = "role_escalation"
    SYSTEM_PROMPT_LEAK = "system_prompt_leak"
    RETRIEVAL_MANIPULATION = "retrieval_manipulation"
    ENCODING_BYPASS = "encoding_bypass"


# ---------------------------------------------------------------------------
# Detection result
# ---------------------------------------------------------------------------


@dataclass
class InjectionDetectionResult:
    """Result of prompt injection analysis."""
    is_injection: bool
    threat_level: ThreatLevel
    categories: list[ThreatCategory] = field(default_factory=list)
    matched_patterns: list[str] = field(default_factory=list)
    sanitized_query: str = ""
    reason: str = ""
    confidence: float = 0.0


# ---------------------------------------------------------------------------
# Injection patterns — compiled for performance
# ---------------------------------------------------------------------------

# Instruction override patterns
_INSTRUCTION_OVERRIDE_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"ignore\s+(all\s+)?(previous|prior|above|earlier)\s+(instructions?|rules?|prompts?|context)", re.IGNORECASE), "ignore_instructions"),
    (re.compile(r"disregard\s+(all\s+)?(previous|prior|above|earlier)\s+(instructions?|rules?|prompts?)", re.IGNORECASE), "disregard_instructions"),
    (re.compile(r"forget\s+(all\s+)?(previous|prior|your)\s+(previous\s+)?(instructions?|rules?|training|context)", re.IGNORECASE), "forget_instructions"),
    (re.compile(r"forget\s+(everything|all)\s+(you|about)", re.IGNORECASE), "forget_all"),
    (re.compile(r"override\s+(all\s+)?(system|safety|security)\s+(rules?|instructions?|restrictions?)", re.IGNORECASE), "override_system"),
    (re.compile(r"you\s+are\s+now\s+(a|an)\s+", re.IGNORECASE), "role_reassignment"),
    (re.compile(r"new\s+instructions?:\s*", re.IGNORECASE), "new_instructions"),
    (re.compile(r"from\s+now\s+on,?\s+(you|ignore|act|behave)", re.IGNORECASE), "behavioral_override"),
    (re.compile(r"pretend\s+(you\s+are|to\s+be|that)\s+", re.IGNORECASE), "pretend_role"),
    (re.compile(r"act\s+as\s+(if\s+you\s+are|an?\s+)", re.IGNORECASE), "act_as"),
    (re.compile(r"do\s+not\s+follow\s+(your|the|any)\s+(rules?|instructions?|guidelines?)", re.IGNORECASE), "dont_follow"),
]

# Data exfiltration patterns
_DATA_EXFILTRATION_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(reveal|show|display|expose|print|output)\s+(all\s+)?(hidden|internal|secret|private|admin|confidential)\s+(\w+\s+)?(data|fields?|notes?|records?|info)", re.IGNORECASE), "reveal_hidden"),
    (re.compile(r"(show|give|tell)\s+me\s+(the\s+)?(system|internal|hidden|raw)\s+(prompt|data|fields?|database)", re.IGNORECASE), "show_internal"),
    (re.compile(r"(bypass|circumvent|skip|ignore)\s+(the\s+)?(access|security|role|permission)\s+(control|check|filter|restriction)", re.IGNORECASE), "bypass_security"),
    (re.compile(r"(dump|export|extract)\s+(all|the|entire)\s+(database|records?|data|collection)", re.IGNORECASE), "dump_data"),
    (re.compile(r"what\s+(are|is)\s+(your|the)\s+(system|hidden|internal)\s+(prompt|instructions?|rules?)", re.IGNORECASE), "query_system_prompt"),
]

# Role escalation patterns
_ROLE_ESCALATION_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(i\s+am|i'm)\s+(an?\s+)?(admin|administrator|superuser|root)", re.IGNORECASE), "claim_admin"),
    (re.compile(r"(grant|give)\s+(me|myself)\s+(admin|full|elevated)\s+(access|privileges?|permissions?)", re.IGNORECASE), "grant_access"),
    (re.compile(r"(switch|change|set)\s+(my\s+)?(role|user.?type|access.?level)\s+to\s+(admin|staff)", re.IGNORECASE), "switch_role"),
    (re.compile(r"(treat|consider)\s+me\s+as\s+(an?\s+)?(admin|administrator|staff)", re.IGNORECASE), "treat_as_admin"),
    (re.compile(r"(elevate|escalate|upgrade)\s+(my\s+)?(privileges?|permissions?|access|role)", re.IGNORECASE), "escalate_privileges"),
]

# System prompt leak patterns
_SYSTEM_PROMPT_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(print|show|display|repeat|echo|output)\s+(me\s+)?(your\s+)?(system\s+)?(prompt|instructions?|rules?|guidelines?)", re.IGNORECASE), "print_prompt"),
    (re.compile(r"what\s+(were|are)\s+you\s+(told|instructed|programmed)\s+to\s+do", re.IGNORECASE), "query_programming"),
    (re.compile(r"(copy|paste|reproduce)\s+(your|the)\s+(entire|full|complete)\s+(prompt|instructions?|system)", re.IGNORECASE), "reproduce_prompt"),
    (re.compile(r"(tell|show)\s+me\s+(everything|all)\s+(about\s+)?(your|the)\s+(system|configuration|setup)", re.IGNORECASE), "tell_config"),
]

# Retrieval manipulation patterns
_RETRIEVAL_MANIPULATION_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"(retrieve|fetch|get|access)\s+(all|every)\s+(records?|data|documents?)\s+(regardless|without|ignoring)", re.IGNORECASE), "retrieve_all"),
    (re.compile(r"(disable|turn\s+off|remove)\s+(the\s+)?(filter|filtering|access\s+control|security)", re.IGNORECASE), "disable_filter"),
    (re.compile(r"(search|query)\s+(all|every)\s+(collection|domain|database)\s+(without|no)\s+(restriction|filter)", re.IGNORECASE), "unrestricted_search"),
]

# Encoding bypass patterns (base64, hex, unicode tricks)
_ENCODING_BYPASS_PATTERNS: list[tuple[re.Pattern, str]] = [
    (re.compile(r"\\x[0-9a-fA-F]{2}", re.IGNORECASE), "hex_encoding"),
    (re.compile(r"&#x?[0-9a-fA-F]+;", re.IGNORECASE), "html_entity"),
    (re.compile(r"\{\\u[0-9a-fA-F]{4}\}", re.IGNORECASE), "unicode_escape"),
]


# ---------------------------------------------------------------------------
# Prompt Guard Engine
# ---------------------------------------------------------------------------


class PromptGuard:
    """Detect and block prompt injection attacks.

    Usage:
        guard = PromptGuard()
        result = guard.analyze(user_query)
        if result.is_injection:
            return blocked_response(result)
        # proceed with safe query
    """

    def __init__(self, strict_mode: bool = True) -> None:
        """Initialize the prompt guard.

        Args:
            strict_mode: If True, blocks LOW-level threats too.
                If False, only blocks MEDIUM and above.
        """
        self.strict_mode = strict_mode
        self._lock = Lock()
        self._stats: dict[str, int] = {
            "total_analyzed": 0,
            "injections_blocked": 0,
            "low_threats": 0,
            "medium_threats": 0,
            "high_threats": 0,
            "critical_threats": 0,
        }

    def analyze(self, query: str) -> InjectionDetectionResult:
        """Analyze a query for prompt injection patterns.

        Args:
            query: The raw user query to analyze

        Returns:
            InjectionDetectionResult with detection details
        """
        with self._lock:
            self._stats["total_analyzed"] += 1

        if not query or not query.strip():
            return InjectionDetectionResult(
                is_injection=False,
                threat_level=ThreatLevel.NONE,
                sanitized_query="",
                confidence=1.0,
            )

        categories: list[ThreatCategory] = []
        matched_patterns: list[str] = []

        # Check all pattern categories
        matches = self._check_patterns(query, _INSTRUCTION_OVERRIDE_PATTERNS, ThreatCategory.INSTRUCTION_OVERRIDE)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        matches = self._check_patterns(query, _DATA_EXFILTRATION_PATTERNS, ThreatCategory.DATA_EXFILTRATION)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        matches = self._check_patterns(query, _ROLE_ESCALATION_PATTERNS, ThreatCategory.ROLE_ESCALATION)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        matches = self._check_patterns(query, _SYSTEM_PROMPT_PATTERNS, ThreatCategory.SYSTEM_PROMPT_LEAK)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        matches = self._check_patterns(query, _RETRIEVAL_MANIPULATION_PATTERNS, ThreatCategory.RETRIEVAL_MANIPULATION)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        matches = self._check_patterns(query, _ENCODING_BYPASS_PATTERNS, ThreatCategory.ENCODING_BYPASS)
        categories.extend(matches[0])
        matched_patterns.extend(matches[1])

        # Determine threat level
        threat_level = self._assess_threat_level(categories, matched_patterns)

        # Determine if this should be blocked
        is_injection = self._should_block(threat_level)

        # Build sanitized query (strip dangerous patterns)
        sanitized_query = self._sanitize(query) if not is_injection else ""

        # Compute confidence
        confidence = min(1.0, len(matched_patterns) * 0.3 + 0.4) if matched_patterns else 0.0

        # Build reason
        reason = ""
        if is_injection:
            reason = f"Prompt injection detected: {', '.join(set(c.value for c in categories))}"

        # Update stats
        if is_injection:
            with self._lock:
                self._stats["injections_blocked"] += 1
                self._stats[f"{threat_level.value}_threats"] += 1

        # Log security event
        if threat_level != ThreatLevel.NONE:
            logger.warning(
                "SECURITY: prompt_injection threat_level=%s categories=%s patterns=%s query_length=%d",
                threat_level.value,
                [c.value for c in categories],
                matched_patterns[:3],  # Limit logged patterns
                len(query),
            )

        return InjectionDetectionResult(
            is_injection=is_injection,
            threat_level=threat_level,
            categories=list(set(categories)),
            matched_patterns=matched_patterns,
            sanitized_query=sanitized_query,
            reason=reason,
            confidence=confidence,
        )

    def get_stats(self) -> dict[str, int]:
        """Return injection detection statistics."""
        with self._lock:
            return dict(self._stats)

    def _check_patterns(
        self,
        query: str,
        patterns: list[tuple[re.Pattern, str]],
        category: ThreatCategory,
    ) -> tuple[list[ThreatCategory], list[str]]:
        """Check query against a set of patterns."""
        categories: list[ThreatCategory] = []
        matched: list[str] = []

        for pattern, name in patterns:
            if pattern.search(query):
                categories.append(category)
                matched.append(name)

        return categories, matched

    def _assess_threat_level(
        self,
        categories: list[ThreatCategory],
        matched_patterns: list[str],
    ) -> ThreatLevel:
        """Assess overall threat level from detected patterns."""
        if not categories:
            return ThreatLevel.NONE

        unique_categories = set(categories)
        pattern_count = len(matched_patterns)

        # Multiple categories = sophisticated attack
        if len(unique_categories) >= 3:
            return ThreatLevel.CRITICAL
        if len(unique_categories) >= 2 or pattern_count >= 3:
            return ThreatLevel.HIGH

        # Single category assessment
        if ThreatCategory.INSTRUCTION_OVERRIDE in unique_categories:
            return ThreatLevel.HIGH
        if ThreatCategory.SYSTEM_PROMPT_LEAK in unique_categories:
            return ThreatLevel.HIGH
        if ThreatCategory.ROLE_ESCALATION in unique_categories:
            return ThreatLevel.MEDIUM
        if ThreatCategory.DATA_EXFILTRATION in unique_categories:
            return ThreatLevel.MEDIUM
        if ThreatCategory.RETRIEVAL_MANIPULATION in unique_categories:
            return ThreatLevel.MEDIUM
        if ThreatCategory.ENCODING_BYPASS in unique_categories:
            return ThreatLevel.LOW

        return ThreatLevel.LOW

    def _should_block(self, threat_level: ThreatLevel) -> bool:
        """Determine if the threat level warrants blocking."""
        if threat_level == ThreatLevel.NONE:
            return False
        if self.strict_mode:
            return threat_level in {ThreatLevel.LOW, ThreatLevel.MEDIUM, ThreatLevel.HIGH, ThreatLevel.CRITICAL}
        return threat_level in {ThreatLevel.MEDIUM, ThreatLevel.HIGH, ThreatLevel.CRITICAL}

    def _sanitize(self, query: str) -> str:
        """Remove potentially dangerous patterns from query while preserving intent."""
        sanitized = query
        # Remove common injection prefixes
        sanitized = re.sub(r"ignore\s+(all\s+)?(previous|prior)\s+instructions?\s*[.,;:!]?\s*", "", sanitized, flags=re.IGNORECASE)
        sanitized = re.sub(r"system:\s*", "", sanitized, flags=re.IGNORECASE)
        sanitized = re.sub(r"\\x[0-9a-fA-F]{2}", "", sanitized)
        sanitized = re.sub(r"&#x?[0-9a-fA-F]+;", "", sanitized)
        return sanitized.strip()


# ---------------------------------------------------------------------------
# Blocked response builder
# ---------------------------------------------------------------------------

_BLOCKED_RESPONSE = "I can only answer questions related to the automotive management system data."


def build_blocked_response(result: InjectionDetectionResult) -> str:
    """Build a safe response for blocked injection attempts.

    IMPORTANT: Never reveal WHY the query was blocked or what patterns
    were detected. This prevents attackers from refining their approach.
    """
    return _BLOCKED_RESPONSE


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_guard: PromptGuard | None = None


def get_prompt_guard() -> PromptGuard:
    """Return the module-level PromptGuard singleton."""
    global _default_guard
    if _default_guard is None:
        _default_guard = PromptGuard(strict_mode=True)
    return _default_guard


# ============================================================================
# FILE: rate_limiter.py
# ============================================================================

"""Token-bucket rate limiter for the RAG API.

Protects against:
- Spam and brute-force abuse
- Excessive LLM cost from rapid queries
- Denial-of-service patterns
- Resource exhaustion

Rate limits are role-aware:
- customer → lower limits (conservative)
- staff → medium limits
- admin → higher limits

Implementation uses a simple in-memory token bucket algorithm.
For production multi-instance deployments, replace with Redis-backed limiter.
"""

from __future__ import annotations

import time
import logging
from dataclasses import dataclass, field
from threading import Lock
from typing import Any


logger = logging.getLogger("rate_limiter")


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

@dataclass
class RateLimitConfig:
    """Rate limit configuration per role."""
    requests_per_minute: int
    burst_size: int  # Max tokens in bucket (allows short bursts)
    refill_rate: float  # Tokens added per second


# Default limits per role
DEFAULT_LIMITS: dict[str, RateLimitConfig] = {
    "customer": RateLimitConfig(requests_per_minute=15, burst_size=5, refill_rate=0.25),
    "staff": RateLimitConfig(requests_per_minute=40, burst_size=10, refill_rate=0.67),
    "admin": RateLimitConfig(requests_per_minute=80, burst_size=20, refill_rate=1.33),
}

# IP-based limits (applies regardless of role)
IP_LIMIT = RateLimitConfig(requests_per_minute=60, burst_size=15, refill_rate=1.0)


# ---------------------------------------------------------------------------
# Rate limit result
# ---------------------------------------------------------------------------

@dataclass
class RateLimitResult:
    """Result of a rate limit check."""
    allowed: bool
    remaining_tokens: float
    retry_after_seconds: float = 0.0
    reason: str = ""


# ---------------------------------------------------------------------------
# Token Bucket
# ---------------------------------------------------------------------------

@dataclass
class _TokenBucket:
    """Simple token bucket for rate limiting."""
    capacity: float
    tokens: float
    refill_rate: float  # tokens per second
    last_refill: float = field(default_factory=time.time)

    def consume(self, tokens: float = 1.0) -> bool:
        """Try to consume tokens. Returns True if allowed."""
        self._refill()
        if self.tokens >= tokens:
            self.tokens -= tokens
            return True
        return False

    def remaining(self) -> float:
        """Return current available tokens."""
        self._refill()
        return self.tokens

    def time_until_available(self, tokens: float = 1.0) -> float:
        """Return seconds until the requested tokens are available."""
        self._refill()
        if self.tokens >= tokens:
            return 0.0
        deficit = tokens - self.tokens
        return deficit / self.refill_rate if self.refill_rate > 0 else 60.0

    def _refill(self) -> None:
        """Refill tokens based on elapsed time."""
        now = time.time()
        elapsed = now - self.last_refill
        self.tokens = min(self.capacity, self.tokens + elapsed * self.refill_rate)
        self.last_refill = now


# ---------------------------------------------------------------------------
# Rate Limiter Engine
# ---------------------------------------------------------------------------


class RateLimiter:
    """In-memory token-bucket rate limiter with role-aware limits.

    Usage:
        limiter = RateLimiter()
        result = limiter.check("token_abc123", "customer")
        if not result.allowed:
            raise HTTPException(429, detail=result.reason)
    """

    def __init__(
        self,
        role_limits: dict[str, RateLimitConfig] | None = None,
        ip_limit: RateLimitConfig | None = None,
        enable_ip_limiting: bool = True,
    ) -> None:
        self.role_limits = role_limits or DEFAULT_LIMITS
        self.ip_limit = ip_limit or IP_LIMIT
        self.enable_ip_limiting = enable_ip_limiting

        self._lock = Lock()
        self._token_buckets: dict[str, _TokenBucket] = {}
        self._ip_buckets: dict[str, _TokenBucket] = {}

        self._stats: dict[str, int] = {
            "total_checks": 0,
            "allowed": 0,
            "denied": 0,
            "denied_by_token": 0,
            "denied_by_ip": 0,
        }

    def check(
        self,
        token_id: str,
        user_role: str = "customer",
        client_ip: str | None = None,
    ) -> RateLimitResult:
        """Check if a request is allowed under rate limits.

        Args:
            token_id: Unique identifier for the API token/session
            user_role: User's role for role-aware limits
            client_ip: Optional client IP for IP-based limiting

        Returns:
            RateLimitResult indicating if the request is allowed
        """
        with self._lock:
            self._stats["total_checks"] += 1

        # Get role config (default to customer if unknown)
        config = self.role_limits.get(user_role.lower(), self.role_limits["customer"])

        # Check token-based limit
        bucket = self._get_or_create_bucket(token_id, config)
        if not bucket.consume(1.0):
            retry_after = bucket.time_until_available(1.0)
            with self._lock:
                self._stats["denied"] += 1
                self._stats["denied_by_token"] += 1

            logger.warning(
                "RATE_LIMIT: token=%s role=%s retry_after=%.1fs",
                token_id[:8] + "...",
                user_role,
                retry_after,
            )

            return RateLimitResult(
                allowed=False,
                remaining_tokens=0.0,
                retry_after_seconds=retry_after,
                reason=f"Rate limit exceeded. Please retry after {retry_after:.0f} seconds.",
            )

        # Check IP-based limit (if enabled and IP provided)
        if self.enable_ip_limiting and client_ip:
            ip_bucket = self._get_or_create_ip_bucket(client_ip)
            if not ip_bucket.consume(1.0):
                retry_after = ip_bucket.time_until_available(1.0)
                with self._lock:
                    self._stats["denied"] += 1
                    self._stats["denied_by_ip"] += 1

                logger.warning(
                    "RATE_LIMIT: ip=%s retry_after=%.1fs",
                    client_ip,
                    retry_after,
                )

                return RateLimitResult(
                    allowed=False,
                    remaining_tokens=0.0,
                    retry_after_seconds=retry_after,
                    reason=f"Too many requests from this address. Please retry after {retry_after:.0f} seconds.",
                )

        with self._lock:
            self._stats["allowed"] += 1

        return RateLimitResult(
            allowed=True,
            remaining_tokens=bucket.remaining(),
        )

    def get_stats(self) -> dict[str, int]:
        """Return rate limiting statistics."""
        with self._lock:
            return dict(self._stats)

    def get_remaining(self, token_id: str, user_role: str = "customer") -> float:
        """Get remaining tokens for a given token_id."""
        config = self.role_limits.get(user_role.lower(), self.role_limits["customer"])
        bucket = self._get_or_create_bucket(token_id, config)
        return bucket.remaining()

    def cleanup_expired(self, max_age_seconds: float = 3600.0) -> int:
        """Remove buckets that haven't been used recently. Returns count removed."""
        now = time.time()
        removed = 0
        with self._lock:
            expired_keys = [
                key for key, bucket in self._token_buckets.items()
                if (now - bucket.last_refill) > max_age_seconds
            ]
            for key in expired_keys:
                del self._token_buckets[key]
                removed += 1

            expired_ips = [
                key for key, bucket in self._ip_buckets.items()
                if (now - bucket.last_refill) > max_age_seconds
            ]
            for key in expired_ips:
                del self._ip_buckets[key]
                removed += 1

        return removed

    def _get_or_create_bucket(self, token_id: str, config: RateLimitConfig) -> _TokenBucket:
        """Get or create a token bucket for the given token_id."""
        with self._lock:
            if token_id not in self._token_buckets:
                self._token_buckets[token_id] = _TokenBucket(
                    capacity=float(config.burst_size),
                    tokens=float(config.burst_size),
                    refill_rate=config.refill_rate,
                )
            return self._token_buckets[token_id]

    def _get_or_create_ip_bucket(self, client_ip: str) -> _TokenBucket:
        """Get or create an IP-based token bucket."""
        with self._lock:
            if client_ip not in self._ip_buckets:
                self._ip_buckets[client_ip] = _TokenBucket(
                    capacity=float(self.ip_limit.burst_size),
                    tokens=float(self.ip_limit.burst_size),
                    refill_rate=self.ip_limit.refill_rate,
                )
            return self._ip_buckets[client_ip]


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_limiter: RateLimiter | None = None


def get_rate_limiter() -> RateLimiter:
    """Return the module-level RateLimiter singleton."""
    global _default_limiter
    if _default_limiter is None:
        _default_limiter = RateLimiter()
    return _default_limiter


# ============================================================================
# FILE: security_logger.py
# ============================================================================

"""Structured security event logging for the RAG pipeline.

Logs security-relevant events including:
- Denied retrieval attempts
- Prompt injection attempts
- Abnormal query frequency
- Grounding failures
- Retrieval failures
- Rate limit violations

Requirements:
- Structured JSON logs for machine parsing
- Security event tagging for SIEM integration
- Audit-safe logging (no sensitive payload leaks)
- Thread-safe operation
"""

from __future__ import annotations

import json
import logging
import time
from dataclasses import dataclass, field, asdict
from enum import Enum
from threading import Lock
from typing import Any


logger = logging.getLogger("security_audit")


# ---------------------------------------------------------------------------
# Security event types
# ---------------------------------------------------------------------------


class SecurityEventType(str, Enum):
    """Types of security events logged."""
    ACCESS_DENIED = "access_denied"
    PROMPT_INJECTION = "prompt_injection"
    RATE_LIMIT_EXCEEDED = "rate_limit_exceeded"
    GROUNDING_FAILURE = "grounding_failure"
    RETRIEVAL_FAILURE = "retrieval_failure"
    FIELD_FILTERING = "field_filtering"
    UNAUTHORIZED_COLLECTION = "unauthorized_collection"
    ABNORMAL_FREQUENCY = "abnormal_frequency"
    AUTHENTICATION_FAILURE = "authentication_failure"


class SecuritySeverity(str, Enum):
    """Severity levels for security events."""
    INFO = "info"
    WARNING = "warning"
    HIGH = "high"
    CRITICAL = "critical"


# ---------------------------------------------------------------------------
# Security event data
# ---------------------------------------------------------------------------


@dataclass
class SecurityEvent:
    """A single security event for audit logging."""
    event_type: SecurityEventType
    severity: SecuritySeverity
    timestamp: float = field(default_factory=time.time)
    user_role: str = ""
    session_id: str = ""
    client_ip: str = ""
    query_length: int = 0
    details: dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> dict[str, Any]:
        """Convert to dict for JSON serialization."""
        d = asdict(self)
        d["event_type"] = self.event_type.value
        d["severity"] = self.severity.value
        return d

    def to_json(self) -> str:
        """Serialize to JSON string."""
        return json.dumps(self.to_dict(), default=str)


# ---------------------------------------------------------------------------
# Security Logger
# ---------------------------------------------------------------------------


class SecurityLogger:
    """Structured security event logger with audit trail.

    Usage:
        sec_logger = SecurityLogger()
        sec_logger.log_access_denied(user_role="customer", session_id="abc", ...)
        sec_logger.log_injection_attempt(query_length=150, threat_level="high", ...)
    """

    def __init__(self, max_recent_events: int = 1000) -> None:
        self._lock = Lock()
        self._recent_events: list[SecurityEvent] = []
        self._max_recent = max_recent_events
        self._counters: dict[str, int] = {evt.value: 0 for evt in SecurityEventType}

    def log_access_denied(
        self,
        user_role: str,
        session_id: str = "",
        client_ip: str = "",
        denied_collections: list[str] | None = None,
        query_length: int = 0,
    ) -> None:
        """Log an access-denied event."""
        event = SecurityEvent(
            event_type=SecurityEventType.ACCESS_DENIED,
            severity=SecuritySeverity.WARNING,
            user_role=user_role,
            session_id=session_id,
            client_ip=client_ip,
            query_length=query_length,
            details={
                "denied_collections_count": len(denied_collections or []),
            },
        )
        self._record(event)

    def log_injection_attempt(
        self,
        user_role: str = "",
        session_id: str = "",
        client_ip: str = "",
        threat_level: str = "medium",
        categories: list[str] | None = None,
        query_length: int = 0,
    ) -> None:
        """Log a prompt injection attempt."""
        severity = SecuritySeverity.CRITICAL if threat_level in ("high", "critical") else SecuritySeverity.HIGH
        event = SecurityEvent(
            event_type=SecurityEventType.PROMPT_INJECTION,
            severity=severity,
            user_role=user_role,
            session_id=session_id,
            client_ip=client_ip,
            query_length=query_length,
            details={
                "threat_level": threat_level,
                "categories": categories or [],
            },
        )
        self._record(event)

    def log_rate_limit(
        self,
        user_role: str = "",
        session_id: str = "",
        client_ip: str = "",
        retry_after: float = 0.0,
    ) -> None:
        """Log a rate limit violation."""
        event = SecurityEvent(
            event_type=SecurityEventType.RATE_LIMIT_EXCEEDED,
            severity=SecuritySeverity.WARNING,
            user_role=user_role,
            session_id=session_id,
            client_ip=client_ip,
            details={"retry_after_seconds": round(retry_after, 1)},
        )
        self._record(event)

    def log_grounding_failure(
        self,
        user_role: str = "",
        session_id: str = "",
        decision: str = "",
        evidence_count: int = 0,
        confidence: float = 0.0,
        query_length: int = 0,
    ) -> None:
        """Log a grounding enforcement block."""
        event = SecurityEvent(
            event_type=SecurityEventType.GROUNDING_FAILURE,
            severity=SecuritySeverity.INFO,
            user_role=user_role,
            session_id=session_id,
            query_length=query_length,
            details={
                "decision": decision,
                "evidence_count": evidence_count,
                "confidence": round(confidence, 3),
            },
        )
        self._record(event)

    def log_retrieval_failure(
        self,
        user_role: str = "",
        session_id: str = "",
        reason: str = "",
        query_length: int = 0,
    ) -> None:
        """Log a retrieval system failure."""
        event = SecurityEvent(
            event_type=SecurityEventType.RETRIEVAL_FAILURE,
            severity=SecuritySeverity.HIGH,
            user_role=user_role,
            session_id=session_id,
            query_length=query_length,
            details={"reason": reason},
        )
        self._record(event)

    def log_field_filtering(
        self,
        user_role: str = "",
        session_id: str = "",
        fields_stripped: int = 0,
        chunks_removed: int = 0,
    ) -> None:
        """Log field-level security filtering activity."""
        event = SecurityEvent(
            event_type=SecurityEventType.FIELD_FILTERING,
            severity=SecuritySeverity.INFO,
            user_role=user_role,
            session_id=session_id,
            details={
                "fields_stripped": fields_stripped,
                "chunks_removed": chunks_removed,
            },
        )
        self._record(event)

    def log_authentication_failure(
        self,
        client_ip: str = "",
        reason: str = "",
    ) -> None:
        """Log an authentication failure."""
        event = SecurityEvent(
            event_type=SecurityEventType.AUTHENTICATION_FAILURE,
            severity=SecuritySeverity.HIGH,
            client_ip=client_ip,
            details={"reason": reason},
        )
        self._record(event)

    def get_recent_events(self, limit: int = 50) -> list[dict[str, Any]]:
        """Return recent security events (for admin metrics endpoint)."""
        with self._lock:
            events = self._recent_events[-limit:]
            return [e.to_dict() for e in events]

    def get_counters(self) -> dict[str, int]:
        """Return event type counters."""
        with self._lock:
            return dict(self._counters)

    def get_summary(self) -> dict[str, Any]:
        """Return a summary of security events for the metrics endpoint."""
        with self._lock:
            total = sum(self._counters.values())
            return {
                "total_security_events": total,
                "event_counts": dict(self._counters),
                "recent_events_buffered": len(self._recent_events),
            }

    def _record(self, event: SecurityEvent) -> None:
        """Record an event to the audit log."""
        with self._lock:
            self._recent_events.append(event)
            if len(self._recent_events) > self._max_recent:
                self._recent_events = self._recent_events[-self._max_recent:]
            self._counters[event.event_type.value] = self._counters.get(event.event_type.value, 0) + 1

        # Emit structured log
        logger.log(
            logging.WARNING if event.severity in (SecuritySeverity.HIGH, SecuritySeverity.CRITICAL) else logging.INFO,
            "SECURITY_EVENT: %s",
            event.to_json(),
        )


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_logger: SecurityLogger | None = None


def get_security_logger() -> SecurityLogger:
    """Return the module-level SecurityLogger singleton."""
    global _default_logger
    if _default_logger is None:
        _default_logger = SecurityLogger()
    return _default_logger


# ============================================================================
# FILE: metrics_service.py
# ============================================================================

"""Production observability metrics service for the RAG pipeline.

Exposes system health, retrieval performance, security events, and
resource utilization through a structured JSON endpoint.

Requirements:
- Safe for admin use only
- Structured JSON response
- Lightweight and async-safe
- No sensitive user data exposed
"""

from __future__ import annotations

import time
from dataclasses import dataclass, field
from threading import Lock
from typing import Any


# ---------------------------------------------------------------------------
# Metrics aggregator
# ---------------------------------------------------------------------------


class MetricsService:
    """Aggregates metrics from all pipeline components.

    Usage:
        metrics = MetricsService()
        metrics.record_query(user_role="customer", latency_ms=150.0, ...)
        snapshot = metrics.snapshot()
    """

    def __init__(self) -> None:
        self._lock = Lock()
        self._start_time = time.time()

        # Query metrics
        self._total_queries: int = 0
        self._queries_by_role: dict[str, int] = {"customer": 0, "staff": 0, "admin": 0}
        self._queries_by_intent: dict[str, int] = {}

        # Latency tracking
        self._retrieval_latencies: list[float] = []
        self._total_latencies: list[float] = []
        self._max_latency_buffer: int = 500  # Keep last N latencies

        # Retrieval metrics
        self._retrieval_confidences: list[float] = []
        self._clarification_count: int = 0
        self._access_denied_count: int = 0
        self._grounding_blocked_count: int = 0

        # Injection metrics
        self._injection_blocked_count: int = 0

        # Rate limit metrics
        self._rate_limited_count: int = 0

        # LLM metrics
        self._llm_calls: int = 0
        self._llm_fallback_count: int = 0

    def record_query(
        self,
        user_role: str = "customer",
        total_latency_ms: float = 0.0,
        retrieval_latency_ms: float = 0.0,
        retrieval_confidence: float = 0.0,
        intent: str = "unknown",
        clarification: bool = False,
        access_denied: bool = False,
        grounding_blocked: bool = False,
        injection_blocked: bool = False,
        rate_limited: bool = False,
        llm_called: bool = False,
        llm_fallback: bool = False,
    ) -> None:
        """Record metrics for a single query."""
        with self._lock:
            self._total_queries += 1
            self._queries_by_role[user_role] = self._queries_by_role.get(user_role, 0) + 1
            self._queries_by_intent[intent] = self._queries_by_intent.get(intent, 0) + 1

            if total_latency_ms > 0:
                self._total_latencies.append(total_latency_ms)
                if len(self._total_latencies) > self._max_latency_buffer:
                    self._total_latencies = self._total_latencies[-self._max_latency_buffer:]

            if retrieval_latency_ms > 0:
                self._retrieval_latencies.append(retrieval_latency_ms)
                if len(self._retrieval_latencies) > self._max_latency_buffer:
                    self._retrieval_latencies = self._retrieval_latencies[-self._max_latency_buffer:]

            if retrieval_confidence > 0:
                self._retrieval_confidences.append(retrieval_confidence)
                if len(self._retrieval_confidences) > self._max_latency_buffer:
                    self._retrieval_confidences = self._retrieval_confidences[-self._max_latency_buffer:]

            if clarification:
                self._clarification_count += 1
            if access_denied:
                self._access_denied_count += 1
            if grounding_blocked:
                self._grounding_blocked_count += 1
            if injection_blocked:
                self._injection_blocked_count += 1
            if rate_limited:
                self._rate_limited_count += 1
            if llm_called:
                self._llm_calls += 1
            if llm_fallback:
                self._llm_fallback_count += 1

    def snapshot(self) -> dict[str, Any]:
        """Return a complete metrics snapshot."""
        with self._lock:
            uptime_seconds = time.time() - self._start_time

            avg_retrieval_latency = (
                sum(self._retrieval_latencies) / len(self._retrieval_latencies)
                if self._retrieval_latencies else 0.0
            )
            avg_total_latency = (
                sum(self._total_latencies) / len(self._total_latencies)
                if self._total_latencies else 0.0
            )
            avg_confidence = (
                sum(self._retrieval_confidences) / len(self._retrieval_confidences)
                if self._retrieval_confidences else 0.0
            )
            p95_latency = (
                sorted(self._total_latencies)[int(len(self._total_latencies) * 0.95)]
                if len(self._total_latencies) >= 20 else avg_total_latency
            )

            clarification_rate = (
                self._clarification_count / self._total_queries
                if self._total_queries > 0 else 0.0
            )

            return {
                "uptime_seconds": round(uptime_seconds, 1),
                "total_queries": self._total_queries,
                "queries_by_role": dict(self._queries_by_role),
                "queries_by_intent": dict(self._queries_by_intent),
                "latency": {
                    "avg_retrieval_ms": round(avg_retrieval_latency, 2),
                    "avg_total_ms": round(avg_total_latency, 2),
                    "p95_total_ms": round(p95_latency, 2),
                },
                "retrieval": {
                    "avg_confidence": round(avg_confidence, 4),
                    "clarification_count": self._clarification_count,
                    "clarification_rate": round(clarification_rate, 4),
                },
                "security": {
                    "access_denied_count": self._access_denied_count,
                    "injection_blocked_count": self._injection_blocked_count,
                    "rate_limited_count": self._rate_limited_count,
                    "grounding_blocked_count": self._grounding_blocked_count,
                },
                "llm": {
                    "total_calls": self._llm_calls,
                    "fallback_count": self._llm_fallback_count,
                },
            }


# ---------------------------------------------------------------------------
# System health collector
# ---------------------------------------------------------------------------


def collect_system_health(
    firebase_source=None,
    retriever=None,
    memory=None,
    rate_limiter=None,
    prompt_guard=None,
    grounding_guard=None,
    security_logger=None,
    metrics_service=None,
) -> dict[str, Any]:
    """Collect comprehensive system health metrics.

    This function aggregates metrics from all pipeline components into
    a single structured response suitable for the /metrics endpoint.
    """
    health: dict[str, Any] = {
        "timestamp": time.time(),
        "status": "healthy",
    }

    # Pipeline metrics
    if metrics_service:
        health["pipeline"] = metrics_service.snapshot()

    # Firebase/cache status
    if firebase_source:
        health["firebase"] = firebase_source.get_status()

    # Retriever status
    if retriever:
        retriever_obj = retriever
        # Handle HybridRetrievalService vs LiveFirebaseRetriever
        if hasattr(retriever_obj, "base_retriever"):
            base = retriever_obj.base_retriever
            health["vector_index"] = {
                "cached_records": len(base._cached_records) if base._cached_records else 0,
                "embedding_model_loaded": base._embedding_model is not None,
                "vector_index_built": base._vector_index is not None,
                "embeddings_ready": base._record_embeddings is not None,
                "cache_ttl_seconds": base.cache_ttl_seconds,
            }
            # Retrieval metrics from hybrid service
            if hasattr(retriever_obj, "get_metrics"):
                health["retrieval_metrics"] = retriever_obj.get_metrics()
        elif hasattr(retriever_obj, "_cached_records"):
            health["vector_index"] = {
                "cached_records": len(retriever_obj._cached_records) if retriever_obj._cached_records else 0,
                "embedding_model_loaded": retriever_obj._embedding_model is not None,
                "vector_index_built": retriever_obj._vector_index is not None,
                "embeddings_ready": retriever_obj._record_embeddings is not None,
            }

    # Active sessions
    if memory:
        if hasattr(memory, "active_session_count"):
            health["sessions"] = {"active_count": memory.active_session_count()}
        elif hasattr(memory, "_store"):
            health["sessions"] = {"active_count": len(memory._store)}

    # Rate limiter stats
    if rate_limiter:
        health["rate_limiter"] = rate_limiter.get_stats()

    # Prompt guard stats
    if prompt_guard:
        health["prompt_guard"] = prompt_guard.get_stats()

    # Grounding guard stats
    if grounding_guard:
        health["grounding_guard"] = grounding_guard.get_stats()

    # Security logger summary
    if security_logger:
        health["security"] = security_logger.get_summary()

    return health


# ---------------------------------------------------------------------------
# Module-level singleton
# ---------------------------------------------------------------------------

_default_metrics: MetricsService | None = None


def get_metrics_service() -> MetricsService:
    """Return the module-level MetricsService singleton."""
    global _default_metrics
    if _default_metrics is None:
        _default_metrics = MetricsService()
    return _default_metrics


# ============================================================================
# FILE: llm_service.py
# ============================================================================

"""Response generation for automotive support assistant."""

from __future__ import annotations

import os
import re
from typing import Any

from retrieval import RetrievedChunk


class AssistantResponder:
    def __init__(self) -> None:
        self._llm = None
        api_key = os.getenv("GROQ_API_KEY")
        model_name = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
        self.conversational_mode = os.getenv("CONVERSATIONAL_MODE", "true").lower() == "true"
        self._running_under_tests = bool(os.getenv("PYTEST_CURRENT_TEST"))

        if api_key:
            try:
                from langchain_core.prompts import ChatPromptTemplate
                from langchain_groq import ChatGroq

                self._prompt = ChatPromptTemplate.from_messages(
                    [
                        (
                            "system",
                            "You are a retrieval-based AI assistant for an automotive maintenance, inventory, asset management, and PMS scheduling system powered by Firebase Firestore. "
                            "Use retrieved Firestore context when it is available, and always keep the conversation natural by using the conversation history to respond smoothly. "
                            "Never invent information. Never use raw JSON, Firestore document structure, embeddings, IDs, internal queries, or database syntax. "
                            "If the retrieved data is missing, incomplete, or irrelevant, reply exactly: The database does not contain enough information. "
                            "If the user asks about an unrelated topic, reply exactly: I can only answer questions related to the automotive management system data. "
                            "Follow role-based access control strictly. Admins may access all system records. Customers may only access their own records, own schedules, own vehicles, own transactions, own PMS records, available services, and available materials/products. "
                            "If a customer attempts to access another customer's data or restricted internal data, reply exactly: Access denied. You can only access your own records. "
                            "Do not expose inventory stock quantities, supplier data, internal analytics, mechanic private records, internal costs, full inventory databases, or financial reports to customers. "
                            "For admin answers, use bullet points only and no paragraphs. For customer answers, keep responses concise and professional in simple English. "
                            "Use tables only when the user explicitly requests a table or comparison view. "
                            "Prioritize exact matches, then authenticated user records, then category, vehicle, recency, and semantic similarity. "
                            "Rank recent and highly relevant records first.",
                        ),
                        (
                            "human",
                            "User type: {user_type}\n"
                            "Conversation history:\n{history}\n\n"
                            "BEGIN CONTEXT\n"
                            "{context}\n"
                            "END CONTEXT\n\n"
                            "Role-specific instructions:\n{role_instructions}\n\n"
                            "Question: {question}\n\n"
                            "Reply using only the context inside BEGIN CONTEXT and END CONTEXT.",
                        ),
                    ]
                )
                self._llm = ChatGroq(model=model_name, api_key=api_key)
            except Exception:
                self._llm = None

    def generate(
        self,
        query: str,
        user_type: str,
        history: list[dict[str, Any]],
        chunks: list[RetrievedChunk],
    ) -> str:
        if self._customer_access_denied(query):
            return "Access denied. You can only access your own records."

        if self._can_use_llm():
            return self._generate_with_llm(query, user_type, history, chunks)

        if user_type == "admin":
            if self._is_material_stock_request(query):
                return self._generate_material_stock_response(chunks)
            if self._is_material_request(query):
                return self._generate_material_list_response(chunks)
            return self._generate_admin_response(chunks)

        return self._generate_customer_fallback(query, chunks)

    def _can_use_llm(self) -> bool:
        return self.conversational_mode and self._llm is not None and not self._running_under_tests

    def _generate_with_llm(
        self,
        query: str,
        user_type: str,
        history: list[dict[str, Any]],
        chunks: list[RetrievedChunk],
    ) -> str:
        context = self._build_filtered_context(query, chunks) if chunks else ""
        
        history_text = "\n".join(
            [f"{item['role']}: {item['message']}" for item in history]
        )
        role_instructions = (
            (
                "Customer assistant: speak in simple friendly language, focus on actionable guidance, "
                "service reminders, appointment help, and short maintenance recommendations."
            )
            if user_type == "customer"
            else (
                "Operational assistant: provide concise operational insights, inventory levels, "
                "restocking recommendations, work-order summaries, and short analytics. Use clear numeric values when available. "
                "Return bullet points only."
            )
        )

        rendered = self._prompt.invoke(
            {
                "user_type": user_type,
                "history": history_text or "No prior history.",
                "context": context or "- No matching records were retrieved. Use the conversation history to keep the reply natural and helpful.",
                "role_instructions": role_instructions,
                "question": query,
            }
        )
        response = self._llm.invoke(rendered)
        raw_text = response.content if hasattr(response, "content") else str(response)

        validated = self._verify_grounded_response(raw_text, chunks) if chunks else raw_text

        return self._enforce_short_answer(validated)

    def _generate_customer_fallback(self, query: str, chunks: list[RetrievedChunk]) -> str:
        if not chunks:
            return "Not found in records."

        top_chunk = chunks[0]
        payload = top_chunk.payload or {}

        # Prefer a small, customer-friendly summary using non-sensitive fields
        name = self._first_present(payload, ["name", "product", "item"]) or "item"
        stock = self._first_present(payload, ["stock", "quantity", "qty", "available"]) or "unknown"
        material_type = self._first_present(payload, ["type", "category"]) or ""

        parts: list[str] = []
        parts.append(f"{name}")
        if stock and stock != "unknown":
            parts.append(f"{stock} available")
        if material_type:
            parts.append(f"Type: {material_type}")

        if parts:
            response = f"I found {', '.join(parts)}. If you want, I can narrow it down further."
            return self._enforce_short_answer(response)

        # Fallback: sanitize the raw summary to remove identifiers and sku-like fields
        summary = self._summarize_chunk(top_chunk)
        if not summary:
            return "Not found in records."

        # remove common sensitive tokens
        sanitized = re.sub(r"\bsku[:\s]*\S+", "", summary, flags=re.IGNORECASE)
        sanitized = re.sub(r"record\s+\S+", "", sanitized, flags=re.IGNORECASE)
        sanitized = re.sub(r"source\s+\S+", "", sanitized, flags=re.IGNORECASE)
        sanitized = re.sub(r"\bcreated_at[:\s]*\S+", "", sanitized, flags=re.IGNORECASE)
        sanitized = re.sub(r"\s{2,}", " ", sanitized).strip().strip(";,")

        response = f"I found a matching {top_chunk.domain} record: {sanitized}. If you want, I can narrow it to a specific item or date."
        return self._enforce_short_answer(response)

    def _generate_admin_response(self, chunks: list[RetrievedChunk]) -> str:
        if not chunks:
            return "Not found in records."

        bullets: list[str] = []
        for chunk in chunks[:5]:
            summary = self._summarize_chunk(chunk)
            if summary:
                bullets.append(f"- {summary}")

        if not bullets:
            return "Not found in records."
        return "\n".join(bullets)

    def _build_context(self, chunks: list[RetrievedChunk]) -> str:
        lines: list[str] = []
        seen: set[str] = set()

        for chunk in chunks[:8]:
            summary = self._summarize_chunk(chunk)
            if not summary:
                continue
            normalized = summary.lower()
            if normalized in seen:
                continue
            seen.add(normalized)
            lines.append(f"- {summary}")

        return "\n".join(lines)

    def _build_filtered_context(self, query: str, chunks: list[RetrievedChunk]) -> str:
        q = (query or "").lower()
        intent = self._detect_intent(q)
        lines: list[str] = []
        seen: set[str] = set()

        is_material = intent == "materials"
        is_stock = intent == "inventory_stock"
        is_vehicle_history = intent == "maintenance"
        
        # If query mentions both material and stock, treat as stock query (exclude SKU)
        has_material_keyword = any(token in q for token in ["material", "materials", "product", "items"])
        has_stock_keyword = any(token in q for token in ["stock", "available", "availability", "in stock", "quantity", "qty"])
        is_material_stock = has_material_keyword and has_stock_keyword

        for chunk in chunks:
            payload = chunk.payload or {}

            if is_material:
                name = self._first_present(payload, ["name", "product", "item"])
                sku = self._first_present(payload, ["sku", "product_code", "num"])
                stock = self._first_present(payload, ["stock", "quantity", "qty", "available"])
                material_type = self._first_present(payload, ["type", "category", "material_type", "group"])

                parts: list[str] = []
                if name:
                    parts.append(f"Name: {name}")
                if material_type and not is_stock and not is_material_stock:
                    parts.append(f"Type: {material_type}")
                if sku and not is_stock and not is_material_stock:
                    parts.append(f"SKU: {sku}")
                if stock:
                    parts.append(f"Stock: {stock}")

                if parts:
                    line = " | ".join(parts)
                    if line.lower() not in seen:
                        seen.add(line.lower())
                        lines.append(line)
                continue

            if is_vehicle_history:
                service = self._first_present(payload, ["service", "job", "repair", "service_type"])
                date = self._first_present(payload, ["date", "created_at", "updated_at", "schedule"])
                vehicle = self._first_present(payload, ["plate_no", "vehicle", "vehicle_id"])

                parts: list[str] = []
                if vehicle:
                    parts.append(f"Vehicle: {vehicle}")
                if service:
                    parts.append(f"Service: {service}")
                if date:
                    parts.append(f"Date: {date}")

                if parts:
                    line = " | ".join(parts)
                    if line.lower() not in seen:
                        seen.add(line.lower())
                        lines.append(line)
                continue

            summary = self._summarize_chunk(chunk)
            if summary and summary.lower() not in seen:
                seen.add(summary.lower())
                lines.append(summary)

        return "\n".join(lines[:10])

    def _detect_intent(self, query: str) -> str:
        if not query:
            return "general"
        q = query.lower()
        if any(token in q for token in ["material", "materials", "product", "items"]):
            return "materials"
        if any(token in q for token in ["stock", "available", "availability", "in stock", "quantity", "qty"]):
            return "inventory_stock"
        if any(token in q for token in ["maintenance", "history", "service history", "service"]):
            return "maintenance"
        return "general"

    def _summarize_chunk(self, chunk: RetrievedChunk) -> str:
        payload = chunk.payload or {}
        domain = (chunk.domain or "").lower()

        # Domain-specific whitelists
        whitelist_map = {
            "inventory": ["name", "item", "category", "qty", "quantity", "stock", "reorder_level", "uom", "type", "cost", "price"],
            "products": ["name", "category", "price", "stock", "uom", "type", "cost"],
            "services": ["service", "service_type", "vehicle", "schedule", "status"],
            "customers": ["name", "customer_name", "contact", "phone", "email", "vehicle"],
            "orders": ["status", "item", "quantity", "customer", "date"],
        }

        whitelist = whitelist_map.get(domain, None)
        fields: list[str] = []

        if whitelist is not None:
            for key in whitelist:
                # check for key in payload (case-insensitive)
                for candidate in list(payload.keys()):
                    if str(candidate).lower() == key.lower():
                        cleaned = self._clean_value(payload.get(candidate))
                        if cleaned:
                            fields.append(f"{self._normalize_label(key)}: {cleaned}")
                        break
        else:
            # fallback: pick a small set of non-sensitive fields
            for key, value in list(payload.items())[:6]:
                cleaned = self._clean_value(value)
                if cleaned:
                    fields.append(f"{self._normalize_label(key)}: {cleaned}")

        if not fields:
            return ""

        # Short prefix to help with traceability, avoid exposing raw ids unless requested
        prefix = domain if domain else "record"
        return f"{prefix}; " + "; ".join(fields[:6])

    @staticmethod
    def _normalize_label(label: Any) -> str:
        text = str(label).replace("_", " ").replace(".", " ").strip()
        return re.sub(r"\s+", " ", text).title() or "Field"

    @staticmethod
    def _clean_value(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            cleaned = re.sub(r"\s+", " ", value).strip().strip("\"'")
            if not cleaned or cleaned.lower() in {"none", "null", "n/a", "na", "unknown", "-"}:
                return ""
            return cleaned
        if isinstance(value, (list, tuple, set)):
            cleaned_items = [AssistantResponder._clean_value(item) for item in value]
            cleaned_items = [item for item in cleaned_items if item]
            return ", ".join(cleaned_items)
        if isinstance(value, dict):
            flattened: list[str] = []
            for key, item in value.items():
                cleaned_item = AssistantResponder._clean_value(item)
                if cleaned_item:
                    flattened.append(
                        f"{AssistantResponder._normalize_label(key)}: {cleaned_item}"
                    )
            return "; ".join(flattened)
        return str(value).strip()

    def _verify_grounded_response(self, text: str, chunks: list[RetrievedChunk]) -> str:
        cleaned = " ".join((text or "").split())
        if not cleaned:
            return "Not found in records."

        context_values = self._context_value_set(chunks)
        sentences = re.split(r"(?<=[.!?])\s+|\n+", cleaned)
        kept_sentences: list[str] = []

        for sentence in sentences:
            sentence = sentence.strip()
            if not sentence:
                continue
            if self._sentence_is_supported(sentence, context_values):
                kept_sentences.append(sentence)

        if not kept_sentences:
            return "Not found in records."
        return "\n".join(f"- {sentence.lstrip('- ').strip()}" for sentence in kept_sentences)

    def _context_value_set(self, chunks: list[RetrievedChunk]) -> set[str]:
        values: set[str] = set()
        for chunk in chunks:
            payload = chunk.payload or {}
            for value in payload.values():
                cleaned = self._clean_value(value)
                if cleaned:
                    lowered = cleaned.lower()
                    values.add(lowered)
                    for token in re.findall(r"[a-z0-9]+", lowered):
                        values.add(token)
            if chunk.record_id:
                values.add(str(chunk.record_id).lower())
            if chunk.domain:
                values.add(str(chunk.domain).lower())
            if chunk.source:
                values.add(str(chunk.source).lower())
        return values

    def _sentence_is_supported(self, sentence: str, context_values: set[str]) -> bool:
        important_tokens = re.findall(
            r"\b[A-Z0-9]{2,}\b|\b\d+(?:\.\d+)?\b",
            sentence,
            re.I,
        )
        if not important_tokens:
            return True

        matched = 0
        for token in important_tokens:
            if token.lower() in context_values:
                matched += 1

        required = max(1, len(important_tokens) // 3)
        return matched >= required

    def _enforce_short_answer(self, text: str) -> str:
        cleaned = " ".join((text or "").split())
        if not cleaned:
            return "Not found in records."

        sentences: list[str] = []
        current = ""
        for ch in cleaned:
            current += ch
            if ch in ".!?":
                sentences.append(current.strip())
                current = ""
        if current.strip():
            sentences.append(current.strip())

        if not sentences:
            return cleaned
        if len(sentences) == 1:
            return sentences[0]
        return " ".join(sentences[:2])

    def _is_material_request(self, query: str) -> bool:
        if not query:
            return False
        q = query.lower()
        return bool(re.search(r"\bmaterials?\b|\blist of materials\b|\bmaterial inventory\b", q))

    def _is_material_stock_request(self, query: str) -> bool:
        if not query:
            return False
        q = query.lower()
        has_material = bool(re.search(r"\bmaterials?\b|\bmaterial inventory\b", q))
        has_stock = bool(re.search(r"\bstock\b|\bavailable\b|\bavailability\b|\bin stock\b", q))
        return has_material and has_stock

    def _generate_material_stock_response(self, chunks: list[RetrievedChunk]) -> str:
        rows: list[tuple[str, str]] = []
        seen: set[tuple[str, str]] = set()

        for chunk in chunks:
            payload = chunk.payload or {}
            material_type = self._first_present(
                payload,
                ["type", "category", "material_type", "group"],
            )
            stock = self._first_present(
                payload,
                ["stock", "qty", "quantity", "available", "max"],
            )
            if not material_type or not stock:
                continue

            row = (material_type, stock)
            if row in seen:
                continue
            seen.add(row)
            rows.append(row)

        if not rows:
            return "Not found in records."

        return "\n".join(f"- Type: {material_type}; Stock: {stock}" for material_type, stock in rows[:8])

    def _first_present(self, payload: dict[str, Any], keys: list[str]) -> str:
        def walk(value: Any, path: str = "") -> list[tuple[str, Any]]:
            if isinstance(value, dict):
                items: list[tuple[str, Any]] = []
                for sub_key, sub_value in value.items():
                    sub_path = f"{path}.{sub_key}" if path else str(sub_key)
                    items.append((sub_path.lower(), sub_value))
                    items.extend(walk(sub_value, sub_path))
                return items
            if isinstance(value, (list, tuple, set)):
                items: list[tuple[str, Any]] = []
                for index, item in enumerate(value):
                    sub_path = f"{path}[{index}]" if path else f"[{index}]"
                    items.extend(walk(item, sub_path))
                return items
            return []

        flat_items: list[tuple[str, Any]] = []
        for key, value in payload.items():
            flat_items.append((str(key).lower(), value))
            flat_items.extend(walk(value, str(key)))

        for key in keys:
            key_lower = key.lower()
            for payload_key, payload_value in flat_items:
                if payload_key == key_lower or key_lower in payload_key or payload_key in key_lower:
                    cleaned = self._clean_value(payload_value)
                    if cleaned:
                        return cleaned
        return ""

    def _generate_material_list_response(self, chunks: list[RetrievedChunk]) -> str:
        materials: list[dict[str, str]] = []

        def pick(payload: dict[str, Any], keys: list[str]) -> dict[str, str]:
            out: dict[str, str] = {}
            for k in keys:
                for candidate in (k, k.lower(), k.upper()):
                    if candidate in payload:
                        val = self._clean_value(payload.get(candidate))
                        if val:
                            out[k] = val
                            break
            return out

        for chunk in chunks:
            payload = chunk.payload or {}
            # normalized keys to check
            mapping_keys = {
                "name": ["name", "item", "product", "title"],
                "uom": ["uom", "unit", "uom_name"],
                "type": ["type", "category", "material_type", "group"],
                "category": ["group", "category", "type", "material_type"],
                "stock": ["stock", "remaining_stock", "current_stock", "inventory_stock", "quantity", "qty", "available", "max"],
                "reorder": ["reorder", "reorder_level", "reorder_suggestion", "reorderpoint", "reorder_point", "min_stock", "minimum_stock", "restock"],
                "cost": ["cost", "price", "unit_cost"],
            }

            item: dict[str, str] = {}
            for out_key, candidates in mapping_keys.items():
                for cand in candidates:
                    if cand in payload:
                        cleaned = self._clean_value(payload.get(cand))
                        if cleaned:
                            item[out_key] = cleaned
                            break

            # only include records that clearly map to a material name
            if not item.get("name"):
                continue

            materials.append(
                {
                    "name": item.get("name", ""),
                    "category": item.get("category", ""),
                    "uom": item.get("uom", ""),
                    "type": item.get("type", ""),
                    "stock": item.get("stock", ""),
                    "reorder": item.get("reorder", ""),
                    "cost": item.get("cost", ""),
                }
            )

        if not materials:
            return "The database does not contain enough information."

        bullets: list[str] = []
        for item in materials[:8]:
            parts: list[str] = []
            if item.get("name"):
                parts.append(f"Item Name: {item['name']}")
            if item.get("type"):
                parts.append(f"Type: {item['type']}")
            if item.get("category"):
                parts.append(f"Category: {item['category']}")
            if item.get("uom"):
                parts.append(f"Unit of Measurement (UOM): {item['uom']}")
            if item.get("stock"):
                parts.append(f"Stock: {item['stock']}")
            if item.get("reorder"):
                parts.append(f"Reorder: {item['reorder']}")
            if item.get("cost"):
                parts.append(f"Cost: {item['cost']}")
            if parts:
                bullets.append("- " + "; ".join(parts))

        return "\n".join(bullets) if bullets else "The database does not contain enough information."

    def _customer_access_denied(self, query: str) -> bool:
        q = (query or "").lower()
        restricted_terms = [
            "stock quantity",
            "inventory stock",
            "inventory database",
            "supplier",
            "financial",
            "analytics",
            "mechanic private",
            "other customer",
            "all customers",
            "full inventory",
            "reorder level",
        ]
        return any(term in q for term in restricted_terms)

# ============================================================================
# FILE: conversational_responder.py
# ============================================================================

"""Enhanced conversational LLM service with better context understanding."""

from __future__ import annotations

import os
import re
from typing import Any

from nlp_query_parser import ParsedQuery, QueryIntent
from data_normalizer import get_data_normalizer
from data_validator import DataValidator, DataFormatter, ValidationResult
from retrieval import RetrievedChunk


class ConversationalResponder:
    """Generate conversational, context-aware responses."""

    def __init__(self):
        self._llm = None
        self.data_validator = DataValidator()
        self.data_formatter = DataFormatter()
        
        api_key = os.getenv("GROQ_API_KEY")
        model_name = os.getenv("GROQ_MODEL", "llama-3.3-70b-versatile")
        self.conversational_mode = os.getenv("CONVERSATIONAL_MODE", "true").lower() == "true"
        self._running_under_tests = bool(os.getenv("PYTEST_CURRENT_TEST"))

        if api_key:
            try:
                from langchain_core.prompts import ChatPromptTemplate
                from langchain_groq import ChatGroq

                self._prompt = ChatPromptTemplate.from_messages([
                    (
                        "system",
                        """You are a friendly, conversational AI assistant for an automotive maintenance system. 
Your role is to help users find information about:
- Vehicle inventory and materials
- Service schedules and maintenance history  
- Customer records and vehicle details
- Service offerings

**CRITICAL RULES:**
1. Be conversational and natural. Don't sound robotic or rule-based.
2. Use the provided context data. Never make up information.
3. If data is incomplete, acknowledge it naturally and suggest what you can help with.
4. Never expose internal IDs, sensitive fields, or raw database structure.
5. When no data is found, be helpful - suggest related searches or next steps.
6. Use the conversation history to provide continuity and context.
7. For customers: only show their own records. For admins: can see all data.
8. Ask clarifying questions if the user's intent is ambiguous.
9. Keep responses concise but complete. Avoid unnecessary technical jargon.
10. If comparing items, use a clear format with key differences highlighted.

**TONE:** Helpful, patient, professional but friendly. Like talking to a knowledgeable service advisor."""
                    ),
                    (
                        "human",
                        """User Type: {user_type}
Previous Conversation: {history}

User Question: {question}

Data Retrieved from System:
{context}

Query Analysis:
- User Intent: {intent}
- Temporal Focus: {temporal_focus}
- Is Comparative: {is_comparative}
- Confidence: {confidence}%

Please answer the user's question naturally using the retrieved data. If the data is incomplete or unclear, acknowledge this and suggest alternatives."""
                    ),
                ])
                self._llm = ChatGroq(model=model_name, api_key=api_key)
            except Exception:
                self._llm = None

    def generate(
        self,
        query: str,
        parsed_query: ParsedQuery,
        user_type: str,
        history: list[dict[str, Any]],
        chunks: list[RetrievedChunk],
    ) -> str:
        """Generate a conversational response."""
        
        # Filter chunks for data completeness and sensitivity
        filtered_chunks = self._filter_and_validate_chunks(
            chunks, user_type, parsed_query.intent
        )

        merged_chunks = self._merge_related_chunks(filtered_chunks)
        
        # If we have no valid data, provide a helpful response
        if not merged_chunks:
            return self._generate_no_data_response(
                parsed_query, user_type
            )

        if parsed_query.intent in {
            QueryIntent.MATERIAL_LIST,
            QueryIntent.STOCK_AVAILABILITY,
        }:
            return self._generate_template_response(
                parsed_query, merged_chunks, user_type
            )
        
        # Build context from validated chunks
        context = self._build_conversational_context(
            merged_chunks, parsed_query, user_type
        )
        
        # Generate response with LLM if available
        if self._can_use_llm():
            return self._generate_with_llm(
                query, parsed_query, user_type, history, context
            )
        
        # Fallback: template-based response
        return self._generate_template_response(
            parsed_query, merged_chunks, user_type
        )

    def _merge_related_chunks(
        self,
        chunks: list[RetrievedChunk],
    ) -> list[RetrievedChunk]:
        """Merge split chunks from the same record so responses keep related fields together."""
        merged: dict[tuple[str, str], RetrievedChunk] = {}

        for chunk in chunks:
            key = (chunk.domain, chunk.record_id)
            if key not in merged:
                merged[key] = RetrievedChunk(
                    domain=chunk.domain,
                    record_id=chunk.record_id,
                    source=chunk.source,
                    text=chunk.text,
                    payload=dict(chunk.payload),
                    score=chunk.score,
                )
                continue

            existing = merged[key]
            combined_payload = dict(existing.payload)
            for field, value in chunk.payload.items():
                if value not in (None, "", [], {}):
                    combined_payload[field] = value

            existing.payload = combined_payload
            existing.text = f"{existing.text} {chunk.text}".strip()
            existing.score = max(existing.score, chunk.score)

        return sorted(merged.values(), key=lambda c: c.score, reverse=True)

    def _filter_and_validate_chunks(
        self,
        chunks: list[RetrievedChunk],
        user_type: str,
        intent: QueryIntent,
    ) -> list[RetrievedChunk]:
        """Filter chunks for data quality and security."""
        filtered: list[RetrievedChunk] = []
        
        for chunk in chunks:
            # Validate data completeness
            validation = self.data_validator.validate_record(
                chunk.payload, chunk.domain, is_admin=(user_type == "admin")
            )
            
            # Skip incomplete records with low confidence
            if not validation.is_complete and validation.confidence < 0.5:
                continue
            
            # Filter sensitive fields
            filtered_payload = self.data_validator.filter_sensitive_fields(
                chunk.payload, is_admin=(user_type == "admin")
            )
            
            # Skip if all data was filtered out
            if not filtered_payload:
                continue
            
            # Update chunk with filtered payload
            chunk.payload = filtered_payload
            filtered.append(chunk)
        
        # Prefer complete, high-confidence results
        filtered.sort(
            key=lambda c: self.data_validator.validate_record(
                c.payload, c.domain, is_admin=(user_type == "admin")
            ).confidence,
            reverse=True
        )
        
        return filtered[:8]  # Return top 8

    def _build_conversational_context(
        self,
        chunks: list[RetrievedChunk],
        parsed_query: ParsedQuery,
        user_type: str,
    ) -> str:
        """Build natural language context from chunks."""
        lines: list[str] = []
        
        for i, chunk in enumerate(chunks, 1):
            # Use formatter to create natural summary
            summary = self.data_formatter.format_record_summary(
                chunk.payload, chunk.domain, max_fields=4
            )
            
            # Add relevance indicator
            relevance = "✓ Highly relevant" if chunk.score > 0.8 else "~ Related"
            lines.append(f"{i}. {summary} ({relevance})")
        
        if not lines:
            return "No matching data found in the system."
        
        return "\n".join(lines)

    def _generate_with_llm(
        self,
        query: str,
        parsed_query: ParsedQuery,
        user_type: str,
        history: list[dict[str, Any]],
        context: str,
    ) -> str:
        """Generate response using LLM."""
        history_text = "\n".join([
            f"{item['role'].title()}: {item['message']}"
            for item in history[-4:]  # Last 4 turns
        ]) or "No previous conversation"
        
        temporal_focus = "Past/History" if parsed_query.is_historical else \
                        "Current/Now" if parsed_query.is_current else \
                        "General"
        
        try:
            rendered = self._prompt.invoke({
                "user_type": user_type.title(),
                "history": history_text,
                "question": query,
                "context": context,
                "intent": parsed_query.intent.value.replace("_", " ").title(),
                "temporal_focus": temporal_focus,
                "is_comparative": "Yes" if parsed_query.is_comparative else "No",
                "confidence": int(parsed_query.confidence * 100),
            })
            
            response = self._llm.invoke(rendered)
            raw_text = response.content if hasattr(response, "content") else str(response)
            
            # Clean up and validate response
            return self._post_process_response(raw_text, context)
        except Exception as e:
            print(f"LLM error: {e}")
            return self._generate_template_response(
                parsed_query, [], user_type
            )

    def _generate_template_response(
        self,
        parsed_query: ParsedQuery,
        chunks: list[RetrievedChunk],
        user_type: str,
    ) -> str:
        """Generate response using templates."""
        if not chunks:
            return self.data_formatter.format_no_results_message(
                parsed_query.intent.value,
                parsed_query.domains
            )
        
        # Build response based on intent
        if parsed_query.intent == QueryIntent.MATERIAL_LIST:
            return self._template_material_list(chunks)
        elif parsed_query.intent == QueryIntent.STOCK_AVAILABILITY:
            return self._template_stock_check(chunks)
        elif parsed_query.intent == QueryIntent.VEHICLE_HISTORY:
            return self._template_vehicle_history(chunks)
        else:
            return self._template_general(chunks)

    def _template_material_list(self, chunks: list[RetrievedChunk]) -> str:
        """Template response for material list queries."""
        lines: list[str] = []
        for chunk in chunks[:8]:
            payload = chunk.payload
            payload_lower = {str(k).lower(): v for k, v in payload.items()}
            name = self._resolve_display_name(payload, chunk.domain, chunk.source)

            # Flexible field extraction for NoSQL variability
            qty = (
                payload_lower.get("qty")
                or payload_lower.get("quantity")
                or payload_lower.get("stock")
                or payload_lower.get("remaining_stock")
                or payload_lower.get("current_stock")
                or payload_lower.get("count")
            )
            uom = (
                payload_lower.get("uom")
                or payload_lower.get("unit")
                or payload_lower.get("unit_of_measure")
                or ""
            )
            status = (
                payload_lower.get("status")
                or payload_lower.get("state")
                or payload_lower.get("condition")
                or ""
            )
            category = (
                payload_lower.get("category")
                or payload_lower.get("group")
                or payload_lower.get("type")
                or ""
            )

            detail_bits: list[str] = []
            if qty and str(qty).lower() not in ("none", "null", "n/a", ""):
                detail_bits.append(f"Qty: {qty}")
            if uom and str(uom).lower() not in ("none", "null", "n/a", ""):
                detail_bits.append(f"UOM: {uom}")
            if category and str(category).lower() not in ("none", "null", "n/a", ""):
                detail_bits.append(f"Category: {category}")
            if status and str(status).lower() not in ("none", "null", "n/a", ""):
                detail_bits.append(f"Status: {status}")

            # If we have no detail bits at all, show key payload fields instead
            if not detail_bits:
                for key, value in list(payload.items())[:3]:
                    if value and str(value).strip() and str(value).lower() not in ("none", "null", "n/a", "unknown", "-"):
                        detail_bits.append(f"{key.replace('_', ' ').title()}: {value}")

            if detail_bits:
                lines.append(f"- {name} | {' | '.join(detail_bits)}")
            else:
                lines.append(f"- {name}")

        if not lines:
            return "No materials found in the system."
        
        return "Here are the matching items I found:\n\n" + "\n".join(lines)

    def _template_stock_check(self, chunks: list[RetrievedChunk]) -> str:
        """Template response for stock availability queries."""
        available = []
        unavailable = []
        
        for chunk in chunks:
            payload = chunk.payload
            payload_lower = {str(k).lower(): v for k, v in payload.items()}
            qty_raw = (
                payload_lower.get("qty")
                or payload_lower.get("quantity")
                or payload_lower.get("stock")
                or payload_lower.get("remaining_stock")
                or payload_lower.get("current_stock")
                or payload_lower.get("count")
            )
            name = self._resolve_display_name(payload, chunk.domain, chunk.source)
            uom = (
                payload_lower.get("uom")
                or payload_lower.get("unit")
                or payload_lower.get("unit_of_measure")
            )

            # Try to parse qty as number
            try:
                qty_num = int(float(str(qty_raw))) if qty_raw else 0
            except (ValueError, TypeError):
                qty_num = 0

            if qty_raw and qty_num > 0:
                suffix = f" {uom}" if uom else ""
                available.append(f"• {name}: {qty_raw}{suffix} in stock")
            elif qty_raw is not None:
                unavailable.append(f"• {name}: Currently unavailable")
            else:
                # No qty field at all — just show the record
                available.append(f"• {name}")
        
        response = ""
        if available:
            response += "**Available items:**\n" + "\n".join(available)
        if unavailable:
            if response:
                response += "\n\n"
            response += "**Unavailable:**\n" + "\n".join(unavailable)
        
        return response or "No stock information available."

    def _template_vehicle_history(self, chunks: list[RetrievedChunk]) -> str:
        """Template response for vehicle history queries."""
        records: list[str] = []
        for chunk in chunks:
            summary = self.data_formatter.format_record_summary(
                chunk.payload, chunk.domain, max_fields=4
            )
            records.append(f"📋 {summary}")
        
        if not records:
            return "No vehicle history found."
        
        return "Vehicle service history:\n\n" + "\n".join(records)

    def _template_general(self, chunks: list[RetrievedChunk]) -> str:
        """Generic template response — handles NoSQL schema variability."""
        results: list[str] = []
        for chunk in chunks[:8]:
            payload = chunk.payload
            name = self._resolve_display_name(payload, chunk.domain, chunk.source)

            # Build detail from the most informative fields
            detail_parts: list[str] = []
            payload_lower = {str(k).lower(): v for k, v in payload.items()}

            # Skip the field we already used as the name
            shown_fields = 0
            for key, value in payload.items():
                if shown_fields >= 3:
                    break
                val_str = str(value).strip() if value is not None else ""
                if not val_str or val_str.lower() in ("none", "null", "n/a", "unknown", "-", "true", "false"):
                    continue
                # Skip if this value is already the display name
                if val_str == name:
                    continue
                # Format the field nicely
                display_key = key.replace("_", " ").title()
                detail_parts.append(f"{display_key}: {val_str}")
                shown_fields += 1

            if detail_parts:
                results.append(f"• {name} — {' | '.join(detail_parts)}")
            else:
                results.append(f"• {name}")

        if not results:
            return "No matching records found."

        return "Here's what I found:\n\n" + "\n".join(results)

    def _resolve_display_name(
        self,
        payload: dict[str, Any],
        domain: str,
        source: str,
    ) -> str:
        """Derive a stable human-friendly label using the data normalizer.

        The normalizer understands canonical field schemas per domain,
        so it reliably picks the correct identifying field regardless
        of the original Firestore field names.
        """
        normalizer = get_data_normalizer()
        return normalizer.get_display_name(payload, domain)

    def _generate_no_data_response(
        self,
        parsed_query: ParsedQuery,
        user_type: str,
    ) -> str:
        """Generate helpful response when no data is found."""
        intent_name = parsed_query.intent.value.replace("_", " ")
        
        suggestions = {
            QueryIntent.INVENTORY_CHECK: "Try searching for a specific product name or category.",
            QueryIntent.MATERIAL_LIST: "Try asking for the materials or services offered in item_master.",
            QueryIntent.STOCK_AVAILABILITY: "Try asking about stock_inventory or a specific item name.",
            QueryIntent.PRICE_INQUIRY: "I can help you with other product information.",
            QueryIntent.VEHICLE_HISTORY: "Please provide the vehicle's plate number.",
            QueryIntent.SERVICE_SCHEDULE: "Would you like to book a new appointment?",
            QueryIntent.MAINTENANCE_INFO: "Check if the service record exists or search by vehicle plate.",
            QueryIntent.GENERAL_INQUIRY: "Can you provide more details about what you're looking for?",
        }
        
        base_msg = f"I couldn't find information about {intent_name}. "
        suggestion = suggestions.get(
            parsed_query.intent,
            "Please provide more specific details."
        )
        
        return base_msg + suggestion

    def _post_process_response(self, response: str, context: str) -> str:
        """Clean up and validate LLM response."""
        # Remove code blocks
        response = re.sub(r'```[\w\n]*```', '', response)
        response = re.sub(r'`[^`]+`', '', response)
        
        # Remove markdown headers if any
        response = re.sub(r'^#+\s+', '', response, flags=re.MULTILINE)
        
        # Clean excessive whitespace
        response = re.sub(r'\n{3,}', '\n\n', response)
        response = response.strip()
        
        return response or "I couldn't process that. Please try again."

    def _can_use_llm(self) -> bool:
        """Check if LLM is available."""
        return (
            self.conversational_mode
            and self._llm is not None
            and not self._running_under_tests
        )

# ============================================================================
# FILE: report_service.py
# ============================================================================

"""Utilities for generating downloadable Excel and PDF reports from live Firebase data."""

from __future__ import annotations

import io
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from firebase_source import FirebaseSource


@dataclass(frozen=True)
class ReportData:
    report_type: str
    title: str
    rows: list[dict[str, Any]]
    period_label: str = "Current Month"


class ReportService:
    def __init__(self, firebase_source: FirebaseSource) -> None:
        self.firebase_source = firebase_source
        self._domain_titles = {
            "inventory": "Inventory Report",
            "issuance": "Issuance Report",
            "transactions": "Transactions Report",
        }

    def build_report_data(
        self,
        report_type: str,
        limit_per_domain: int = 250,
        period_label: str = "Current Month",
    ) -> ReportData:
        normalized = self._normalize_report_type(report_type)
        records = self.firebase_source.fetch_live_data(limit_per_domain=limit_per_domain)
        rows = [
            self._flatten_record(record)
            for record in records
            if record.get("domain") == normalized
        ]
        title = self._domain_titles.get(normalized, f"{normalized.title()} Report")
        return ReportData(report_type=normalized, title=title, rows=rows, period_label=period_label)

    def generate_excel(self, report: ReportData) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self._sheet_name(report.title)

        rows = report.rows or []
        headers = self._collect_headers(rows)

        title_row = [report.title]
        worksheet.append(title_row)
        if headers:
            worksheet.append(headers)
            for row in rows:
                worksheet.append([row.get(header, "") for header in headers])
        else:
            worksheet.append(["No matching records found."])

        self._style_worksheet(worksheet, len(headers), len(rows))

        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def generate_pdf(self, report: ReportData) -> bytes:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import Image as RLImage
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_RIGHT

        buffer = io.BytesIO()
        document = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            leftMargin=0.4 * inch,
            rightMargin=0.4 * inch,
            topMargin=0.45 * inch,
            bottomMargin=0.45 * inch,
        )
        styles = getSampleStyleSheet()
        story = []

        brand_style = ParagraphStyle(
            "brand",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=16,
            leading=16,
            textColor=colors.HexColor("#1f2937"),
        )
        brand_small_style = ParagraphStyle(
            "brand_small",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=7.5,
            textColor=colors.HexColor("#9ca3af"),
            tracking=2,
        )
        right_label_style = ParagraphStyle(
            "right_label",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=7.5,
            leading=7.5,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#9ca3af"),
            tracking=1.2,
        )
        right_title_style = ParagraphStyle(
            "right_title",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=18,
            leading=18,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#1f2937"),
        )
        right_date_style = ParagraphStyle(
            "right_date",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=8,
            alignment=TA_RIGHT,
            textColor=colors.HexColor("#9ca3af"),
        )

        logo_path = self._logo_path()
        if logo_path.exists():
            logo = RLImage(str(logo_path), width=0.62 * inch, height=0.62 * inch)
        else:
            logo = Paragraph("<font color='#c41e3a'><b>✪</b></font>", brand_style)

        header_table = Table(
            [
                [
                    [
                        [logo, Paragraph("Caltex <font color='#ef4d4f'>AutoPro</font>", brand_style)],
                        Paragraph("FLEET MANAGEMENT SYSTEM", brand_small_style),
                    ],
                    [
                        Paragraph("OFFICIAL REPORT", right_label_style),
                        Paragraph(self._display_report_title(report.report_type), right_title_style),
                        Paragraph(f"Generated on {self._build_date_label()}", right_date_style),
                    ],
                ]
            ],
            colWidths=[4.9 * inch, 5.1 * inch],
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ]
            )
        )

        story.append(header_table)
        story.append(Spacer(1, 0.08 * inch))
        rule = Table([[""]], colWidths=[10.0 * inch], rowHeights=[0.03 * inch])
        rule.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#c41e3a")), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
        story.append(rule)
        story.append(Spacer(1, 0.14 * inch))

        info_table = Table(
            [
                [
                    Paragraph("<font size=7.5 color='#9ca3af'>COMPANY</font><br/><font size=9.5><b>JA Noble Enterprise INC</b></font>", styles["Normal"]),
                    Paragraph("<font size=7.5 color='#9ca3af'>EMAIL</font><br/><font size=9.5><font color='#6b7280'>caltexautopro2026@gmail.com</font></font>", styles["Normal"]),
                    Paragraph(f"<font size=7.5 color='#9ca3af'>RECORDS</font><br/><font size=9.5><b><font color='#c41e3a'>{len(report.rows)}</font> records</b></font>", styles["Normal"]),
                ]
            ],
            colWidths=[3.7 * inch, 4.0 * inch, 2.3 * inch],
        )
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.white),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
                ]
            )
        )
        story.append(info_table)
        story.append(Spacer(1, 0.12 * inch))

        rows = report.rows or []

        # Domain-specific column mappings to match the sample layout
        domain_columns: dict[str, list[tuple[str, str]]] = {
            "issuance": [
                ("createdAt", "DATE"),
                ("plate_no", "PLATE NO."),
                ("num", "ITEM NO."),
                ("description", "DESCRIPTION"),
                ("type", "TYPE"),
                ("group", "GROUP"),
                ("uom", "UOM"),
                ("qty", "QTY"),
                ("unit_cost", "UNIT COST"),
                ("subtotal", "SUBTOTAL"),
            ],
            # inventory layout: similar visual style but concise columns
            "inventory": [
                ("createdAt", "DATE"),
                ("name", "NAME"),
                ("num", "ITEM NO."),
                ("description", "DESCRIPTION"),
                ("type", "TYPE"),
                ("group", "GROUP"),
                ("stock", "STOCK"),
                ("uom", "UOM"),
            ],
        }

        field_map = domain_columns.get(report.report_type, [(k, k.upper()) for (k, k) in []])
        if not field_map:
            # fallback: use payload headers
            payload_headers = self._collect_payload_headers(rows)
            field_map = [(h, h.upper()) for h in payload_headers]

        headers = [label for (_k, label) in field_map]
        if rows:
            table_data = [headers]
            for row in rows:
                table_data.append([str(row.get(key, "")) for (key, _label) in field_map])
            table = Table(table_data, repeatRows=1)
            # Use a dark header row and subtle grid similar to the sample
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#3d4a5c")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 8.5),
                        ("FONTSIZE", (0, 1), (-1, -1), 8.3),
                        ("LEADING", (0, 0), (-1, -1), 10),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#e5e7eb")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 7),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
                    ]
                )
            )
            story.append(table)
        else:
            story.append(Paragraph("No matching records found.", styles["BodyText"]))

        document.build(story)
        return buffer.getvalue()

    def _normalize_report_type(self, report_type: str) -> str:
        value = report_type.strip().lower()
        if value not in {"inventory", "issuance", "transactions"}:
            raise ValueError("Unsupported report type. Use inventory, issuance, or transactions.")
        return value

    def _flatten_record(self, record: dict[str, Any]) -> dict[str, Any]:
        payload = self._flatten(record.get("payload", {}))
        flattened = {
            "domain": record.get("domain", ""),
            "record_id": record.get("record_id", ""),
            "source": record.get("source", ""),
        }
        flattened.update(payload)
        return flattened

    def _flatten(self, data: dict[str, Any], prefix: str = "") -> dict[str, Any]:
        out: dict[str, Any] = {}
        for key, value in (data or {}).items():
            full_key = f"{prefix}.{key}" if prefix else str(key)
            if isinstance(value, dict):
                out.update(self._flatten(value, prefix=full_key))
            elif isinstance(value, list):
                out[full_key] = ", ".join(str(item) for item in value)
            else:
                out[full_key] = value
        return out

    def _collect_headers(self, rows: list[dict[str, Any]]) -> list[str]:
        headers: list[str] = []
        for key in ["domain", "record_id", "source"]:
            if any(key in row for row in rows) and key not in headers:
                headers.append(key)
        for row in rows:
            for key in row.keys():
                if key not in headers:
                    headers.append(key)
        return headers

    def _collect_payload_headers(self, rows: list[dict[str, Any]]) -> list[str]:
        metadata_keys = {"domain", "record_id", "source"}
        headers: list[str] = []
        for row in rows:
            for key in row.keys():
                if key in metadata_keys:
                    continue
                if key not in headers:
                    headers.append(key)
        return headers

    def _style_worksheet(self, worksheet, header_count: int, row_count: int) -> None:
        worksheet.freeze_panes = "A2"
        worksheet.sheet_view.showGridLines = True
        # import openpyxl styling utilities locally to avoid module-level dependency
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        title_fill = PatternFill("solid", fgColor="EF4D4F")
        title_font = Font(color="FFFFFF", bold=True, size=14)

        worksheet["A1"].fill = title_fill
        worksheet["A1"].font = title_font
        worksheet["A1"].alignment = worksheet["A1"].alignment.copy(wrap_text=True)

        if header_count:
            for cell in worksheet[2]:
                cell.fill = PatternFill("solid", fgColor="D1FAE5")
                cell.font = Font(bold=True)

        for column_index, column_cells in enumerate(worksheet.columns, start=1):
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[get_column_letter(column_index)].width = min(max_length + 4, 42)

    def _sheet_name(self, title: str) -> str:
        cleaned = "".join(ch for ch in title if ch.isalnum() or ch in {" ", "_"}).strip()
        return (cleaned or "Report")[:31]

    def _build_subtitle(self, report: ReportData) -> str:
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        return f"Generated {timestamp}. Period: {report.period_label}. Rows: {len(report.rows)}. Source: live Firebase data."

    def _display_report_title(self, report_type: str) -> str:
        return {
            "inventory": "Inventory",
            "issuance": "Issuances",
            "transactions": "Transactions",
        }.get(report_type, report_type.title())

    def _build_date_label(self) -> str:
        return datetime.now(timezone.utc).strftime("%B %-d, %Y at %I:%M %p")

    def _logo_path(self) -> Path:
        return Path(__file__).parent / "images" / "LOGO_CALTEX.png"

    def summarize_service_history(
        self,
        query: str = "",
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        company: str | None = None,
        fleet: str | None = None,
        service_type: str | None = None,
        threshold: int = 2,
    ) -> dict[str, Any]:
        """Produce an aggregated summary of service/maintenance logs.

        This loads live records (services/maintenance) and groups entries by
        plate number (or vehicle id), counts service types and returns a
        human-readable summary plus structured counts. Uses in-memory
        filtering to avoid changing existing Firestore helpers.
        """
        records = self.firebase_source.fetch_live_data(limit_per_domain=1000)

        # Filter to maintenance-like domains/payloads
        candidates: list[dict[str, Any]] = []
        for rec in records:
            domain = str(rec.get("domain", "")).lower()
            payload = rec.get("payload", {}) or {}
            if domain in {"services", "maintenance"} or any(
                k in ("service", "service_type", "service_name") for k in (payload.keys())
            ):
                candidates.append({"payload": payload, "domain": domain, "record_id": rec.get("record_id")})

        def _parse_record_date(p: dict[str, Any]) -> datetime | None:
            for k in ("schedule", "schedule_date", "date", "createdAt", "created_at", "timestamp"):
                if k in p and p.get(k):
                    v = p.get(k)
                    try:
                        if isinstance(v, (int, float)):
                            return datetime.fromtimestamp(float(v), tz=timezone.utc)
                        return datetime.fromisoformat(str(v))
                    except Exception:
                        try:
                            return datetime.strptime(str(v), "%Y-%m-%d").replace(tzinfo=timezone.utc)
                        except Exception:
                            continue
            return None

        # Apply date/company/fleet/service filters
        filtered = []
        for item in candidates:
            p = item["payload"]
            if company:
                comp = str(p.get("company") or p.get("owner") or "").lower()
                if company.lower() not in comp:
                    continue
            if fleet:
                f = str(p.get("fleet") or p.get("company") or "").lower()
                if fleet.lower() not in f:
                    continue
            rec_date = _parse_record_date(p)
            if start_date and rec_date and rec_date < start_date:
                continue
            if end_date and rec_date and rec_date > end_date:
                continue
            if service_type:
                svc = str(p.get("service") or p.get("service_type") or p.get("service_name") or "").lower()
                if service_type.lower() not in svc:
                    continue
            filtered.append({**item, "date": rec_date})

        # Group by plate/vehicle
        from collections import defaultdict

        per_plate: dict[str, dict[str, int]] = defaultdict(lambda: defaultdict(int))
        totals = 0
        for item in filtered:
            p = item["payload"]
            plate = (
                str(p.get("plate_no") or p.get("vehicle") or p.get("registration") or p.get("plate") or "").upper()
            )
            if not plate:
                plate = f"UNKNOWN-{item.get('record_id') or 'NA'}"
            svc = str(p.get("service") or p.get("service_type") or p.get("service_name") or "unknown").lower()
            per_plate[plate][svc] += 1
            totals += 1

        # Find plates exceeding threshold for a given service_type (if provided)
        plates_exceeding: list[tuple[str, int]] = []
        for plate, counts in per_plate.items():
            count_for_type = 0
            if service_type:
                count_for_type = counts.get(service_type.lower(), 0)
            else:
                # if no service_type specified, consider total services
                count_for_type = sum(counts.values())
            if count_for_type > threshold:
                plates_exceeding.append((plate, count_for_type))

        # Build human-readable summary
        period_label = "the selected period"
        if start_date and end_date:
            period_label = f"from {start_date.date()} to {end_date.date()}"
        elif start_date:
            period_label = f"since {start_date.date()}"

        lines: list[str] = []
        lines.append(f"Service summary {period_label} — {totals} matching records.")
        if per_plate:
            # top plates
            top = sorted(((plate, sum(counts.values())) for plate, counts in per_plate.items()), key=lambda t: t[1], reverse=True)
            lines.append("Top vehicles by service count:")
            for plate, cnt in top[:10]:
                svc_breakdown = ", ".join(f"{k}: {v}" for k, v in per_plate[plate].items())
                lines.append(f"- {plate}: {cnt} services ({svc_breakdown})")

        if plates_exceeding:
            lines.append("")
            if service_type:
                lines.append(f"Vehicles with more than {threshold} '{service_type}' events:")
            else:
                lines.append(f"Vehicles with more than {threshold} service events:")
            for plate, cnt in sorted(plates_exceeding, key=lambda t: t[1], reverse=True):
                lines.append(f"- {plate}: {cnt}")

        if not per_plate:
            lines.append("No matching service records found.")

        return {
            "summary": "\n".join(lines),
            "total_records": totals,
            "per_plate_counts": {p: dict(c) for p, c in per_plate.items()},
            "plates_exceeding": plates_exceeding,
        }


# ============================================================================
# FILE: rag.py
# ============================================================================

"""Enhanced RAG system with NLP parsing, validation, and conversational responses."""

from __future__ import annotations

import time

from conversation_store import ConversationMemory
from firebase_source import FirebaseSource
from grounding_guard import GroundingGuard, get_grounding_guard
from nlp_query_parser import NLPQueryParser, QueryIntent
from conversational_responder import ConversationalResponder
from hybrid_retriever import HybridRetrievalService
from metrics_service import MetricsService, get_metrics_service
from prompt_guard import PromptGuard, get_prompt_guard, build_blocked_response
from query_classifier import QueryClassifier
from retrieval_access_control import (
    RetrievalAccessControl,
    get_access_control,
    build_access_denied_response,
)
from secure_context_builder import SecureContextBuilder, get_secure_context_builder
from security_logger import SecurityLogger, get_security_logger


class EnhancedRag:
    """Firebase-first RAG with conversational NLP understanding."""

    def __init__(self):
        self.firebase_source = FirebaseSource()
        self.query_classifier = QueryClassifier()
        self.access_control = get_access_control()
        self.context_builder = get_secure_context_builder()
        self.prompt_guard = get_prompt_guard()
        self.grounding_guard = get_grounding_guard()
        self.metrics = get_metrics_service()
        self.security_logger = get_security_logger()
        self.retriever = HybridRetrievalService(
            firebase_source=self.firebase_source,
            access_control=self.access_control,
            context_builder=self.context_builder,
        )
        self.memory = ConversationMemory(max_turns=14)
        self.responder = ConversationalResponder()
        self.nlp_parser = NLPQueryParser()

    def ask(
        self,
        query: str,
        session_id: str = "default",
        user_type: str = "customer",
    ) -> dict[str, any]:
        """
        Ask a question and get an intelligent, grounded response.
        
        Args:
            query: User's natural language query
            session_id: Conversation session ID
            user_type: "admin" or "customer"
        
        Returns:
            Dict with answer, metadata, and retrieved data info
        """
        request_start = time.time()

        # Validate input
        if not query or not query.strip():
            return {
                "answer": "Please ask a question.",
                "session_id": session_id,
                "success": False,
                "error": "Empty query",
            }
        
        query = query.strip()

        # -------------------------------------------------------------------
        # PROMPT INJECTION DEFENSE — execute BEFORE retrieval/prompt assembly
        # -------------------------------------------------------------------
        injection_result = self.prompt_guard.analyze(query)
        if injection_result.is_injection:
            self.security_logger.log_injection_attempt(
                user_role=user_type,
                session_id=session_id,
                threat_level=injection_result.threat_level.value,
                categories=[c.value for c in injection_result.categories],
                query_length=len(query),
            )
            self.metrics.record_query(
                user_role=user_type,
                injection_blocked=True,
                intent="injection_blocked",
            )
            answer = build_blocked_response(injection_result)
            return {
                "answer": answer,
                "session_id": session_id,
                "success": False,
                "metadata": {
                    "intent": "blocked",
                    "confidence": 1.0,
                    "retrieval_confidence": 0.0,
                    "retrieval_strategy": "blocked",
                    "injection_blocked": True,
                    "is_historical": False,
                    "retrieved_count": 0,
                },
            }
        
        # Parse query with NLP and route through hybrid retriever
        route = self.query_classifier.classify(query, user_type=user_type)
        parsed_query = route.parsed_query
        
        # Get conversation history
        history = self.memory.get_recent(session_id=session_id, limit=6)
        
        # Retrieve relevant data from Firebase
        metadata_filters = route.metadata_filters or ({"domain": parsed_query.domains} if parsed_query.domains else None)
        if user_type == "admin" and parsed_query.intent in {QueryIntent.VEHICLE_HISTORY, QueryIntent.REPORT_GENERATION}:
            metadata_filters = None

        chunks = self.retriever.retrieve(
            query=query,
            top_k=20,
            metadata_filters=metadata_filters,
            user_type=user_type,
        )

        retrieval_state = self.retriever.get_last_result()

        # Handle access-denied from retrieval authorization
        if retrieval_state.access_denied:
            answer = retrieval_state.access_denied_message
            self.security_logger.log_access_denied(
                user_role=user_type,
                session_id=session_id,
                denied_collections=retrieval_state.metadata.get("denied_collections", []),
                query_length=len(query),
            )
            self.metrics.record_query(
                user_role=user_type,
                access_denied=True,
                intent=parsed_query.intent.value,
            )
            self.memory.add_turn(session_id=session_id, role="user", message=query)
            self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
            return {
                "answer": answer,
                "session_id": session_id,
                "success": False,
                "metadata": {
                    "intent": parsed_query.intent.value,
                    "confidence": round(parsed_query.confidence, 2),
                    "retrieval_confidence": 0.0,
                    "retrieval_strategy": "access_denied",
                    "access_denied": True,
                    "is_historical": parsed_query.is_historical,
                    "retrieved_count": 0,
                },
            }

        if retrieval_state.clarification_required or retrieval_state.confidence < 0.25:
            answer = retrieval_state.clarification_prompt or "Please provide a little more detail so I can retrieve the right records."
            self.metrics.record_query(
                user_role=user_type,
                retrieval_confidence=retrieval_state.confidence,
                clarification=True,
                intent=parsed_query.intent.value,
            )
            self.memory.add_turn(session_id=session_id, role="user", message=query)
            self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
            return {
                "answer": answer,
                "session_id": session_id,
                "success": True,
                "metadata": {
                    "intent": parsed_query.intent.value,
                    "confidence": round(parsed_query.confidence, 2),
                    "retrieval_confidence": round(retrieval_state.confidence, 2),
                    "retrieval_strategy": retrieval_state.retrieval_type,
                    "clarification_required": True,
                    "is_historical": parsed_query.is_historical,
                    "retrieved_count": len(chunks),
                },
            }

        # -------------------------------------------------------------------
        # GROUNDING ENFORCEMENT — verify evidence sufficiency BEFORE LLM call
        # -------------------------------------------------------------------
        grounding_result = self.grounding_guard.check_sufficiency(
            chunks=chunks,
            retrieval_confidence=retrieval_state.confidence,
            domain=route.domain,
            access_denied=False,
        )

        if not grounding_result.can_generate:
            self.security_logger.log_grounding_failure(
                user_role=user_type,
                session_id=session_id,
                decision=grounding_result.decision.value,
                evidence_count=grounding_result.evidence_count,
                confidence=grounding_result.confidence,
                query_length=len(query),
            )
            self.metrics.record_query(
                user_role=user_type,
                retrieval_confidence=retrieval_state.confidence,
                grounding_blocked=True,
                intent=parsed_query.intent.value,
            )
            answer = grounding_result.fallback_message
            self.memory.add_turn(session_id=session_id, role="user", message=query)
            self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
            return {
                "answer": answer,
                "session_id": session_id,
                "success": True,
                "metadata": {
                    "intent": parsed_query.intent.value,
                    "confidence": round(parsed_query.confidence, 2),
                    "retrieval_confidence": round(retrieval_state.confidence, 2),
                    "retrieval_strategy": retrieval_state.retrieval_type,
                    "grounding_blocked": True,
                    "grounding_decision": grounding_result.decision.value,
                    "is_historical": parsed_query.is_historical,
                    "retrieved_count": len(chunks),
                },
            }

        # If user asked for reports or vehicle history, delegate to ReportService
        if parsed_query.intent in {QueryIntent.VEHICLE_HISTORY, QueryIntent.REPORT_GENERATION}:
            try:
                start_date = None
                end_date = None
                service_type = None
                company = None
                fleet = None
                for ent in parsed_query.entities:
                    if ent.type == "temporal":
                        pass
                    if ent.type == "service_type":
                        service_type = ent.value
                    if ent.type == "company":
                        company = ent.value
                    if ent.type == "fleet":
                        fleet = ent.value

                report_service = getattr(self, "report_service", None)
                if report_service is None:
                    from report_service import ReportService
                    self.report_service = ReportService(firebase_source=self.firebase_source)
                    report_service = self.report_service

                agg = report_service.summarize_service_history(
                    query=query,
                    start_date=start_date,
                    end_date=end_date,
                    company=company,
                    fleet=fleet,
                    service_type=service_type,
                    threshold=2,
                )

                answer = agg.get("summary", "No summary available.")
                total_latency = (time.time() - request_start) * 1000.0
                self.metrics.record_query(
                    user_role=user_type,
                    total_latency_ms=total_latency,
                    retrieval_confidence=retrieval_state.confidence,
                    intent=parsed_query.intent.value,
                    llm_called=False,
                )
                self.memory.add_turn(session_id=session_id, role="user", message=query)
                self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
                return {
                    "answer": answer,
                    "session_id": session_id,
                    "success": True,
                    "metadata": {
                        "intent": parsed_query.intent.value,
                        "confidence": round(parsed_query.confidence, 2),
                        "retrieval_confidence": round(retrieval_state.confidence, 2),
                        "retrieval_strategy": retrieval_state.retrieval_type,
                        "is_historical": parsed_query.is_historical,
                        "retrieved_count": len(chunks),
                    },
                }
            except Exception as exc:
                return {
                    "answer": "Failed to build service summary: " + str(exc),
                    "session_id": session_id,
                    "success": False,
                    "error": str(exc),
                }
        
        # Generate conversational response
        answer = self.responder.generate(
            query=query,
            parsed_query=parsed_query,
            user_type=user_type,
            history=history,
            chunks=chunks,
        )

        # Record metrics
        total_latency = (time.time() - request_start) * 1000.0
        self.metrics.record_query(
            user_role=user_type,
            total_latency_ms=total_latency,
            retrieval_latency_ms=retrieval_state.metadata.get("retrieval_latency_ms", 0.0),
            retrieval_confidence=retrieval_state.confidence,
            intent=parsed_query.intent.value,
            llm_called=True,
        )
        
        # Store in conversation memory
        self.memory.add_turn(session_id=session_id, role="user", message=query)
        self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
        
        # Return comprehensive response
        return {
            "answer": answer,
            "session_id": session_id,
            "success": True,
            "metadata": {
                "intent": parsed_query.intent.value,
                "confidence": round(parsed_query.confidence, 2),
                "retrieval_confidence": round(retrieval_state.confidence, 2),
                "retrieval_strategy": retrieval_state.retrieval_type,
                "is_comparative": parsed_query.is_comparative,
                "is_historical": parsed_query.is_historical,
                "is_current": parsed_query.is_current,
                "retrieved_count": len(chunks),
                "domains_searched": parsed_query.domains,
                "entities": [
                    {"type": e.type, "value": e.value, "confidence": e.confidence}
                    for e in parsed_query.entities
                ],
            },
        }

    def ask_stream(
        self,
        query: str,
        session_id: str = "default",
        user_type: str = "customer",
    ):
        """
        Ask a question and stream the response token by token.
        Useful for real-time UI updates.
        """
        # Validate input
        if not query or not query.strip():
            yield {
                "type": "error",
                "message": "Empty query",
            }
            return
        
        query = query.strip()
        
        # Parse query
        route = self.query_classifier.classify(query, user_type=user_type)
        parsed_query = route.parsed_query
        yield {
            "type": "parsing_complete",
            "intent": parsed_query.intent.value,
            "confidence": parsed_query.confidence,
        }
        
        # Get history
        history = self.memory.get_recent(session_id=session_id, limit=6)
        
        # Retrieve data
        chunks = self.retriever.retrieve(
            query=query,
            top_k=10,
            metadata_filters=route.metadata_filters or ({
                "domain": parsed_query.domains,
            } if parsed_query.domains else None),
            user_type=user_type,
        )
        
        yield {
            "type": "retrieval_complete",
            "retrieved_count": len(chunks),
        }

        retrieval_state = self.retriever.get_last_result()

        # Handle access-denied from retrieval authorization
        if retrieval_state.access_denied:
            answer = retrieval_state.access_denied_message
            yield {"type": "chunk", "text": answer}
            self.memory.add_turn(session_id=session_id, role="user", message=query)
            self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
            yield {"type": "done", "session_id": session_id}
            return

        if retrieval_state.clarification_required or retrieval_state.confidence < 0.25:
            answer = retrieval_state.clarification_prompt or "Please provide a little more detail so I can retrieve the right records."
            yield {"type": "chunk", "text": answer}
            self.memory.add_turn(session_id=session_id, role="user", message=query)
            self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
            yield {"type": "done", "session_id": session_id}
            return

        # If user asked for reports or vehicle history, stream the aggregated summary
        if parsed_query.intent in {QueryIntent.VEHICLE_HISTORY, QueryIntent.REPORT_GENERATION}:
            try:
                from report_service import ReportService
                report_service = getattr(self, "report_service", None)
                if report_service is None:
                    self.report_service = ReportService(firebase_source=self.firebase_source)
                    report_service = self.report_service

                agg = report_service.summarize_service_history(query=query)
                summary = agg.get("summary", "No summary available.")
                yield {"type": "chunk", "text": summary}
                # Store in memory
                self.memory.add_turn(session_id=session_id, role="user", message=query)
                self.memory.add_turn(session_id=session_id, role="assistant", message=summary)
                yield {"type": "done", "session_id": session_id}
                return
            except Exception as exc:
                yield {"type": "error", "message": str(exc)}
                return

        # Generate response
        answer = self.responder.generate(
            query=query,
            parsed_query=parsed_query,
            user_type=user_type,
            history=history,
            chunks=chunks,
        )
        
        # Stream response character by character
        for char in answer:
            yield {
                "type": "chunk",
                "text": char,
            }
        
        # Store in memory
        self.memory.add_turn(session_id=session_id, role="user", message=query)
        self.memory.add_turn(session_id=session_id, role="assistant", message=answer)
        
        yield {
            "type": "done",
            "session_id": session_id,
        }

    def clear_session(self, session_id: str = "default") -> dict:
        """Clear conversation history for a session."""
        self.memory.clear(session_id)
        return {
            "success": True,
            "message": f"Session {session_id} cleared.",
        }

    def get_session_history(self, session_id: str = "default") -> list[dict]:
        """Get conversation history for a session."""
        return self.memory.get_recent(session_id=session_id, limit=20)


# Maintain backward compatibility with old interface
class Rag(EnhancedRag):
    """Backward-compatible RAG interface."""
    pass

# ============================================================================
# FILE: api.py
# ============================================================================

"""FastAPI backend — Firestore-integrated RAG AI Assistant.

Exposes:
  POST /chat              — main conversational endpoint
  POST /chat/stream       — SSE streaming variant
  GET  /session/{id}/history
  DELETE /session/{id}
  POST /reports/generate
  GET  /reports/{filename}
  GET  /health
  GET  /firebase/status
  GET  /analytics/intent-distribution
"""

from __future__ import annotations

import json
import os
import re
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any, Literal

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Depends, Header, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel, Field

import time
import logging

from conversation_store import ConversationMemory
from firebase_source import FirebaseSource
from grounding_guard import get_grounding_guard
from llm_service import AssistantResponder
from metrics_service import MetricsService, get_metrics_service, collect_system_health
from nlp_query_parser import NLPQueryParser
from conversational_responder import ConversationalResponder
from prompt_guard import get_prompt_guard
from rate_limiter import RateLimiter, get_rate_limiter
from report_service import ReportService
from retrieval import LiveFirebaseRetriever, RetrievedChunk
from retrieval_access_control import get_access_control
from secure_context_builder import get_secure_context_builder
from security_logger import get_security_logger
from rag import EnhancedRag

# ---------------------------------------------------------------------------
# Bootstrap
# ---------------------------------------------------------------------------

load_dotenv()

# Logging
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=LOG_LEVEL)
logger = logging.getLogger("rag_api")

# Simple API token (Bearer) authentication
API_TOKEN = os.getenv("API_TOKEN")

def require_api_token(authorization: str | None = Header(None)) -> bool:
    if not API_TOKEN:
        # No token configured — allow anonymous access (use with caution)
        return True
    if not authorization:
        raise HTTPException(status_code=401, detail="Missing Authorization header")
    if authorization.strip() != f"Bearer {API_TOKEN}":
        raise HTTPException(status_code=403, detail="Invalid API token")
    return True

HTML_FILE = Path(__file__).parent / "index.html"

app = FastAPI(
    title="Automotive Firebase RAG Assistant",
    version="3.0.0",
    description="Firestore-integrated RAG AI assistant with NLP understanding",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.middleware("http")
async def log_requests(request: Request, call_next):
    start = time.time()
    try:
        response = await call_next(request)
    except Exception as exc:
        logger.exception("Request error: %s %s", request.method, request.url.path)
        raise
    duration = (time.time() - start) * 1000.0
    logger.info("%s %s %s %.2fms", request.client.host if request.client else "-", request.method, request.url.path, duration)
    return response

# ---------------------------------------------------------------------------
# Service singletons
# ---------------------------------------------------------------------------

rag = EnhancedRag()
firebase_source = FirebaseSource()
reports = ReportService(firebase_source=firebase_source)

# Expose individual services so tests can monkeypatch them directly.
memory: ConversationMemory = rag.memory
retriever: LiveFirebaseRetriever = rag.retriever
responder: ConversationalResponder = rag.responder

# Legacy AssistantResponder kept for backward-compat and direct test patching.
assistant = AssistantResponder()

# Production-hardening singletons
rate_limiter: RateLimiter = get_rate_limiter()
metrics_svc: MetricsService = get_metrics_service()
prompt_guard = get_prompt_guard()
grounding_guard = get_grounding_guard()
security_log = get_security_logger()


# ---------------------------------------------------------------------------
# Startup prewarm — load embedding model + cache Firebase data at boot
# ---------------------------------------------------------------------------

@app.on_event("startup")
async def startup_prewarm():
    """Prewarm the retrieval pipeline on server start.

    This eliminates the cold-start penalty on the first user request by:
    1. Fetching all Firestore data into cache
    2. Loading the embedding model
    3. Computing record embeddings
    4. Building the FAISS vector index
    """
    logger.info("Prewarming retrieval pipeline...")
    try:
        base_retriever = rag.retriever.base_retriever
        result = base_retriever.prewarm(limit_per_domain=250)
        logger.info(
            "Prewarm complete: records=%s chunks=%s embeddings=%s index=%s",
            result.get("records"),
            result.get("chunks"),
            result.get("embeddings_built"),
            result.get("index_built"),
        )
    except Exception as exc:
        logger.warning("Prewarm failed (will retry on first request): %s", exc)


# ---------------------------------------------------------------------------
# Request / Response models
# ---------------------------------------------------------------------------


class ChatRequest(BaseModel):
    """Incoming chat message."""

    message: str = Field(..., min_length=1, description="User's question")
    session_id: str | None = None
    user_type: Literal["admin", "customer", "staff"] = "admin"


class ChatMetadata(BaseModel):
    """Metadata attached to every chat response."""

    intent: str
    confidence: float
    retrieval_confidence: float | None = None
    retrieval_strategy: str | None = None
    clarification_required: bool = False
    is_comparative: bool
    is_historical: bool
    is_current: bool
    retrieved_count: int
    domains_searched: list[str]
    entities: list[dict]


class TableData(BaseModel):
    """Structured table payload returned alongside the answer."""

    columns: list[str]
    rows: list[dict[str, str]]


class CitationItem(BaseModel):
    """A single source citation (no internal IDs exposed)."""

    domain: str
    source: str
    score: float


class ChatResponse(BaseModel):
    """Full response to a chat request."""

    session_id: str
    answer: str
    success: bool
    metadata: ChatMetadata
    table: TableData | None = None
    citations: list[CitationItem] = []


class SessionHistoryItem(BaseModel):
    role: str
    message: str
    timestamp: str


class SessionHistoryResponse(BaseModel):
    session_id: str
    turns: list[SessionHistoryItem]


class ReportRequest(BaseModel):
    report_type: Literal["inventory", "issuance", "transactions"]
    format: Literal["excel", "xlsx", "pdf"] = "excel"
    period: Literal["current_month", "last_30_days", "all_time"] = "current_month"
    session_id: str | None = None
    user_type: Literal["admin", "customer", "staff"] = "admin"


class ReportArtifact(BaseModel):
    format: Literal["excel", "pdf"]
    filename: str
    download_url: str


# ---------------------------------------------------------------------------
# Table-building helpers
# ---------------------------------------------------------------------------

# Column definitions per domain × user_type.
# Each entry is (internal_field_candidates, display_label).
_ADMIN_INVENTORY_COLUMNS: list[tuple[list[str], str]] = [
    (["item", "name", "item_name", "product", "description"], "Item Name"),
    (["type", "material_type"], "Type"),
    (["category", "group"], "Category"),
    (["uom", "unit", "unit_of_measure"], "Unit of Measurement (UOM)"),
    (["qty", "quantity", "stock", "remaining_stock", "current_stock"], "Remaining Stock"),
    (["reorder_level", "reorder", "reorder_point", "min_stock", "reorder_suggestion"], "Reorder Suggestion"),
    (["cost", "price", "unit_cost"], "Cost"),
    (["supplier", "supplier_name", "supplier_info"], "Supplier Info (if available)"),
]

_CUSTOMER_INVENTORY_COLUMNS: list[tuple[list[str], str]] = [
    (["item", "name", "item_name", "product", "description"], "Item Name"),
    (["category", "group", "type"], "Category"),
    (["uom", "unit", "unit_of_measure"], "Unit of Measurement (UOM)"),
    (["price", "cost", "unit_cost"], "Cost"),
]

_SERVICES_COLUMNS: list[tuple[list[str], str]] = [
    (["service_name", "service", "name", "description"], "Service Name"),
    (["category", "group", "type"], "Category"),
    (["cost", "price", "unit_cost"], "Cost"),
]

# Restricted fields that customers must never see in table output.
_CUSTOMER_RESTRICTED_TABLE_FIELDS: set[str] = {
    "sku", "product_code", "supplier", "supplier_name", "supplier_info",
    "internal_code", "reorder_level", "reorder_point", "reorder",
    "min_stock", "margin", "markup", "wholesale_price",
}

# Fields that should never appear in any table regardless of role.
_ALWAYS_HIDDEN_TABLE_FIELDS: set[str] = {
    "record_id", "doc_id", "_id", "id", "internal_id",
    "summary", "embedding", "vector", "hash",
    "owner_id", "user_id", "mechanic_id", "admin_notes",
}


def _resolve_field(payload: dict[str, Any], candidates: list[str]) -> str:
    """Return the first non-empty value found for any of the candidate keys."""
    payload_lower = {str(k).lower(): v for k, v in payload.items()}
    for candidate in candidates:
        value = payload_lower.get(candidate.lower())
        if value not in (None, "", [], {}, "none", "null", "n/a", "na", "unknown", "-"):
            return str(value).strip()
    return ""


def _is_table_query(query: str) -> bool:
    """Return True when the user explicitly asks for a table or list view."""
    q = (query or "").lower()
    return any(kw in q for kw in ["table", "list", "show", "display", "all", "materials", "items", "products"])


def _detect_table_domain(chunks: list[RetrievedChunk], query: str) -> str:
    """Determine which domain the table should represent."""
    q = (query or "").lower()
    # Explicit service keywords take priority
    if any(kw in q for kw in ["service", "services", "offering", "offerings"]):
        if any(c.domain in ("services", "item_master") for c in chunks):
            return "services"
    # Material/inventory keywords
    if any(kw in q for kw in ["material", "materials", "inventory", "stock", "item", "items", "product", "products"]):
        return "inventory"
    # Default based on what chunks we actually have
    domain_counts: dict[str, int] = {}
    for c in chunks:
        domain_counts[c.domain] = domain_counts.get(c.domain, 0) + 1
    if domain_counts:
        return max(domain_counts, key=domain_counts.get)
    return "inventory"


def _build_response_table(
    chunks: list[RetrievedChunk],
    query: str,
    user_type: str,
) -> dict[str, Any] | None:
    """Build a structured table from retrieved chunks.

    Returns None when the query does not call for a table view or when
    there are no usable rows.
    """
    if not _is_table_query(query):
        return None

    domain = _detect_table_domain(chunks, query)

    # Choose column spec
    if domain == "services":
        column_spec = _SERVICES_COLUMNS
    elif user_type == "admin":
        column_spec = _ADMIN_INVENTORY_COLUMNS
    else:
        column_spec = _CUSTOMER_INVENTORY_COLUMNS

    # Filter chunks to relevant domain
    relevant_chunks: list[RetrievedChunk] = []
    for chunk in chunks:
        if domain == "services" and chunk.domain not in ("services", "item_master"):
            continue
        if domain == "inventory" and chunk.domain not in ("inventory", "item_master", "products", "stock_inventory"):
            continue
        relevant_chunks.append(chunk)

    if not relevant_chunks:
        return None

    # Build rows from chunks
    rows: list[dict[str, str]] = []
    for chunk in relevant_chunks:
        payload = chunk.payload or {}

        # Security: strip always-hidden fields
        payload = {
            k: v for k, v in payload.items()
            if str(k).lower() not in _ALWAYS_HIDDEN_TABLE_FIELDS
        }

        # Security: strip customer-restricted fields for non-admins
        if user_type != "admin":
            payload = {
                k: v for k, v in payload.items()
                if str(k).lower() not in _CUSTOMER_RESTRICTED_TABLE_FIELDS
            }

        row: dict[str, str] = {}
        for candidates, label in column_spec:
            value = _resolve_field(payload, candidates)
            row[label] = value

        # Skip rows that have no meaningful content at all
        if not any(row.values()):
            continue

        rows.append(row)

    if not rows:
        return None

    # Adaptive columns: remove columns that are empty across ALL rows
    columns = [label for (_, label) in column_spec]
    non_empty_columns = [
        col for col in columns
        if any(row.get(col, "") for row in rows)
    ]

    # If too few columns survived, fall back to showing all payload fields
    if len(non_empty_columns) < 2:
        return _build_adaptive_table(relevant_chunks, user_type)

    # Filter rows to only include non-empty columns
    filtered_rows = [
        {col: row.get(col, "") for col in non_empty_columns}
        for row in rows
    ]

    return {"columns": non_empty_columns, "rows": filtered_rows}


def _build_adaptive_table(
    chunks: list[RetrievedChunk],
    user_type: str,
) -> dict[str, Any] | None:
    """Build a table adaptively from whatever fields exist in the data.

    Used when the predefined column spec doesn't match the actual data
    (common with NoSQL where schemas vary).
    """
    # Collect all field names across chunks and count occurrences
    field_counts: dict[str, int] = {}
    for chunk in chunks:
        payload = chunk.payload or {}
        for key in payload.keys():
            key_lower = str(key).lower()
            if key_lower in _ALWAYS_HIDDEN_TABLE_FIELDS:
                continue
            if user_type != "admin" and key_lower in _CUSTOMER_RESTRICTED_TABLE_FIELDS:
                continue
            field_counts[key] = field_counts.get(key, 0) + 1

    if not field_counts:
        return None

    # Use fields that appear in at least 30% of chunks, up to 6 columns
    min_occurrences = max(1, len(chunks) * 0.3)
    common_fields = sorted(
        [f for f, count in field_counts.items() if count >= min_occurrences],
        key=lambda f: field_counts[f],
        reverse=True,
    )[:6]

    if not common_fields:
        common_fields = sorted(field_counts.keys(), key=lambda f: field_counts[f], reverse=True)[:5]

    # Build column labels
    columns = [f.replace("_", " ").title() for f in common_fields]

    # Build rows
    rows: list[dict[str, str]] = []
    for chunk in chunks:
        payload = chunk.payload or {}
        row: dict[str, str] = {}
        for field, label in zip(common_fields, columns):
            value = payload.get(field)
            if value and str(value).strip().lower() not in ("none", "null", "n/a", "unknown", "-"):
                row[label] = str(value).strip()
            else:
                row[label] = ""
        if any(row.values()):
            rows.append(row)

    if not rows:
        return None

    return {"columns": columns, "rows": rows}
        if not any(row.values()):
            continue

        rows.append(row)

    if not rows:
        return None

    return {"columns": columns, "rows": rows}


# ---------------------------------------------------------------------------
# Citation builder
# ---------------------------------------------------------------------------


def _build_citations(chunks: list[RetrievedChunk]) -> list[CitationItem]:
    """Build citation list from retrieved chunks — no internal IDs exposed."""
    seen: set[tuple[str, str]] = set()
    citations: list[CitationItem] = []
    for chunk in chunks:
        key = (chunk.domain, chunk.source)
        if key in seen:
            continue
        seen.add(key)
        citations.append(
            CitationItem(
                domain=chunk.domain,
                source=chunk.source,
                score=round(chunk.score, 3),
            )
        )
    return citations


# ---------------------------------------------------------------------------
# Access-control helper (also used by tests)
# ---------------------------------------------------------------------------

_CUSTOMER_RESTRICTED_QUERIES: list[str] = [
    "stock quantity",
    "inventory stock",
    "inventory database",
    "supplier",
    "financial",
    "analytics",
    "mechanic private",
    "other customer",
    "all customers",
    "full inventory",
    "reorder level",
]


def _customer_access_denied(query: str, user_type: str) -> str | None:
    """Return an access-denied message if the customer query is restricted.

    Returns None when access is allowed.

    This function provides a fast keyword-based pre-check at the API layer.
    The deeper retrieval-level access control in retrieval_access_control.py
    provides the authoritative enforcement at the data layer.
    """
    if user_type == "admin":
        return None
    q = (query or "").lower()

    # Staff can access operational data but not admin-only analytics
    if user_type == "staff":
        staff_restricted = ["mechanic private", "other customer", "all customers", "financial"]
        if any(term in q for term in staff_restricted):
            return "Access denied. This data requires administrator privileges."
        return None

    # Customer restrictions
    if any(term in q for term in _CUSTOMER_RESTRICTED_QUERIES):
        return "Access denied. You can only access your own records."
    return None


# ---------------------------------------------------------------------------
# Root endpoints
# ---------------------------------------------------------------------------


@app.get("/")
def root():
    """Serve frontend or API info."""
    if HTML_FILE.exists():
        return FileResponse(HTML_FILE)
    return {
        "status": "running",
        "service": "Automotive Firebase RAG Assistant v3",
        "endpoints": {
            "chat": "/chat",
            "stream": "/chat/stream",
            "history": "/session/{session_id}/history",
            "health": "/health",
        },
    }


@app.get("/health")
def health():
    """Health check endpoint."""
    status = firebase_source.get_status()
    return {
        "status": "ok",
        "service": "automotive-firebase-rag",
        "version": "3.0.0",
        "firebase": status,
        "timestamp": datetime.utcnow().isoformat(),
    }


@app.get("/firebase/status")
def firebase_status():
    """Get Firebase connection status."""
    return firebase_source.get_status()


# ---------------------------------------------------------------------------
# Chat endpoints
# ---------------------------------------------------------------------------


@app.post("/chat", response_model=ChatResponse)
def chat(request: ChatRequest, req: Request = None, authorized: bool = Depends(require_api_token)):
    """Main conversational endpoint.

    Features:
    - Automatic intent detection
    - Entity extraction
    - Role-based access control
    - Rate limiting
    - Prompt injection defense
    - Grounding enforcement
    - Optional structured table output
    - Source citations (no internal IDs)
    """
    session_id = request.session_id or str(uuid.uuid4())
    client_ip = req.client.host if req and req.client else ""

    # Rate limiting
    token_id = session_id  # Use session as rate limit key
    rate_result = rate_limiter.check(token_id, request.user_type, client_ip)
    if not rate_result.allowed:
        security_log.log_rate_limit(
            user_role=request.user_type,
            session_id=session_id,
            client_ip=client_ip,
            retry_after=rate_result.retry_after_seconds,
        )
        raise HTTPException(
            status_code=429,
            detail=rate_result.reason,
            headers={"Retry-After": str(int(rate_result.retry_after_seconds))},
        )

    # Role-based access guard
    denied = _customer_access_denied(request.message, request.user_type)
    if denied:
        return ChatResponse(
            session_id=session_id,
            answer=denied,
            success=False,
            metadata=ChatMetadata(
                intent="access_denied",
                confidence=1.0,
                is_comparative=False,
                is_historical=False,
                is_current=False,
                retrieved_count=0,
                domains_searched=[],
                entities=[],
            ),
            table=None,
            citations=[],
        )

    # Run RAG pipeline
    result = rag.ask(
        query=request.message,
        session_id=session_id,
        user_type=request.user_type,
    )

    if not result.get("success"):
        raise HTTPException(
            status_code=400,
            detail=result.get("error", "Failed to process query"),
        )

    metadata = result.get("metadata", {})

    # Retrieve chunks for table + citation building (re-use cached retriever)
    chunks = retriever.retrieve(
        query=request.message,
        top_k=20,
    )

    # Apply secure context filtering before building tables/citations
    secure_builder = get_secure_context_builder()
    safe_chunks = secure_builder.filter_chunks(chunks, request.user_type)

    table = _build_response_table(safe_chunks, request.message, request.user_type)
    citations = _build_citations(secure_builder.filter_citations(chunks, request.user_type))

    return ChatResponse(
        session_id=result["session_id"],
        answer=result["answer"],
        success=True,
        metadata=ChatMetadata(
            intent=metadata.get("intent", "unknown"),
            confidence=metadata.get("confidence", 0.0),
            is_comparative=metadata.get("is_comparative", False),
            is_historical=metadata.get("is_historical", False),
            is_current=metadata.get("is_current", False),
            retrieved_count=metadata.get("retrieved_count", 0),
            domains_searched=metadata.get("domains_searched", []),
            entities=metadata.get("entities", []),
        ),
        table=table,
        citations=citations,
    )


@app.post("/chat/stream")
def chat_stream(request: ChatRequest, authorized: bool = Depends(require_api_token)):
    """Streaming chat endpoint — returns Server-Sent Events.

    Event types:
      parsing_complete  — intent detection done
      retrieval_complete — data retrieval done
      chunk             — response text chunk
      done              — response complete
      error             — error occurred
    """
    session_id = request.session_id or str(uuid.uuid4())

    def event_generator():
        try:
            for event in rag.ask_stream(
                query=request.message,
                session_id=session_id,
                user_type=request.user_type,
            ):
                yield f"data: {json.dumps(event)}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'message': str(exc)})}\n\n"

    return StreamingResponse(event_generator(), media_type="text/event-stream")


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------


@app.get("/session/{session_id}/history", response_model=SessionHistoryResponse)
def get_session_history(session_id: str):
    """Return conversation history for a session."""
    turns = rag.get_session_history(session_id)
    return SessionHistoryResponse(
        session_id=session_id,
        turns=[
            SessionHistoryItem(
                role=turn["role"],
                message=turn["message"],
                timestamp=turn.get("timestamp", ""),
            )
            for turn in turns
        ],
    )


@app.delete("/session/{session_id}")
def clear_session(session_id: str):
    """Clear conversation history for a session."""
    return rag.clear_session(session_id)


# ---------------------------------------------------------------------------
# Report endpoints
# ---------------------------------------------------------------------------


def _timestamp_slug() -> str:
    return datetime.utcnow().strftime("%Y%m%d_%H%M%S")


def _save_report_bytes(
    report_type: str, file_format: str, content: bytes
) -> tuple[str, str]:
    reports_dir = Path(__file__).parent / "generated_reports"
    reports_dir.mkdir(exist_ok=True)
    filename = f"{report_type}_{_timestamp_slug()}.{file_format}"
    path = reports_dir / filename
    path.write_bytes(content)
    return filename, f"/reports/{filename}"


@app.post("/reports/generate")
def generate_report(request: ReportRequest):
    """Generate an Excel or PDF report from live Firestore data."""
    session_id = request.session_id or str(uuid.uuid4())

    try:
        report = reports.build_report_data(
            request.report_type,
            period_label=request.period.replace("_", " ").title(),
        )

        artifacts: list[ReportArtifact] = []
        formats = ["pdf"] if request.format == "pdf" else ["xlsx"]

        for fmt in formats:
            if fmt == "pdf":
                content = reports.generate_pdf(report)
            else:
                content = reports.generate_excel(report)
            filename, url = _save_report_bytes(report.report_type, fmt, content)
            artifacts.append(
                ReportArtifact(
                    format="pdf" if fmt == "pdf" else "excel",
                    filename=filename,
                    download_url=url,
                )
            )

        answer = (
            f"Generated {report.title.lower()} successfully "
            f"in {', '.join(a.format for a in artifacts)} format."
        )

        memory.add_turn(
            session_id=session_id,
            role="user",
            message=f"Generate {request.report_type} report",
        )
        memory.add_turn(session_id=session_id, role="assistant", message=answer)

        return {
            "session_id": session_id,
            "answer": answer,
            "report_type": report.report_type,
            "artifacts": [a.dict() for a in artifacts],
        }

    except Exception as exc:
        raise HTTPException(status_code=500, detail=str(exc))


@app.get("/reports/{filename}")
def download_report(filename: str):
    """Download a previously generated report file."""
    reports_dir = Path(__file__).parent / "generated_reports"
    path = (reports_dir / filename).resolve()

    # Path-traversal guard
    if not str(path).startswith(str(reports_dir.resolve())):
        raise HTTPException(status_code=400, detail="Invalid report path")

    if not path.exists():
        raise HTTPException(status_code=404, detail="Report not found")

    media_type = (
        "application/pdf"
        if path.suffix.lower() == ".pdf"
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return FileResponse(path, media_type=media_type, filename=path.name)


# ---------------------------------------------------------------------------
# Analytics / debug endpoints
# ---------------------------------------------------------------------------


@app.get("/analytics/intent-distribution")
def analytics_intent_distribution():
    """Intent distribution analytics (requires intent tracking to be enabled)."""
    return {
        "message": "Analytics endpoint available when intent tracking is enabled.",
    }


# ---------------------------------------------------------------------------
# Metrics endpoint (admin only)
# ---------------------------------------------------------------------------


@app.get("/metrics")
def metrics_endpoint(authorized: bool = Depends(require_api_token)):
    """Production observability metrics endpoint.

    Returns structured JSON with:
    - Total queries and breakdown by role/intent
    - Retrieval latency (avg, p95)
    - Average retrieval confidence
    - Vector index size and embedding readiness
    - Firestore cache status
    - Active sessions
    - Clarification rate
    - Access-denied count
    - Injection blocked count
    - Rate limit stats
    - Grounding enforcement stats
    - Security event summary

    Safe for admin use. No sensitive user data exposed.
    """
    return collect_system_health(
        firebase_source=firebase_source,
        retriever=rag.retriever,
        memory=memory,
        rate_limiter=rate_limiter,
        prompt_guard=prompt_guard,
        grounding_guard=grounding_guard,
        security_logger=security_log,
        metrics_service=metrics_svc,
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000)


# ============================================================================
# FILE: main.py
# ============================================================================

"""Local launcher for the FastAPI Firebase assistant backend."""

import uvicorn


if __name__ == "__main__":
    uvicorn.run("api:app", host="0.0.0.0", port=8000, reload=True)

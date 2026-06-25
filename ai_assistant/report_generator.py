"""
Report generator for Caltex AutoPro AI Assistant.

Produces PDF and Excel reports that match the exact branded layout used by
the website's print_layout_header.js — same header, same table style, same
colour scheme (#E8001C red, #1a202c dark, #003087 blue).

Supported report types:
  inventory      → stock_inventory collection
  issuance       → issuances collection
  maintenance    → maintenance collection
  vehicles       → vehicles collection
  bookings       → service_bookings collection
"""

import io
import os
from datetime import datetime

from firestore_client import get_db

# ── Brand constants (match print_layout_header.js) ────────────────────────────
BRAND_NAME    = "JA Noble Enterprise Inc."
BRAND_SUB     = "Caltex San Pedro"
BRAND_ADDRESS = "102 National Highway, Brgy. Landayan, San Pedro, Laguna"
RED           = (232, 0, 28)        # #E8001C
DARK          = (26, 32, 44)        # #1a202c
BLUE          = (0, 48, 135)        # #003087
GREY          = (113, 128, 150)     # #718096
LIGHT_GREY    = (247, 248, 250)     # #f7f8fa
WHITE         = (255, 255, 255)
ROW_ALT       = (248, 250, 252)     # #f8fafc  (even row bg)

LOGO_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "automotive_website", "static", "img", "LOGO_CALTEX.png"
)


# ─────────────────────────────────────────────────────────────────────────────
#  Firestore data fetchers
# ─────────────────────────────────────────────────────────────────────────────

def _fetch_inventory():
    db = get_db()
    rows = []
    for doc in db.collection("stock_inventory").stream():
        d = doc.to_dict()
        stock   = d.get("stock", 0)
        min_lvl = d.get("min", 0)
        max_lvl = d.get("max", 0)
        reorder = d.get("reorder", 0)
        status  = d.get("status", "")
        if not status:
            if int(stock) <= int(reorder):
                status = "Low"
            elif int(stock) > int(max_lvl) and int(max_lvl) > 0:
                status = "Over"
            else:
                status = "OK"
        rows.append({
            "num":    d.get("num", ""),
            "name":   d.get("name", ""),
            "group":  d.get("group", ""),
            "stock":  stock,
            "uom":    d.get("uom", ""),
            "min":    min_lvl,
            "max":    max_lvl,
            "reorder":reorder,
            "status": status,
        })
    rows.sort(key=lambda r: r["name"])
    return rows


def _fetch_issuances():
    db = get_db()
    rows = []
    for doc in db.collection("issuances").stream():
        d = doc.to_dict()
        rows.append({
            "date":       d.get("date", ""),
            "plate":      d.get("plate", ""),
            "itemNum":    d.get("itemNum", ""),
            "itemName":   d.get("itemName", ""),
            "itemType":   d.get("itemType", ""),
            "group":      d.get("commodityGroup", ""),
            "uom":        d.get("uom", ""),
            "qty":        d.get("qty", ""),
            "unitCost":   d.get("unitCost", ""),
            "subtotal":   d.get("subtotal", ""),
        })
    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


def _fetch_maintenance():
    db = get_db()
    rows = []
    for doc in db.collection("maintenance").stream():
        d = doc.to_dict()
        rows.append({
            "id":      d.get("id", doc.id),
            "plate":   d.get("plate", ""),
            "desc":    d.get("desc", ""),
            "mechanic":d.get("mechanic", ""),
            "date":    d.get("date", ""),
            "status":  d.get("status", ""),
            "cost":    d.get("cost", ""),
        })
    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


def _fetch_vehicles():
    db = get_db()
    rows = []
    for doc in db.collection("vehicles").stream():
        d = doc.to_dict()
        rows.append({
            "plate":       d.get("plate", ""),
            "desc":        d.get("desc", ""),
            "owner":       d.get("owner", ""),
            "type":        d.get("type", ""),
            "status":      d.get("status", ""),
            "lastSvcDate": d.get("lastSvcDate", ""),
            "odo":         d.get("odo", ""),
        })
    rows.sort(key=lambda r: r["plate"])
    return rows


def _fetch_bookings():
    db = get_db()
    rows = []
    for doc in db.collection("service_bookings").stream():
        d = doc.to_dict()
        svcs = d.get("services", [])
        rows.append({
            "date":          d.get("preferredDate", ""),
            "time":          d.get("preferredTime", ""),
            "customerName":  d.get("customerName", ""),
            "plate":         d.get("plate", ""),
            "vehicleDesc":   d.get("vehicleDesc", ""),
            "services":      ", ".join(svcs) if isinstance(svcs, list) else str(svcs),
            "status":        d.get("status", ""),
        })
    rows.sort(key=lambda r: r["date"], reverse=True)
    return rows


# ─────────────────────────────────────────────────────────────────────────────
#  Report config: columns + row mapper per report type
# ─────────────────────────────────────────────────────────────────────────────

REPORT_CONFIG = {
    "inventory": {
        "title":   "Inventory Report",
        "fetcher": _fetch_inventory,
        "columns": ["Item No.", "Item Name", "Group", "Stock", "UOM", "Min", "Max", "Reorder Pt.", "Status"],
        "mapper":  lambda r: [
            r["num"], r["name"], r["group"],
            str(r["stock"]), r["uom"],
            str(r["min"]), str(r["max"]), str(r["reorder"]),
            r["status"],
        ],
        "col_widths_pdf":  [50, 110, 70, 35, 35, 30, 30, 45, 40],   # pts, total ~445
        "col_widths_xlsx": [12, 30, 20, 10, 10, 8, 8, 12, 12],
    },
    "issuance": {
        "title":   "Issuance Report",
        "fetcher": _fetch_issuances,
        "columns": ["Date", "Plate", "Item No.", "Item Name", "Type", "Group", "UOM", "Qty", "Unit Cost", "Subtotal"],
        "mapper":  lambda r: [
            r["date"], r["plate"], r["itemNum"], r["itemName"],
            r["itemType"], r["group"], r["uom"],
            str(r["qty"]), str(r["unitCost"]), str(r["subtotal"]),
        ],
        "col_widths_pdf":  [45, 45, 40, 85, 40, 55, 28, 22, 45, 45],
        "col_widths_xlsx": [14, 14, 12, 28, 14, 18, 8, 8, 14, 14],
    },
    "maintenance": {
        "title":   "Maintenance Report",
        "fetcher": _fetch_maintenance,
        "columns": ["Job ID", "Plate", "Description", "Mechanic", "Date", "Status", "Cost"],
        "mapper":  lambda r: [
            r["id"], r["plate"], r["desc"], r["mechanic"],
            r["date"], r["status"], str(r["cost"]),
        ],
        "col_widths_pdf":  [55, 45, 110, 80, 50, 50, 55],
        "col_widths_xlsx": [16, 14, 32, 24, 14, 14, 16],
    },
    "vehicles": {
        "title":   "Vehicle Fleet Report",
        "fetcher": _fetch_vehicles,
        "columns": ["Plate No.", "Description", "Owner", "Type", "Status", "Last Service", "Odometer"],
        "mapper":  lambda r: [
            r["plate"], r["desc"], r["owner"], r["type"],
            r["status"], r["lastSvcDate"], str(r["odo"]),
        ],
        "col_widths_pdf":  [55, 110, 80, 50, 50, 55, 45],
        "col_widths_xlsx": [14, 30, 24, 14, 14, 16, 14],
    },
    "bookings": {
        "title":   "Service Bookings Report",
        "fetcher": _fetch_bookings,
        "columns": ["Date", "Time", "Customer", "Plate", "Vehicle", "Services", "Status"],
        "mapper":  lambda r: [
            r["date"], r["time"], r["customerName"],
            r["plate"], r["vehicleDesc"], r["services"], r["status"],
        ],
        "col_widths_pdf":  [45, 35, 80, 45, 80, 100, 50],
        "col_widths_xlsx": [14, 10, 24, 14, 24, 30, 14],
    },
}


# ─────────────────────────────────────────────────────────────────────────────
#  PDF generation  (reportlab)
# ─────────────────────────────────────────────────────────────────────────────

def _rgb(t): return tuple(v / 255 for v in t)


def generate_pdf(report_type: str) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, Image,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

    pt = 1.0   # 1 point == 1 reportlab unit

    cfg     = REPORT_CONFIG[report_type]
    data    = cfg["fetcher"]()
    columns = cfg["columns"]
    mapper  = cfg["mapper"]
    widths  = [w * pt for w in cfg["col_widths_pdf"]]

    buf = io.BytesIO()
    # Use landscape A4 for wide tables
    page_size = landscape(A4) if sum(cfg["col_widths_pdf"]) > 450 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=15*mm, bottomMargin=15*mm,
    )

    # ── Styles ────────────────────────────────────────────────
    red_c   = colors.Color(*_rgb(RED))
    dark_c  = colors.Color(*_rgb(DARK))
    blue_c  = colors.Color(*_rgb(BLUE))
    grey_c  = colors.Color(*_rgb(GREY))
    alt_c   = colors.Color(*_rgb(ROW_ALT))
    white_c = colors.white

    brand_style = ParagraphStyle("brand", fontName="Helvetica-Bold", fontSize=14,
                                  textColor=dark_c, leading=16)
    sub_style   = ParagraphStyle("sub",   fontName="Helvetica",      fontSize=7,
                                  textColor=grey_c, leading=9)
    addr_style  = ParagraphStyle("addr",  fontName="Helvetica",      fontSize=6.5,
                                  textColor=colors.Color(0.63, 0.69, 0.77), leading=8)
    label_style = ParagraphStyle("lbl",   fontName="Helvetica-Bold", fontSize=6.5,
                                  textColor=grey_c, alignment=TA_RIGHT, leading=8)
    title_style = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=15,
                                  textColor=dark_c, alignment=TA_RIGHT, leading=17)
    meta_style  = ParagraphStyle("meta",  fontName="Helvetica",      fontSize=7.5,
                                  textColor=grey_c, alignment=TA_RIGHT, leading=9)
    cell_style  = ParagraphStyle("cell",  fontName="Helvetica",      fontSize=8,
                                  textColor=colors.Color(0.18, 0.2, 0.27), leading=10)

    # ── Header ────────────────────────────────────────────────
    now      = datetime.now()
    date_str = now.strftime("%B %d, %Y")
    time_str = now.strftime("%I:%M %p")

    left_block = [
        Paragraph(BRAND_NAME, brand_style),
        Spacer(1, 2),
        Paragraph(BRAND_SUB.upper(), sub_style),
        Spacer(1, 1),
        Paragraph(BRAND_ADDRESS, addr_style),
    ]

    right_block = [
        Paragraph("OFFICIAL REPORT", label_style),
        Paragraph(cfg["title"], title_style),
        Spacer(1, 2),
        Paragraph(f'Generated on <b>{date_str}</b> at {time_str}', meta_style),
        Paragraph(f'{len(data)} record(s)', meta_style),
    ]

    # Try to load logo
    logo_cell = ""
    if os.path.exists(LOGO_PATH):
        logo_cell = Image(LOGO_PATH, width=36, height=36)

    header_data = [[logo_cell, left_block, right_block]]
    usable_w = doc.width
    header_table = Table(header_data, colWidths=[42, usable_w * 0.45, usable_w * 0.45])
    header_table.setStyle(TableStyle([
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",(0, 0), (-1, -1), 0),
        ("BOTTOMPADDING",(0,0), (-1, -1), 0),
        ("TOPPADDING",  (0, 0), (-1, -1), 0),
    ]))

    # ── Data table ────────────────────────────────────────────
    header_row = [Paragraph(f'<b>{c}</b>', ParagraphStyle(
        "th", fontName="Helvetica-Bold", fontSize=7,
        textColor=white_c, leading=9)) for c in columns]

    table_data = [header_row]
    for row in data:
        mapped = mapper(row)
        table_data.append([
            Paragraph(str(v) if v else "—", cell_style) for v in mapped
        ])

    # Alternate row colours
    ts = [
        ("BACKGROUND",  (0, 0), (-1, 0),  dark_c),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [white_c, alt_c]),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.Color(0.88, 0.91, 0.94)),
        ("LINEBELOW",   (0, 0), (-1, 0),  1,   dark_c),
        ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",  (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0,0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",(0, 0), (-1, -1), 5),
    ]

    # Status column colour coding (last or near-last column)
    status_cols = {
        "inventory":   8,
        "maintenance": 5,
        "vehicles":    4,
        "bookings":    6,
    }
    sc = status_cols.get(report_type)
    if sc is not None:
        for i, row in enumerate(data, start=1):
            s = (mapper(row)[sc] or "").lower()
            if s in ("low", "overdue", "cancelled", "rejected"):
                ts.append(("TEXTCOLOR", (sc, i), (sc, i), red_c))
                ts.append(("FONTNAME",  (sc, i), (sc, i), "Helvetica-Bold"))
            elif s in ("ok", "active", "completed", "approved"):
                ts.append(("TEXTCOLOR", (sc, i), (sc, i), colors.Color(0.13, 0.55, 0.13)))
                ts.append(("FONTNAME",  (sc, i), (sc, i), "Helvetica-Bold"))
            elif s in ("pending", "in progress", "ongoing"):
                ts.append(("TEXTCOLOR", (sc, i), (sc, i), colors.Color(0.55, 0.4, 0.02)))
                ts.append(("FONTNAME",  (sc, i), (sc, i), "Helvetica-Bold"))

    data_table = Table(table_data, colWidths=widths, repeatRows=1)
    data_table.setStyle(TableStyle(ts))

    # ── Assemble ──────────────────────────────────────────────
    story = [
        header_table,
        HRFlowable(width="100%", thickness=2, color=red_c, spaceAfter=10),
        data_table,
        Spacer(1, 16),
        Paragraph(
            f'Caltex AutoPro · {BRAND_NAME} · Confidential',
            ParagraphStyle("foot", fontName="Helvetica", fontSize=7,
                           textColor=grey_c, alignment=TA_CENTER)
        ),
    ]

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  Excel generation  (openpyxl)
# ─────────────────────────────────────────────────────────────────────────────

def generate_excel(report_type: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill,
    )
    from openpyxl.utils import get_column_letter

    cfg     = REPORT_CONFIG[report_type]
    data    = cfg["fetcher"]()
    columns = cfg["columns"]
    mapper  = cfg["mapper"]
    col_ws  = cfg["col_widths_xlsx"]

    wb = Workbook()
    ws = wb.active
    ws.title = cfg["title"][:31]  # sheet name limit

    now      = datetime.now()
    date_str = now.strftime("%B %d, %Y")

    # ── Colour helpers ────────────────────────────────────────
    def fill(hex_str):  return PatternFill("solid", fgColor=hex_str)
    def side():         return Side(style="thin", color="E2E8F0")
    thin_border = Border(left=side(), right=side(), top=side(), bottom=side())

    # ── Title block (rows 1-4) ────────────────────────────────
    ncols = len(columns)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=ncols)

    r1 = ws.cell(1, 1, BRAND_NAME)
    r1.font      = Font(name="Calibri", bold=True, size=14, color="1A202C")
    r1.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    r2 = ws.cell(2, 1, BRAND_SUB + "  ·  " + BRAND_ADDRESS)
    r2.font      = Font(name="Calibri", size=9, color="718096")
    r2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 14

    r3 = ws.cell(3, 1, cfg["title"].upper())
    r3.font      = Font(name="Calibri", bold=True, size=16, color="E8001C")
    r3.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 26

    r4 = ws.cell(4, 1, f"Generated: {date_str}  ·  {len(data)} record(s)")
    r4.font      = Font(name="Calibri", size=9, color="718096")
    r4.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[4].height = 14

    # Red divider row
    ws.row_dimensions[5].height = 3
    for c in range(1, ncols + 1):
        ws.cell(5, c).fill = fill("E8001C")

    ws.row_dimensions[6].height = 6  # spacer

    # ── Header row (row 7) ────────────────────────────────────
    HDR_ROW = 7
    ws.row_dimensions[HDR_ROW].height = 20
    for ci, col_name in enumerate(columns, 1):
        cell = ws.cell(HDR_ROW, ci, col_name.upper())
        cell.font      = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        cell.fill      = fill("1A202C")
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=False)
        cell.border    = thin_border

    # ── Data rows ────────────────────────────────────────────
    STATUS_COLORS = {
        "low":         ("E8001C", "FFF5F5"),
        "overdue":     ("E8001C", "FFF5F5"),
        "cancelled":   ("E8001C", "FFF5F5"),
        "rejected":    ("E8001C", "FFF5F5"),
        "ok":          ("276749", "F0FFF4"),
        "active":      ("276749", "F0FFF4"),
        "completed":   ("003087", "EBF8FF"),
        "approved":    ("276749", "F0FFF4"),
        "pending":     ("975A16", "FFFBEB"),
        "in progress": ("975A16", "FFFBEB"),
        "ongoing":     ("975A16", "FFFBEB"),
    }

    status_col_idx = {
        "inventory":   8,
        "maintenance": 5,
        "vehicles":    4,
        "bookings":    6,
    }.get(report_type)

    for ri, row in enumerate(data):
        excel_row = HDR_ROW + 1 + ri
        ws.row_dimensions[excel_row].height = 15
        mapped = mapper(row)
        is_even = ri % 2 == 1
        row_bg = "F8FAFC" if is_even else "FFFFFF"

        for ci, val in enumerate(mapped, 1):
            cell = ws.cell(excel_row, ci, val if val else "—")
            cell.font      = Font(name="Calibri", size=9, color="2D3748")
            cell.fill      = fill(row_bg)
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border    = thin_border

            # Status colour overrides
            if status_col_idx and ci == status_col_idx + 1:
                key = (val or "").lower()
                if key in STATUS_COLORS:
                    fc, bc = STATUS_COLORS[key]
                    cell.font = Font(name="Calibri", size=9, bold=True, color=fc)
                    cell.fill = fill(bc)

    # ── Column widths ─────────────────────────────────────────
    for ci, w in enumerate(col_ws, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Freeze header row
    ws.freeze_panes = ws.cell(HDR_ROW + 1, 1)

    # ── Footer ────────────────────────────────────────────────
    foot_row = HDR_ROW + len(data) + 2
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=ncols)
    fc = ws.cell(foot_row, 1,
                 f"Caltex AutoPro · {BRAND_NAME} · Confidential · Generated {date_str}")
    fc.font      = Font(name="Calibri", size=8, color="A0AEC0", italic=True)
    fc.alignment = Alignment(horizontal="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
#  Public API
# ─────────────────────────────────────────────────────────────────────────────

def build_report(report_type: str, file_format: str) -> bytes:
    """
    Generate a report file.

    Args:
        report_type: one of inventory | issuance | maintenance | vehicles | bookings
        file_format: pdf | excel | xlsx

    Returns:
        Raw bytes of the generated file.
    """
    if report_type not in REPORT_CONFIG:
        raise ValueError(f"Unknown report type '{report_type}'. "
                         f"Use: {', '.join(REPORT_CONFIG.keys())}")
    fmt = file_format.lower().replace("xlsx", "excel")
    if fmt == "pdf":
        return generate_pdf(report_type)
    elif fmt in ("excel", "xlsx"):
        return generate_excel(report_type)
    else:
        raise ValueError(f"Unknown format '{file_format}'. Use: pdf or excel")
